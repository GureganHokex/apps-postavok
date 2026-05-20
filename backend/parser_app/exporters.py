"""
Экспорт заказов в PDF и Excel.
"""

import logging
import os
from pathlib import Path
from datetime import datetime
from django.conf import settings
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from typing import Optional
from .models import Order, ParsedItem, TapLocation, Tap
import pandas as pd
import copy

logger = logging.getLogger(__name__)


class OrderExporter:
    """
    Класс для экспорта заказов в различные форматы.
    
    Поддерживает экспорт в Excel и PDF форматы.
    """
    
    def __init__(self, order: Order):
        """
        Инициализация экспортера.
        
        Args:
            order: Объект заказа для экспорта
        """
        self.order = order

    @staticmethod
    def _order_item_id(order_item: dict):
        """ID позиции в заказе (поддержка item_id и id)."""
        if not order_item:
            return None
        raw = order_item.get('item_id', order_item.get('id'))
        if raw is None:
            return None
        try:
            return int(raw)
        except (TypeError, ValueError):
            return None

    def _first_order_item_id(self) -> Optional[int]:
        for order_item in self.order.items or []:
            item_id = self._order_item_id(order_item)
            if item_id is not None:
                return item_id
        return None

    @staticmethod
    def _safe_float(value):
        if value is None or value == '':
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _safe_row_num(location) -> Optional[int]:
        """Номер строки в источнике; без int() при None — иначе 500 при экспорте."""
        if not location:
            return None
        r = location.get('row')
        if r is None:
            return None
        try:
            return int(r)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _sheet_name_in_workbook(location, wb) -> Optional[str]:
        """
        Имя листа, как в файле xlsx. У brewery-парсера в JSON бывает sheet='Банки'|'Кеги',
        а реальный лист — original_sheet ('Актуальный Прайс' и т.д.); иначе openpyxl не находит строку.
        """
        if not location or not wb.sheetnames:
            return None
        names = list(wb.sheetnames)
        original = location.get('original_sheet')
        explicit = location.get('sheet')
        for candidate in (original, explicit):
            if candidate and candidate in names:
                return candidate

        def norm(s):
            return (s or '').strip().casefold()

        for candidate in (original, explicit):
            if not candidate:
                continue
            cn = norm(candidate)
            for n in names:
                if norm(n) == cn:
                    return n
        return names[0]

    @staticmethod
    def _merge_anchor(ws, row: int, col: int) -> tuple[int, int]:
        """
        openpyxl: only the top-left cell of a merged range is writable; other cells are MergedCell (read-only).
        """
        for cell_range in ws.merged_cells.ranges:
            if (
                cell_range.min_row <= row <= cell_range.max_row
                and cell_range.min_col <= col <= cell_range.max_col
            ):
                return cell_range.min_row, cell_range.min_col
        return row, col

    def export(self) -> str:
        """
        Экспортирует заказ в выбранный формат.
        
        Returns:
            Путь к экспортированному файлу
        """
        if self.order.export_format == 'excel':
            return self.export_to_excel()
        elif self.order.export_format == 'pdf':
            return self.export_to_pdf()
        else:
            raise ValueError(f"Неизвестный формат экспорта: {self.order.export_format}")
    
    def export_to_excel(self) -> str:
        """
        Экспортирует заказ в Excel файл, используя исходный файл как шаблон.
        
        Returns:
            Путь к созданному Excel файлу
        """
        order_items = self._get_order_items()
        if not order_items:
            raise ValueError(
                "Заказ не содержит позиций: записи в прайсе удалены или недоступны."
            )

        first_item_id = self._first_order_item_id()
        if not first_item_id:
            raise ValueError("В заказе нет корректных позиций (нужен item_id).")

        try:
            parsed_item = ParsedItem.objects.select_related('file').get(id=first_item_id)
        except ParsedItem.DoesNotExist:
            return self._export_pdf_to_excel(None, order_items)

        source_file = parsed_item.file
        source_file_path = Path(settings.MEDIA_ROOT) / source_file.file_path

        if source_file.file_type == 'pdf' or not source_file_path.exists():
            if not source_file_path.exists():
                logger.warning(
                    "Исходный прайс не найден на диске (%s), экспорт заказа #%s в отдельный Excel",
                    source_file_path,
                    self.order.id,
                )
            return self._export_pdf_to_excel(source_file, order_items)

        try:
            return self._export_to_excel_from_template(
                source_file, source_file_path, order_items,
            )
        except Exception:
            logger.exception(
                "Ошибка экспорта заказа #%s по шаблону прайса, fallback в отдельный Excel",
                self.order.id,
            )
            return self._export_pdf_to_excel(source_file, order_items)

    def _export_to_excel_from_template(
        self, source_file, source_file_path: Path, order_items: list,
    ) -> str:
        """Заполняет колонку «Заказ» в копии исходного Excel-прайса."""
        
        # Загружаем исходный файл Excel
        # Используем data_only=False чтобы работать с формулами и иметь возможность их удалять
        # Создаем копию файла для редактирования
        import shutil
        import tempfile
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file_path = temp_file.name
        temp_file.close()
        
        shutil.copy2(str(source_file_path), temp_file_path)
        wb = load_workbook(temp_file_path, data_only=False, keep_vba=False)
        
        # Создаем словарь для быстрого поиска позиций по их расположению
        # В raw_source_location['row'] хранится номер строки в DataFrame после header=1
        # Но нужно учесть, что header_row может быть разным для разных листов
        # Поэтому сначала создаем словарь с row_num, а excel_row будем вычислять позже
        order_items_dict = {}
        for order_item in self.order.items:
            item_id = self._order_item_id(order_item)
            quantity = order_item.get('quantity')
            if item_id is None:
                continue
            try:
                parsed_item = ParsedItem.objects.get(id=item_id)
                location = parsed_item.raw_source_location
                if location:
                    ws_name = self._sheet_name_in_workbook(location, wb)
                    row_num = self._safe_row_num(location)
                    if ws_name is not None and row_num is not None:
                        key = (ws_name, row_num)
                        order_items_dict[key] = quantity
            except ParsedItem.DoesNotExist:
                continue
        
        # Обрабатываем каждый лист
        # Создаем filled_rows для всех листов сразу
        all_filled_rows = {}
        # Сохраняем индексы колонок "Заказ" для каждого листа
        order_col_indices = {}
        for sheet_name in wb.sheetnames:
            all_filled_rows[sheet_name] = set()
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Определяем номер строки с заголовками
            # Ищем строку с колонкой "Кол-во" (но не "Тип фасовки / кол-во в уп")
            # Это самый надежный способ найти заголовки таблицы
            # Важно: исключаем строку 1, так как там обычно находится информационный текст
            header_row = None
            
            # Сначала ищем строку с "Кол-во" (но не "Тип фасовки / кол-во в уп")
            # Начинаем поиск со строки 2, чтобы исключить строку 1 с информационным текстом
            for row_idx in range(2, min(6, ws.max_row + 1)):
                for col_idx in range(1, min(ws.max_column + 1, 15)):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        cell_str = str(cell_value).lower()
                        # Ищем "кол-во", но не "тип фасовки / кол-во в уп"
                        if 'кол-во' in cell_str and 'тип фасовки' not in cell_str:
                            # Проверяем, что это действительно заголовок таблицы
                            # В строке должны быть и другие заголовки (наименование, стиль и т.д.)
                            header_keywords_found = 0
                            for check_col in range(1, min(ws.max_column + 1, 15)):
                                check_cell = ws.cell(row=row_idx, column=check_col).value
                                if check_cell:
                                    check_str = str(check_cell).lower()
                                    if any(keyword in check_str for keyword in ['наименование пивоварни', 'наименование', 'стиль', 'abv', 'цена']):
                                        header_keywords_found += 1
                            
                            # Если найдено хотя бы 2 других заголовка, это заголовки таблицы
                            if header_keywords_found >= 2:
                                header_row = row_idx
                                break
                if header_row:
                    break
            
            # Если не нашли по "Кол-во", ищем строку с "Заказ"
            # Начинаем поиск со строки 2, чтобы исключить строку 1 с информационным текстом
            if header_row is None:
                for row_idx in range(2, min(6, ws.max_row + 1)):
                    for col_idx in range(1, min(ws.max_column + 1, 15)):
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            cell_str = str(cell_value).lower()
                            # Ищем "заказ", но проверяем, что это заголовок таблицы
                            if 'заказ' in cell_str:
                                # Проверяем, что это действительно заголовок таблицы
                                header_keywords_found = 0
                                for check_col in range(1, min(ws.max_column + 1, 15)):
                                    check_cell = ws.cell(row=row_idx, column=check_col).value
                                    if check_cell:
                                        check_str = str(check_cell).lower()
                                        if any(keyword in check_str for keyword in ['наименование пивоварни', 'наименование', 'стиль', 'abv', 'цена']):
                                            header_keywords_found += 1
                                
                                if header_keywords_found >= 2:
                                    header_row = row_idx
                                    break
                    if header_row:
                        break
            
            # Если не нашли по "Кол-во" или "Заказ", ищем строку с несколькими типичными заголовками
            # Начинаем поиск со строки 2, чтобы исключить строку 1 с информационным текстом
            if header_row is None:
                for row_idx in range(2, min(6, ws.max_row + 1)):
                    header_keywords_found = 0
                    for col_idx in range(1, min(ws.max_column + 1, 15)):
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            cell_str = str(cell_value).lower()
                            if any(keyword in cell_str for keyword in ['наименование пивоварни', 'наименование', 'стиль', 'abv', 'цена']):
                                header_keywords_found += 1
                    
                    if header_keywords_found >= 2:
                        header_row = row_idx
                        break
            
            # Если не нашли заголовки, используем строку 2 по умолчанию
            if header_row is None:
                header_row = 2
            
            # Ищем колонку "Заказ" или "Кол-во" (но не "Тип фасовки / кол-во в уп")
            order_col_idx = None
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=header_row, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).lower()
                    # Ищем "заказ" или "кол-во", но не "тип фасовки"
                    if 'заказ' in cell_str:
                        order_col_idx = col_idx
                        break
                    elif 'кол-во' in cell_str and 'тип фасовки' not in cell_str:
                        order_col_idx = col_idx
                        break
            
            # Если колонки "Заказ" нет, добавляем её справа
            if order_col_idx is None:
                order_col_idx = ws.max_column + 1
                # Записываем заголовок "Заказ"
                hr, hc = self._merge_anchor(ws, header_row, order_col_idx)
                ws.cell(row=hr, column=hc, value='Заказ')
            
            # Сохраняем индекс колонки для этого листа
            order_col_indices[sheet_name] = order_col_idx
            
            # Заполняем колонку "Заказ" для позиций из заказа
            # Сначала заполняем позиции из заказа
            filled_rows = all_filled_rows[sheet_name]
            
            # Заполняем позиции из заказа ТОЛЬКО для этого листа
            for key, quantity in order_items_dict.items():
                if key[0] == sheet_name:
                    row_num = key[1]  # Это row из raw_source_location
                    # Преобразуем row_num в excel_row
                    # Парсер использует header=1, что означает заголовки в строке 2 Excel
                    # row_num - это индекс строки в DataFrame после header=1
                    # Формула: excel_row = row_num + 2 (работает для всех листов, так как все имеют header_row=2)
                    excel_row = row_num + 2
                    
                    if header_row + 1 <= excel_row <= ws.max_row:
                        ar, ac = self._merge_anchor(ws, excel_row, order_col_idx)
                        # Удаляем формулу, если она есть
                        cell = ws.cell(row=ar, column=ac)
                        if cell.data_type == 'f':
                            cell.value = None
                        ws.cell(row=ar, column=ac, value=quantity)
                        filled_rows.add(excel_row)
            
        # Восстанавливаем все значения для всех листов после обработки всех листов
        for sheet_name in wb.sheetnames:
            
            ws_check = wb[sheet_name]
            order_col_idx = order_col_indices.get(sheet_name)
            
            if order_col_idx is None:
                continue
            
            # Восстанавливаем значения для этого листа
            for item in self.order.items:
                item_id = self._order_item_id(item)
                quantity = item.get('quantity')
                if item_id is None:
                    continue

                try:
                    parsed_item = ParsedItem.objects.get(id=item_id)
                    location = parsed_item.raw_source_location
                    if not location or self._sheet_name_in_workbook(location, wb) != sheet_name:
                        continue
                    row_num = self._safe_row_num(location)
                    if row_num is None:
                        continue
                    excel_row = row_num + 2

                    if excel_row <= ws_check.max_row:
                        ar, ac = self._merge_anchor(ws_check, excel_row, order_col_idx)
                        cell_check = ws_check.cell(row=ar, column=ac)
                        # Сравниваем значения, преобразуя в числа для надежности
                        cell_value = cell_check.value
                        if cell_value is not None:
                            try:
                                cell_value = float(cell_value)
                            except (ValueError, TypeError):
                                pass

                        try:
                            quantity_float = float(quantity)
                        except (ValueError, TypeError):
                            quantity_float = quantity

                        if cell_value != quantity_float:
                            # Очищаем формулу, если она есть
                            if cell_check.data_type == 'f':
                                cell_check.value = None
                            ws_check.cell(row=ar, column=ac, value=quantity)
                except ParsedItem.DoesNotExist:
                    continue
        
        # Финальная проверка и исправление значений перед сохранением
        for sheet_name in wb.sheetnames:
            
            ws_final = wb[sheet_name]
            order_col_idx = order_col_indices.get(sheet_name)
            
            if order_col_idx is None:
                continue
            
            # Проверяем и исправляем значения для всех позиций
            for item in self.order.items:
                item_id = self._order_item_id(item)
                quantity = item.get('quantity')
                if item_id is None:
                    continue

                try:
                    parsed_item = ParsedItem.objects.get(id=item_id)
                    location = parsed_item.raw_source_location
                    if not location or self._sheet_name_in_workbook(location, wb) != sheet_name:
                        continue
                    row_num = self._safe_row_num(location)
                    if row_num is None:
                        continue
                    excel_row = row_num + 2

                    if excel_row <= ws_final.max_row:
                        ar, ac = self._merge_anchor(ws_final, excel_row, order_col_idx)
                        cell_final = ws_final.cell(row=ar, column=ac)
                        # Сравниваем значения, преобразуя в числа для надежности
                        cell_value = cell_final.value
                        if cell_value is not None:
                            try:
                                cell_value = float(cell_value)
                            except (ValueError, TypeError):
                                pass

                        try:
                            quantity_float = float(quantity)
                        except (ValueError, TypeError):
                            quantity_float = quantity

                        # Формируем значение с указанием формата
                        format_type = parsed_item.format_type or ''
                        if format_type:
                            # Сокращаем название формата
                            format_short = format_type.lower()
                            if 'кег' in format_short or 'keg' in format_short:
                                format_label = 'кег'
                            elif 'банк' in format_short or 'can' in format_short:
                                format_label = 'банка'
                            elif 'бут' in format_short or 'bottle' in format_short:
                                format_label = 'бут'
                            else:
                                format_label = format_type[:10]
                            order_value = f"{quantity} {format_label}"
                        else:
                            order_value = quantity

                        # Если значение неверное, исправляем
                        if cell_final.data_type == 'f':
                            cell_final.value = None
                        ws_final.cell(row=ar, column=ac, value=order_value)
                except ParsedItem.DoesNotExist:
                    continue
        
        # Сохраняем файл
        export_dir = Path(settings.MEDIA_ROOT) / 'exports'
        export_dir.mkdir(parents=True, exist_ok=True)
        
        # Формируем имя файла: название_исходного_файла-день-месяц.xlsx
        source_filename = Path(source_file.original_filename).stem
        today = datetime.now()
        date_suffix = f"{today.day:02d}-{today.month:02d}"
        filename = f"{source_filename}-{date_suffix}.xlsx"
        file_path = export_dir / filename
        
        # Удаляем старый файл, если он существует
        if file_path.exists():
            file_path.unlink()
        
        # Сохраняем файл
        wb.save(str(file_path))
        
        # Удаляем временный файл
        try:
            os.unlink(temp_file_path)
        except:
            pass
        
        # Сохраняем путь в заказе
        self.order.export_file_path = str(file_path.relative_to(settings.MEDIA_ROOT))
        self.order.save()
        
        return str(file_path)
    
    def _export_pdf_to_excel(self, source_file, order_items) -> str:
        """
        Создаёт отдельный Excel с позициями заказа (PDF, нет прайса на диске или fallback).
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Заказ"
        
        # Заголовки
        headers = ['Пивоварня', 'Название', 'Стиль', 'Цена', 'Валюта', 
                   'Объём', 'Формат', 'Кол-во', 'Сумма']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Данные заказа
        row_num = 2
        total_sum = 0
        total_qty = 0
        
        for order_item in self.order.items:
            item_id = self._order_item_id(order_item)
            quantity = order_item.get('quantity', 1)
            if item_id is None:
                continue

            try:
                parsed_item = ParsedItem.objects.get(id=item_id)
                price = self._safe_float(parsed_item.price) or 0
                item_sum = price * quantity
                total_sum += item_sum
                total_qty += quantity
                
                # Формат
                format_type = parsed_item.format_type or ''
                format_short = format_type.lower()
                if 'кег' in format_short or 'keg' in format_short:
                    format_label = 'кег'
                elif 'банк' in format_short or 'can' in format_short:
                    format_label = 'банка'
                else:
                    format_label = format_type
                
                ws.cell(row=row_num, column=1, value=parsed_item.brewery or '')
                ws.cell(row=row_num, column=2, value=parsed_item.beer_name or '')
                ws.cell(row=row_num, column=3, value=parsed_item.style or '')
                ws.cell(row=row_num, column=4, value=price)
                ws.cell(row=row_num, column=5, value=parsed_item.currency or 'RUB')
                ws.cell(row=row_num, column=6, value=self._safe_float(parsed_item.volume) or '')
                ws.cell(row=row_num, column=7, value=format_label)
                ws.cell(row=row_num, column=8, value=quantity)
                
                # Сумма
                sum_cell = ws.cell(row=row_num, column=9, value=item_sum)
                sum_cell.font = Font(color="008000", bold=True)
                
                row_num += 1
            except ParsedItem.DoesNotExist:
                continue

        if row_num <= 2:
            raise ValueError(
                "Не удалось собрать позиции заказа: записи в прайсе удалены или недоступны."
            )

        # Пустая строка
        row_num += 1

        # Итого
        ws.cell(row=row_num, column=7, value="Итого:")
        ws.cell(row=row_num, column=7).font = Font(bold=True)
        ws.cell(row=row_num, column=7).alignment = Alignment(horizontal='right')
        
        qty_cell = ws.cell(row=row_num, column=8, value=total_qty)
        qty_cell.font = Font(bold=True, color="3498DB")
        
        sum_cell = ws.cell(row=row_num, column=9, value=total_sum)
        sum_cell.font = Font(bold=True, color="008000")
        
        # Автоширина колонок
        col_widths = [15, 35, 20, 10, 8, 8, 10, 10, 12]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Сохраняем
        export_dir = Path(settings.MEDIA_ROOT) / 'exports'
        export_dir.mkdir(parents=True, exist_ok=True)
        
        if source_file and source_file.original_filename:
            source_filename = Path(source_file.original_filename).stem
        else:
            source_filename = f"order_{self.order.id}"
        today = datetime.now()
        date_suffix = f"{today.day:02d}-{today.month:02d}"
        filename = f"{source_filename}-{date_suffix}.xlsx"
        file_path = export_dir / filename

        if file_path.exists():
            file_path.unlink()

        wb.save(str(file_path))

        self.order.export_file_path = str(file_path.relative_to(settings.MEDIA_ROOT))
        self.order.save()

        return str(file_path)

    def export_to_pdf(self) -> str:
        """
        Экспортирует заказ в PDF файл.
        
        Returns:
            Путь к созданному PDF файлу
        """
        # Создаем PDF документ
        export_dir = Path(settings.MEDIA_ROOT) / 'exports'
        export_dir.mkdir(parents=True, exist_ok=True)
        
        # Формируем имя файла: название_исходного_файла-день-месяц.pdf
        source_filename = f"order_{self.order.id}"
        first_item_id = self._first_order_item_id()
        if first_item_id:
            try:
                parsed_item = ParsedItem.objects.get(id=first_item_id)
                source_filename = Path(parsed_item.file.original_filename).stem
            except ParsedItem.DoesNotExist:
                pass
        today = datetime.now()
        date_suffix = f"{today.day:02d}-{today.month:02d}"
        filename = f"{source_filename}-{date_suffix}.pdf"
        file_path = export_dir / filename
        
        doc = SimpleDocTemplate(str(file_path), pagesize=A4)
        story = []
        
        # Стили
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        normal_style = styles['Normal']
        
        # Заголовок
        title = Paragraph("Заказ", title_style)
        story.append(title)
        story.append(Spacer(1, 0.2 * inch))
        
        # Дата создания
        date_text = f"Дата создания: {self.order.created_at.strftime('%d.%m.%Y %H:%M')}"
        story.append(Paragraph(date_text, normal_style))
        story.append(Spacer(1, 0.3 * inch))
        
        # Получаем позиции заказа
        order_items = self._get_order_items()
        
        # Заголовки таблицы
        table_data = [['Пивоварня', 'Название', 'Стиль', 'Крепость', 
                       'Цена', 'Валюта', 'Объём', 'Формат', 'Количество']]
        
        # Данные таблицы
        for item_data in order_items:
            row = [
                item_data.get('brewery', ''),
                item_data.get('beer_name', ''),
                item_data.get('style', ''),
                str(item_data.get('abv') or ''),
                str(item_data.get('price') or ''),
                item_data.get('currency', ''),
                str(item_data.get('volume') or ''),
                item_data.get('format_type', ''),
                str(item_data.get('quantity', ''))
            ]
            table_data.append(row)
        
        # Создаем таблицу
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        
        story.append(table)
        
        # Собираем PDF
        doc.build(story)
        
        # Сохраняем путь в заказе
        self.order.export_file_path = str(file_path.relative_to(settings.MEDIA_ROOT))
        self.order.save()
        
        return str(file_path)
    
    def _get_order_items(self) -> list:
        """
        Получает данные позиций заказа.
        
        Returns:
            Список словарей с данными позиций
        """
        from .normalizers import DataNormalizer
        
        normalizer = DataNormalizer()
        order_items = []
        
        for order_item in self.order.items:
            item_id = self._order_item_id(order_item)
            quantity = order_item.get('quantity')
            if item_id is None:
                continue

            try:
                parsed_item = ParsedItem.objects.get(id=item_id)
                # Нормализуем brewery (удаляем город) перед добавлением в заказ
                brewery_normalized = normalizer.normalize_brewery(parsed_item.brewery) if parsed_item.brewery else ''
                
                item_data = {
                    'brewery': brewery_normalized,
                    'beer_name': parsed_item.beer_name,
                    'style': parsed_item.style,
                    'abv': self._safe_float(parsed_item.abv),
                    'price': self._safe_float(parsed_item.price),
                    'currency': parsed_item.currency,
                    'volume': self._safe_float(parsed_item.volume),
                    'format_type': parsed_item.format_type,
                    'quantity': quantity,
                }
                order_items.append(item_data)
            except ParsedItem.DoesNotExist:
                continue
        
        return order_items


class TapsExporter:
    """
    Класс для экспорта кранов в Excel.
    """
    
    def __init__(self, location: TapLocation):
        """
        Инициализация экспортера.
        
        Args:
            location: Объект локации с кранами
        """
        self.location = location
    
    def export_to_excel(self) -> str:
        """
        Экспортирует краны локации в Excel файл.
        
        Returns:
            Путь к созданному Excel файлу
        """
        # Получаем все краны локации
        taps = Tap.objects.filter(location=self.location).order_by('position')
        
        # Создаем новый Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Краны"
        
        # Заголовки
        headers = ['№', 'Пивоварня', 'Название', 'Цена/л', 'Объём/цена', 'IBU', 'ABV', 'След 1', 'След 2', 'Статус']
        ws.append(headers)
        
        # Стили для заголовков
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Данные кранов
        for tap in taps:
            current_beer = f"{tap.brewery} {tap.beer_name}".strip() if (tap.brewery or tap.beer_name) else "—"
            price = float(tap.price_per_liter) if tap.price_per_liter else None
            status_display = dict(Tap.STATUS_CHOICES).get(tap.status, tap.status)
            
            row = [
                tap.position,
                tap.brewery or '',
                tap.beer_name or '',
                f"{price:.2f}" if price else '',
                tap.volume_price_text or '',
                tap.bitterness_ibu or '',
                tap.abv_text or '',
                tap.next_beer_1 or '',
                tap.next_beer_2 or '',
                status_display,
            ]
            ws.append(row)
        
        # Автоматическая ширина колонок
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(str(header))
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Сохраняем файл
        date_suffix = datetime.now().strftime('%d-%m')
        filename = f"{self.location.name}-краны-{date_suffix}.xlsx"
        export_dir = Path(settings.MEDIA_ROOT) / 'exports'
        export_dir.mkdir(exist_ok=True)
        file_path = export_dir / filename
        
        wb.save(str(file_path))
        
        return str(file_path.relative_to(settings.MEDIA_ROOT))
