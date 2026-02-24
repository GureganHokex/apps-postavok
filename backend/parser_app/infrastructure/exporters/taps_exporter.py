"""
Экспорт кранов в Excel.
"""

import os
from pathlib import Path
from datetime import datetime
from django.conf import settings
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import Optional
from parser_app.models import TapLocation, Tap
import pandas as pd


class TapsExporter:
    """
    Класс для экспорта кранов в Excel.
    """
    
    def __init__(self, location: TapLocation):
        """
        Инициализация экспортера.
        
        Args:
            location: Объект локации с кранами
        """
        self.location = location
    
    def export_to_excel(self) -> str:
        """
        Экспортирует краны локации в Excel файл.
        
        Returns:
            Путь к созданному Excel файлу
        """
        # Получаем все краны локации
        taps = Tap.objects.filter(location=self.location).order_by('position')
        
        # Создаем новый Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Краны"
        
        # Заголовки
        headers = ['№', 'Пивоварня', 'Название', 'Цена/л', 'След 1', 'След 2', 'Статус']
        ws.append(headers)
        
        # Стили для заголовков
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Данные кранов
        for tap in taps:
            current_beer = f"{tap.brewery} {tap.beer_name}".strip() if (tap.brewery or tap.beer_name) else "—"
            price = float(tap.price_per_liter) if tap.price_per_liter else None
            status_display = dict(Tap.STATUS_CHOICES).get(tap.status, tap.status)
            
            row = [
                tap.position,
                tap.brewery or '',
                tap.beer_name or '',
                f"{price:.2f}" if price else '',
                tap.next_beer_1 or '',
                tap.next_beer_2 or '',
                status_display,
            ]
            ws.append(row)
        
        # Автоматическая ширина колонок
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(str(header))
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Сохраняем файл
        date_suffix = datetime.now().strftime('%d-%m')
        filename = f"{self.location.name}-краны-{date_suffix}.xlsx"
        export_dir = Path(settings.MEDIA_ROOT) / 'exports'
        export_dir.mkdir(exist_ok=True)
        file_path = export_dir / filename
        
        wb.save(str(file_path))
        
        return str(file_path.relative_to(settings.MEDIA_ROOT))
