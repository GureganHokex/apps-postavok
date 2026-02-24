"""
Модуль для проверки форматирования Excel файлов.

Определяет скрытые или неактуальные строки по форматированию:
- Скрытые строки (hidden)
- Серый цвет текста
- Зачеркивание текста (strikethrough)
- Серая заливка ячеек
"""

import logging
from typing import Optional

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExcelFormattingChecker:
    """Проверяет форматирование Excel для определения неактуальных позиций."""
    
    def __init__(self, file_path: str):
        """
        Инициализация проверщика форматирования.
        
        Args:
            file_path: Путь к Excel файлу
        """
        self.file_path = file_path
        self._workbook_cache = None
        self._worksheet_cache = {}
    
    def _is_gray_color(self, color) -> bool:
        """
        Проверяет, является ли цвет серым (RGB значения близки друг к другу).
        
        Args:
            color: Объект цвета openpyxl
            
        Returns:
            True если цвет серый
        """
        if not color or not hasattr(color, 'rgb'):
            return False
        rgb = color.rgb
        if not rgb or len(rgb) < 6:
            return False
        try:
            # Конвертируем hex в RGB
            r = int(rgb[2:4], 16)
            g = int(rgb[4:6], 16)
            b = int(rgb[6:8], 16)
            # Серый цвет: все компоненты близки друг к другу
            avg = (r + g + b) / 3
            diff = max(abs(r - avg), abs(g - avg), abs(b - avg))
            # Если разница меньше 30 - это серый цвет
            return diff < 30 and avg < 200  # Темно-серый или светло-серый
        except (ValueError, IndexError):
            return False
    
    def _get_worksheet(self, sheet_name: str):
        """
        Получает worksheet с кэшированием workbook.
        
        Args:
            sheet_name: Имя листа
            
        Returns:
            Worksheet объект или None
        """
        if not OPENPYXL_AVAILABLE:
            return None
        
        if sheet_name in self._worksheet_cache:
            return self._worksheet_cache[sheet_name]
        
        try:
            if self._workbook_cache is None:
                self._workbook_cache = load_workbook(
                    self.file_path, 
                    data_only=True, 
                    read_only=True
                )
            
            if sheet_name not in self._workbook_cache.sheetnames:
                return None
            
            ws = self._workbook_cache[sheet_name]
            self._worksheet_cache[sheet_name] = ws
            return ws
        except Exception as e:
            logger.debug(f"Ошибка при открытии workbook: {str(e)}")
            return None
    
    def is_row_hidden_or_inactive(self, sheet_name: str, excel_row_num: int) -> bool:
        """
        Проверяет, является ли строка скрытой или неактуальной.
        
        Args:
            sheet_name: Имя листа
            excel_row_num: Номер строки в Excel (1-based)
            
        Returns:
            True если строка скрыта или неактуальна
        """
        if not OPENPYXL_AVAILABLE:
            return False
        
        ws = self._get_worksheet(sheet_name)
        if not ws:
            return False
        
        try:
            # Проверяем, скрыта ли строка
            if ws.row_dimensions[excel_row_num].hidden:
                return True
            
            # Проверяем форматирование ячеек в строке (только первые 10 колонок для скорости)
            row = ws[excel_row_num]
            has_gray_text = False
            has_strikethrough = False
            has_gray_fill = False
            
            for cell in list(row)[:10]:
                if cell.value is None:
                    continue
                
                # Проверяем цвет текста
                if cell.font and cell.font.color:
                    if self._is_gray_color(cell.font.color):
                        has_gray_text = True
                
                # Проверяем зачеркивание
                if cell.font and cell.font.strike:
                    has_strikethrough = True
                
                # Проверяем цвет заливки
                if cell.fill and cell.fill.start_color:
                    if self._is_gray_color(cell.fill.start_color):
                        has_gray_fill = True
                
                # Если уже нашли признаки неактуальности, можно прервать проверку
                if has_gray_text or has_strikethrough or has_gray_fill:
                    break
            
            # Если строка имеет признаки неактуальности
            if has_gray_text or has_strikethrough or has_gray_fill:
                reasons = []
                if has_gray_text:
                    reasons.append("серый текст")
                if has_strikethrough:
                    reasons.append("зачеркивание")
                if has_gray_fill:
                    reasons.append("серая заливка")
                logger.debug(f"Строка {excel_row_num} помечена как неактуальная: {', '.join(reasons)}")
                return True
            
            return False
        
        except Exception as e:
            logger.debug(f"Ошибка при проверке форматирования строки {excel_row_num}: {str(e)}")
            return False
    
    def close(self):
        """Закрывает открытые ресурсы."""
        if self._workbook_cache:
            try:
                self._workbook_cache.close()
            except Exception:
                pass
            self._workbook_cache = None
            self._worksheet_cache = {}
