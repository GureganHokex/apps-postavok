"""
Парсер для Excel файлов.
Поддерживает конфигурируемые синонимы колонок и усиленный поиск заголовков.
"""

import json
import pandas as pd
import re
import logging
from typing import List, Dict, Optional, Set
from .base_parser import BaseParser
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl не установлен, проверка форматирования Excel недоступна")
from .supplier_profiles import (
    SupplierProfileDetector, SupplierType,
    DistributorProfile, BreweryProfile
)
from parser_app.infrastructure.config.column_config_loader import (
    get_field_patterns,
    get_header_patterns,
    get_price_header_keywords,
)
from parser_app.shared.utils import normalize_number_str

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """
    Парсер для Excel файлов (.xls, .xlsx).
    
    Использует pandas для чтения и парсинга данных.
    Обрабатывает несколько листов в файле.
    """
    
    def parse(self, supplier_type=None, brewery_name=None) -> List[Dict]:
        """
        Парсинг Excel файла.
        
        Args:
            supplier_type: 'distributor' или 'brewery' (если передан, используется вместо автоопределения)
            brewery_name: Название пивоварни (для типа 'brewery')
        
        Returns:
            Список словарей с данными о позициях
        """
        items = []
        original_file_path = self.file_path
        sheet_names = []
        self.parse_report = {}
        
        # Проверяем, нужен ли временный файл (пробуем прочитать один лист)
        temp_file_path = None
        try:
            # Пробуем прочитать первый лист для проверки
            test_sheet = pd.read_excel(self.file_path, sheet_name=0, header=None, engine='openpyxl', nrows=1)
        except (ValueError, Exception) as e:
            error_msg = str(e)
            if "Value must be either numerical" in error_msg or "wildcard" in error_msg or "Unable to read workbook" in error_msg:
                # Файл содержит проблемные фильтры, создаем временный файл
                logger.warning(f"Обнаружены проблемные фильтры в файле, создаем временный файл без фильтров")
                temp_file_path = self._create_temp_file_without_filters()
                if temp_file_path:
                    self.file_path = temp_file_path
                    logger.info(f"Создан временный файл без фильтров: {temp_file_path}")
                else:
                    logger.error("Не удалось создать временный файл без фильтров")
                    return items
        
        try:
            # Получаем список листов
            try:
                excel_file = pd.ExcelFile(self.file_path, engine='openpyxl')
                sheet_names = excel_file.sheet_names
                excel_file.close()
            except Exception as e:
                # Если не удалось прочитать через ExcelFile, используем обходной путь
                logger.warning(f"Не удалось прочитать через ExcelFile, используем обходной путь: {str(e)}")
                sheet_names = self._get_sheet_names_workaround()
            
            logger.info(f"Открыт файл Excel: {self.file_path}, листов: {len(sheet_names)}")
            
            for sheet_name in sheet_names:
                try:
                    logger.info(f"Обработка листа: {sheet_name}")
                    # Пробуем несколько стратегий чтения
                    parsed_items = self._parse_sheet(None, sheet_name, supplier_type=supplier_type, brewery_name=brewery_name)
                    if parsed_items:
                        items.extend(parsed_items)
                        logger.info(f"Извлечено {len(parsed_items)} позиций из листа {sheet_name}")
                    else:
                        logger.warning(f"Не удалось извлечь позиции из листа {sheet_name}")
                except Exception as e:
                    logger.error(f"Ошибка при обработке листа {sheet_name}: {str(e)}", exc_info=True)
                    continue
        except Exception as e:
            logger.error(f"Ошибка при чтении Excel файла {self.file_path}: {str(e)}", exc_info=True)
            # Если не удалось прочитать как Excel, пробуем как CSV
            try:
                logger.info("Пробуем прочитать как CSV")
                df = pd.read_csv(self.file_path)
                parsed_items = self._parse_dataframe(df, 'Sheet1')
                items.extend(parsed_items)
            except Exception as csv_error:
                logger.error(f"Ошибка при чтении как CSV: {str(csv_error)}")
        finally:
            # Восстанавливаем оригинальный путь и удаляем временный файл
            if temp_file_path:
                self.file_path = original_file_path
                try:
                    import os
                    if os.path.exists(temp_file_path):
                        os.remove(temp_file_path)
                        logger.debug(f"Временный файл удален: {temp_file_path}")
                except Exception as e:
                    logger.warning(f"Не удалось удалить временный файл {temp_file_path}: {str(e)}")
        
        self.parse_report = {'sheets_processed': len(sheet_names) if sheet_names else 0, 'total_items': len(items)}
        return items
    
    def _get_sheet_names_workaround(self) -> List[str]:
        """
        Получает список листов из Excel файла, обходя проблемные фильтры.
        Использует прямое чтение XML из ZIP архива.
        
        Returns:
            Список имен листов
        """
        import zipfile
        import xml.etree.ElementTree as ET
        
        sheet_names = []
        try:
            with zipfile.ZipFile(self.file_path, 'r') as zip_ref:
                # Читаем workbook.xml для получения списка листов
                workbook_xml = zip_ref.read('xl/workbook.xml')
                root = ET.fromstring(workbook_xml)
                
                # Находим все элементы sheet
                ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                for sheet in root.findall('.//main:sheet', ns):
                    sheet_name = sheet.get('name')
                    if sheet_name:
                        sheet_names.append(sheet_name)
        except Exception as e:
            logger.warning(f"Не удалось получить список листов через обходной путь: {str(e)}")
            # Fallback: пробуем стандартные имена
            sheet_names = ['Фасовка', 'Розлив', 'БА']
        
        return sheet_names if sheet_names else ['Sheet1']
    
    def _create_temp_file_without_filters(self) -> Optional[str]:
        """
        Создает временный файл Excel без проблемных фильтров.
        Удаляет элементы autoFilter из всех листов.
        
        Returns:
            Путь к временному файлу или None в случае ошибки
        """
        import zipfile
        import xml.etree.ElementTree as ET
        import tempfile
        import os
        import shutil
        
        try:
            # Создаем временный файл
            temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
            os.close(temp_fd)
            
            # Копируем исходный файл во временный
            shutil.copy2(self.file_path, temp_path)
            
            # Открываем временный файл как ZIP
            with zipfile.ZipFile(temp_path, 'r') as zip_read:
                # Создаем новый ZIP файл без фильтров
                temp_zip_path = temp_path + '.new'
                with zipfile.ZipFile(temp_zip_path, 'w', zipfile.ZIP_DEFLATED) as zip_write:
                    # Копируем все файлы, кроме проблемных worksheet файлов
                    for item in zip_read.infolist():
                        data = zip_read.read(item.filename)
                        
                        # Если это worksheet файл, удаляем фильтры
                        if item.filename.startswith('xl/worksheets/sheet') and item.filename.endswith('.xml'):
                            try:
                                root = ET.fromstring(data)
                                # Удаляем все элементы autoFilter (используем рекурсивный поиск)
                                auto_filter_ns = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}autoFilter'
                                for parent in root.iter():
                                    for child in list(parent):
                                        if child.tag == auto_filter_ns:
                                            parent.remove(child)
                                
                                # Сохраняем измененный XML
                                data = ET.tostring(root, encoding='utf-8', xml_declaration=True)
                            except Exception as e:
                                logger.warning(f"Не удалось обработать {item.filename}: {str(e)}")
                                # Используем исходные данные
                        
                        zip_write.writestr(item, data)
            
            # Заменяем временный файл новым
            os.replace(temp_zip_path, temp_path)
            
            logger.info(f"Создан временный файл без фильтров: {temp_path}")
            return temp_path
            
        except Exception as e:
            logger.error(f"Ошибка при создании временного файла: {str(e)}", exc_info=True)
            # Удаляем временный файл в случае ошибки
            try:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
            except:
                pass
            return None
    
    def _parse_sheet(self, excel_file: Optional[pd.ExcelFile], sheet_name: str, supplier_type=None, brewery_name=None) -> List[Dict]:
        """
        Парсит один лист Excel файла, пробуя разные стратегии.
        
        Args:
            excel_file: Объект ExcelFile
            sheet_name: Имя листа
            
        Returns:
            Список словарей с данными позиций
        """
        # Получаем имя файла для анализа
        import os
        file_name = os.path.basename(self.file_path)
        
        # Сначала определяем тип поставщика
        detector = SupplierProfileDetector()
        
        # Стратегия 1: Ищем строку заголовков в первых 20 строках
        try:
            # Используем прямой путь к файлу вместо объекта ExcelFile для избежания ошибок
            logger.debug(f"Стратегия 1: Читаем лист '{sheet_name}' из файла '{self.file_path}'")
            df_full = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None, engine='openpyxl')
            logger.debug(f"Стратегия 1: Прочитано {len(df_full)} строк")
            df_sample = df_full.head(20) if len(df_full) > 20 else df_full
            header_row = self._find_header_row(df_sample)
            
            if header_row is not None:
                logger.debug(f"Найдена строка заголовков на строке {header_row + 1}")
                df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=header_row, engine='openpyxl')
                logger.debug(f"Стратегия 1: DataFrame с заголовками создан, строк: {len(df)}")
                
                # Определяем тип поставщика (если не передан)
                if supplier_type is None:
                    try:
                        supplier_type, characteristics = detector.detect(df, sheet_name, file_name=file_name)
                        logger.info(f"Лист {sheet_name}: тип поставщика = {supplier_type.value}")
                    except Exception as e:
                        logger.warning(f"Ошибка при определении типа поставщика: {str(e)}", exc_info=True)
                        supplier_type = SupplierType.UNKNOWN
                        characteristics = {}
                else:
                    # Используем переданный тип поставщика
                    if isinstance(supplier_type, str):
                        if supplier_type == 'distributor':
                            supplier_type_enum = SupplierType.DISTRIBUTOR
                        elif supplier_type == 'brewery':
                            supplier_type_enum = SupplierType.BREWERY
                        else:
                            supplier_type_enum = SupplierType.UNKNOWN
                    else:
                        supplier_type_enum = supplier_type
                    supplier_type = supplier_type_enum
                    characteristics = {'single_brewery_name': brewery_name} if brewery_name else {}
                    logger.info(f"Лист {sheet_name}: используется переданный тип поставщика = {supplier_type.value}")
                
                parsed_items = self._parse_dataframe(df, sheet_name, supplier_type=supplier_type, 
                                           characteristics=characteristics, brewery_name=brewery_name)
                logger.debug(f"Стратегия 1: Извлечено {len(parsed_items)} позиций")
                if parsed_items:
                    return parsed_items
            else:
                logger.debug("Стратегия 1: Строка заголовков не найдена")
        except (ValueError, Exception) as e:
            error_msg = str(e)
            if "Value must be either numerical" in error_msg or "wildcard" in error_msg or "Unable to read workbook" in error_msg:
                # Проблемные фильтры - создаем временный файл и пробуем снова
                logger.warning(f"Обнаружены проблемные фильтры при чтении листа {sheet_name}, создаем временный файл")
                temp_file_path = self._create_temp_file_without_filters()
                if temp_file_path:
                    original_path = self.file_path
                    self.file_path = temp_file_path
                    try:
                        # Пробуем снова с временным файлом
                        df_full = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None, engine='openpyxl')
                        df_sample = df_full.head(20) if len(df_full) > 20 else df_full
                        header_row = self._find_header_row(df_sample)
                        
                        if header_row is not None:
                            df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=header_row, engine='openpyxl')
                            
                            if supplier_type is None:
                                try:
                                    supplier_type, characteristics = detector.detect(df, sheet_name, file_name=file_name)
                                    logger.info(f"Лист {sheet_name}: тип поставщика = {supplier_type.value}")
                                except Exception as e2:
                                    logger.warning(f"Ошибка при определении типа поставщика: {str(e2)}")
                                    supplier_type = SupplierType.UNKNOWN
                                    characteristics = {}
                            else:
                                if isinstance(supplier_type, str):
                                    supplier_type = SupplierType.DISTRIBUTOR if supplier_type == 'distributor' else (SupplierType.BREWERY if supplier_type == 'brewery' else SupplierType.UNKNOWN)
                                characteristics = {'single_brewery_name': brewery_name} if brewery_name else {}
                                logger.info(f"Лист {sheet_name}: используется переданный тип поставщика = {supplier_type.value}")
                            
                            parsed_items = self._parse_dataframe(df, sheet_name, supplier_type=supplier_type,
                                                brewery_name=brewery_name, 
                                                   characteristics=characteristics)
                            if parsed_items:
                                # НЕ удаляем временный файл здесь - он будет удален в finally блоке parse()
                                # НЕ восстанавливаем путь здесь - это сделает parse()
                                return parsed_items
                    except Exception as e4:
                        logger.error(f"Ошибка при чтении временного файла: {str(e4)}", exc_info=True)
                    finally:
                        # НЕ удаляем временный файл здесь - он будет удален в finally блоке parse()
                        # НЕ восстанавливаем путь здесь - это сделает parse()
                        pass
            logger.debug(f"Стратегия 1 не сработала: {error_msg}")
        
        # Стратегия 2: Пробуем header=1 (вторая строка)
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=1, engine='openpyxl')
            
            if supplier_type is None:
                try:
                    supplier_type, characteristics = detector.detect(df, sheet_name, file_name=file_name)
                    logger.info(f"Лист {sheet_name}: тип поставщика = {supplier_type.value}")
                except Exception as e:
                    logger.warning(f"Ошибка при определении типа поставщика: {str(e)}", exc_info=True)
                    supplier_type = SupplierType.UNKNOWN
                    characteristics = {}
            else:
                if isinstance(supplier_type, str):
                    supplier_type = SupplierType.DISTRIBUTOR if supplier_type == 'distributor' else (SupplierType.BREWERY if supplier_type == 'brewery' else SupplierType.UNKNOWN)
                characteristics = {'single_brewery_name': brewery_name} if brewery_name else {}
                logger.info(f"Лист {sheet_name}: используется переданный тип поставщика = {supplier_type.value}")
            
            parsed_items = self._parse_dataframe(df, sheet_name, supplier_type=supplier_type,
                                                characteristics=characteristics, brewery_name=brewery_name)
            if parsed_items:
                logger.debug("Стратегия 2 (header=1) успешна")
                return parsed_items
        except Exception as e:
            logger.debug(f"Стратегия 2 не сработала: {str(e)}")
        
        # Стратегия 3: Пробуем header=0 (первая строка)
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=0, engine='openpyxl')
            
            if supplier_type is None:
                try:
                    supplier_type, characteristics = detector.detect(df, sheet_name, file_name=file_name)
                    logger.info(f"Лист {sheet_name}: тип поставщика = {supplier_type.value}")
                except Exception as e:
                    logger.warning(f"Ошибка при определении типа поставщика: {str(e)}", exc_info=True)
                    supplier_type = SupplierType.UNKNOWN
                    characteristics = {}
            else:
                if isinstance(supplier_type, str):
                    supplier_type = SupplierType.DISTRIBUTOR if supplier_type == 'distributor' else (SupplierType.BREWERY if supplier_type == 'brewery' else SupplierType.UNKNOWN)
                characteristics = {'single_brewery_name': brewery_name} if brewery_name else {}
                logger.info(f"Лист {sheet_name}: используется переданный тип поставщика = {supplier_type.value}")
            
            parsed_items = self._parse_dataframe(df, sheet_name, supplier_type=supplier_type,
                                                characteristics=characteristics, brewery_name=brewery_name)
            if parsed_items:
                logger.debug("Стратегия 3 (header=0) успешна")
                return parsed_items
        except Exception as e:
            logger.debug(f"Стратегия 3 не сработала: {str(e)}")
        
        # Стратегия 4: Пробуем header=2 и header=3 (заголовок на 3–4 строке — частый случай в прайсах)
        for header_row in [2, 3]:
            try:
                df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=header_row, engine='openpyxl')
                if df.empty or len(df.columns) < 2:
                    continue
                try:
                    st, ch = detector.detect(df, sheet_name, file_name=file_name)
                    if supplier_type is None:
                        supplier_type, characteristics = st, ch
                except Exception:
                    if supplier_type is None:
                        supplier_type, characteristics = SupplierType.UNKNOWN, {}
                parsed_items = self._parse_dataframe(df, sheet_name, supplier_type=supplier_type,
                                                    characteristics=characteristics, brewery_name=brewery_name)
                if parsed_items:
                    logger.info(f"Стратегия 4 (header={header_row}) успешна для листа {sheet_name}")
                    return parsed_items
            except Exception as e:
                logger.debug(f"Стратегия 4 header={header_row} не сработала: {str(e)}")
        
        # Стратегия 5: Читаем без заголовков и ищем их в данных
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None, engine='openpyxl')
            header_row = self._find_header_row(df)
            if header_row is not None:
                col_mapping = self._map_columns(df.iloc[header_row].tolist(), df)
                if col_mapping:
                    df_data = df.iloc[header_row + 1:].reset_index(drop=True)
                    
                    # Определяем тип поставщика (если не передан)
                    if supplier_type is None:
                        try:
                            supplier_type, characteristics = detector.detect(df_data, sheet_name, file_name=file_name)
                            logger.info(f"Лист {sheet_name}: тип поставщика = {supplier_type.value}")
                        except Exception as e:
                            logger.warning(f"Ошибка при определении типа поставщика: {str(e)}", exc_info=True)
                            supplier_type = SupplierType.UNKNOWN
                            characteristics = {}
                    else:
                        if isinstance(supplier_type, str):
                            supplier_type = SupplierType.DISTRIBUTOR if supplier_type == 'distributor' else (SupplierType.BREWERY if supplier_type == 'brewery' else SupplierType.UNKNOWN)
                        characteristics = {'single_brewery_name': brewery_name} if brewery_name else {}
                        logger.info(f"Лист {sheet_name}: используется переданный тип поставщика = {supplier_type.value}")
                    
                    return self._parse_dataframe(df_data, sheet_name, col_mapping=col_mapping, 
                                                supplier_type=supplier_type,
                                                characteristics=characteristics,
                                                brewery_name=brewery_name)
        except Exception as e:
            logger.debug(f"Стратегия 5 не сработала: {str(e)}")
        
        return []
    
    def _is_row_hidden_or_inactive_deprecated(self, sheet_name: str, excel_row_num: int, ws=None) -> bool:
        """
        Проверяет, является ли строка скрытой или неактуальной по форматированию Excel.
        
        Args:
            sheet_name: Имя листа
            excel_row_num: Номер строки в Excel (1-based)
            ws: Открытый worksheet (опционально, для оптимизации)
            
        Returns:
            True если строка скрыта или неактуальна
        """
        if not OPENPYXL_AVAILABLE:
            return False
        
        # Функция для определения, является ли цвет серым
        def is_gray_color(color):
            """Проверяет, является ли цвет серым (RGB значения близки друг к другу)"""
            if not color or not hasattr(color, 'rgb'):
                return False
            rgb = color.rgb
            if not rgb or len(rgb) < 6:
                return False
            try:
                # Конвертируем hex в RGB
                r = int(rgb[2:4], 16)
                g = int(rgb[4:6], 16)
                b = int(rgb[6:8], 16)
                # Серый цвет: все компоненты близки друг к другу
                avg = (r + g + b) / 3
                diff = max(abs(r - avg), abs(g - avg), abs(b - avg))
                # Если разница меньше 30 - это серый цвет
                return diff < 30 and avg < 200  # Темно-серый или светло-серый
            except (ValueError, IndexError):
                return False
        
        try:
            # Если worksheet передан, используем его (для оптимизации)
            if ws is None:
                wb = load_workbook(self.file_path, data_only=True, read_only=True)
                if sheet_name not in wb.sheetnames:
                    wb.close()
                    return False
                ws = wb[sheet_name]
                should_close = True
            else:
                should_close = False
            
            # Проверяем, скрыта ли строка
            if ws.row_dimensions[excel_row_num].hidden:
                if should_close:
                    wb.close()
                return True
            
            # Проверяем форматирование ячеек в строке (только первые несколько колонок для скорости)
            row = ws[excel_row_num]
            has_gray_text = False
            has_strikethrough = False
            has_gray_fill = False
            
            # Проверяем только первые 10 колонок для оптимизации
            for cell in list(row)[:10]:
                if cell.value is None:
                    continue
                
                # Проверяем цвет текста
                if cell.font and cell.font.color:
                    if is_gray_color(cell.font.color):
                        has_gray_text = True
                
                # Проверяем зачеркивание
                if cell.font and cell.font.strike:
                    has_strikethrough = True
                
                # Проверяем цвет заливки
                if cell.fill and cell.fill.start_color:
                    if is_gray_color(cell.fill.start_color):
                        has_gray_fill = True
                
                # Если уже нашли признаки неактуальности, можно прервать проверку
                if has_gray_text or has_strikethrough or has_gray_fill:
                    break
            
            if should_close:
                wb.close()
            
            # Если строка имеет признаки неактуальности
            if has_gray_text or has_strikethrough or has_gray_fill:
                reasons = []
                if has_gray_text:
                    reasons.append("серый текст")
                if has_strikethrough:
                    reasons.append("зачеркивание")
                if has_gray_fill:
                    reasons.append("серая заливка")
                logger.debug(f"Строка {excel_row_num} помечена как неактуальная: {', '.join(reasons)}")
                return True
            
            return False
        
        except Exception as e:
            logger.debug(f"Ошибка при проверке форматирования строки {excel_row_num}: {str(e)}")
            return False
    
    def _get_hidden_or_inactive_rows(self, sheet_name: str) -> Set[int]:
        """
        Определяет скрытые или неактуальные строки по форматированию Excel.
        
        Проверяет:
        - Скрытые строки (hidden)
        - Серый цвет текста (неактуальные позиции)
        - Зачеркивание текста (strikethrough)
        - Серый цвет заливки ячеек
        
        Args:
            sheet_name: Имя листа
            
        Returns:
            Множество индексов строк (0-based), которые нужно пропустить
        """
        hidden_rows = set()
        
        if not OPENPYXL_AVAILABLE:
            return hidden_rows
        
        try:
            wb = load_workbook(self.file_path, data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return hidden_rows
            
            ws = wb[sheet_name]
            
            # Функция для определения, является ли цвет серым
            def is_gray_color(color):
                """Проверяет, является ли цвет серым (RGB значения близки друг к другу)"""
                if not color or not hasattr(color, 'rgb'):
                    return False
                rgb = color.rgb
                if not rgb or len(rgb) < 6:
                    return False
                try:
                    # Конвертируем hex в RGB
                    r = int(rgb[2:4], 16)
                    g = int(rgb[4:6], 16)
                    b = int(rgb[6:8], 16)
                    # Серый цвет: все компоненты близки друг к другу
                    avg = (r + g + b) / 3
                    diff = max(abs(r - avg), abs(g - avg), abs(b - avg))
                    # Если разница меньше 30 - это серый цвет
                    return diff < 30 and avg < 200  # Темно-серый или светло-серый
                except (ValueError, IndexError):
                    return False
            
            # Проверяем каждую строку
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=0):
                # Проверяем, скрыта ли строка
                if ws.row_dimensions[row_idx + 1].hidden:
                    hidden_rows.add(row_idx)
                    logger.debug(f"Строка {row_idx + 1} скрыта (hidden)")
                    continue
                
                # Проверяем форматирование ячеек в строке
                has_gray_text = False
                has_strikethrough = False
                has_gray_fill = False
                
                for cell in row:
                    if cell.value is None:
                        continue
                    
                    # Проверяем цвет текста
                    if cell.font and cell.font.color:
                        if is_gray_color(cell.font.color):
                            has_gray_text = True
                    
                    # Проверяем зачеркивание
                    if cell.font and cell.font.strike:
                        has_strikethrough = True
                    
                    # Проверяем цвет заливки
                    if cell.fill and cell.fill.start_color:
                        if is_gray_color(cell.fill.start_color):
                            has_gray_fill = True
                
                # Если строка имеет признаки неактуальности - добавляем в список
                if has_gray_text or has_strikethrough or has_gray_fill:
                    hidden_rows.add(row_idx)
                    reasons = []
                    if has_gray_text:
                        reasons.append("серый текст")
                    if has_strikethrough:
                        reasons.append("зачеркивание")
                    if has_gray_fill:
                        reasons.append("серая заливка")
                    logger.debug(f"Строка {row_idx + 1} помечена как неактуальная: {', '.join(reasons)}")
            
            wb.close()
            
            if hidden_rows:
                logger.info(f"Найдено {len(hidden_rows)} скрытых/неактуальных строк в листе '{sheet_name}'")
        
        except Exception as e:
            logger.warning(f"Ошибка при проверке форматирования листа '{sheet_name}': {str(e)}")
        
        return hidden_rows
    
    def _parse_dataframe(self, df: pd.DataFrame, sheet_name: str, 
                        col_mapping: Optional[Dict[str, int]] = None,
                        supplier_type: SupplierType = SupplierType.UNKNOWN,
                        characteristics: Dict = None,
                        brewery_name: Optional[str] = None) -> List[Dict]:
        """
        Парсинг DataFrame из Excel.
        
        Args:
            df: DataFrame с данными
            sheet_name: Имя листа
            col_mapping: Предопределенный маппинг колонок (опционально)
            
        Returns:
            Список словарей с данными позиций
        """
        items = []
        
        if df.empty:
            logger.debug(f"DataFrame пуст для листа {sheet_name}")
            return items
        
        # Обрабатываем переданный supplier_type (может быть строка или enum)
        supplier_type_enum = None
        if isinstance(supplier_type, str):
            if supplier_type == 'distributor':
                supplier_type_enum = SupplierType.DISTRIBUTOR
            elif supplier_type == 'brewery':
                supplier_type_enum = SupplierType.BREWERY
            else:
                supplier_type_enum = SupplierType.UNKNOWN
        elif supplier_type is not None:
            supplier_type_enum = supplier_type
        
        # Определяем тип поставщика, если не передан
        if supplier_type_enum is None:
            try:
                import os
                file_name = os.path.basename(self.file_path)
                detector = SupplierProfileDetector()
                supplier_type_enum, characteristics = detector.detect(df, sheet_name, file_name=file_name)
                logger.info(f"Автоопределение типа поставщика: {supplier_type_enum.value}")
            except Exception as e:
                logger.warning(f"Ошибка при автоопределении типа поставщика: {str(e)}", exc_info=True)
                supplier_type_enum = SupplierType.UNKNOWN
                characteristics = {}
        
        # Если передан brewery_name, используем его для создания профиля
        if brewery_name and supplier_type_enum == SupplierType.BREWERY:
            characteristics = {'single_brewery_name': brewery_name}
            logger.info(f"Используется переданное название пивоварни: {brewery_name}")
        
        # Убеждаемся, что supplier_type не None
        if supplier_type_enum is None:
            supplier_type_enum = SupplierType.UNKNOWN
        
        if characteristics is None:
            characteristics = {}
        
        # Если маппинг не передан, определяем его с учетом типа поставщика
        if col_mapping is None:
            col_mapping = {}
            # Применяем профиль поставщика
            if supplier_type_enum:
                try:
                    if supplier_type_enum == SupplierType.DISTRIBUTOR:
                        profile = DistributorProfile()
                        profile_mapping = profile.get_column_mapping_strategy(df, characteristics)
                        col_mapping.update(profile_mapping)
                        logger.debug(f"Применен профиль дистрибьютора, начальный маппинг: {col_mapping}")
                    elif supplier_type_enum == SupplierType.BREWERY:
                        brewery_name_to_use = brewery_name or characteristics.get('single_brewery_name')
                        if brewery_name_to_use:
                            profile = BreweryProfile(brewery_name_to_use)
                            profile_mapping = profile.get_column_mapping_strategy(df, characteristics)
                            col_mapping.update(profile_mapping)
                            logger.debug(f"Применен профиль пивоварни ({brewery_name_to_use}), начальный маппинг: {col_mapping}")
                except Exception as e:
                    logger.warning(f"Ошибка при применении профиля поставщика: {str(e)}", exc_info=True)
            
            # Проверяем, есть ли названия колонок (если читали с header=None, колонки будут 0, 1, 2...)
            has_named_columns = not all(isinstance(col, (int, float)) for col in df.columns)
            
            if has_named_columns:
                # Если колонки именованные, пробуем определить маппинг по названиям
                mapping_by_names = self._map_columns(df.columns.tolist(), df)
                # КРИТИЧЕСКИ ВАЖНО: mapping_by_names имеет приоритет над профилем
                # Особенно для beer_name - колонка "Название" должна иметь приоритет над профилем
                # Объединяем: сначала профиль, потом mapping_by_names (он перезапишет профиль)
                col_mapping = {**col_mapping, **mapping_by_names}
                logger.debug(f"Маппинг по названиям колонок: {mapping_by_names}, объединенный: {col_mapping}")
            
            # Если не нашли маппинг по названиям колонок, ищем заголовки в данных
            if not col_mapping:
                header_row = self._find_header_row(df)
                if header_row is not None:
                    mapping_by_header = self._map_columns(
                        df.iloc[header_row].tolist(), df
                    )
                    col_mapping = {**mapping_by_header, **col_mapping}
                    df = df.iloc[header_row + 1:].reset_index(drop=True)
                    logger.debug(f"Найден заголовок в строке {header_row}, маппинг: {mapping_by_header}")
        
        # Если маппинг все еще не найден, используем эвристику по позициям колонок
        if not col_mapping:
            col_mapping = self._guess_column_mapping(df)
            logger.debug(f"Маппинг по эвристике: {col_mapping}")
        
        # Если все еще нет маппинга, пробуем определить по содержимому первых строк
        if not col_mapping:
            col_mapping = self._analyze_data_content(df)
            logger.debug(f"Маппинг по анализу содержимого: {col_mapping}")
        
        # Для частных поставщиков: при неполном маппинге дополняем универсальным (название + цена)
        if supplier_type_enum == SupplierType.BREWERY and col_mapping:
            if 'price' not in col_mapping or 'beer_name' not in col_mapping:
                fallback = self._fallback_generic_mapping(df)
                for key in ('beer_name', 'price'):
                    if key not in col_mapping and key in fallback:
                        col_mapping[key] = fallback[key]
                        logger.info(f"Лист {sheet_name}: дополнен маппинг для частного поставщика: {key}={fallback[key]}")
        
        # Если маппинг всё ещё пуст — последняя попытка универсальным fallback
        if not col_mapping:
            col_mapping = self._fallback_generic_mapping(df)
            logger.debug(f"Маппинг fallback: {col_mapping}")
        
        # Проверяем, не была ли колонка stock неправильно определена
        # Это нужно делать для всех листов, даже если маппинг уже определен
        if 'stock' in col_mapping and not df.empty:
            stock_col_idx = col_mapping['stock']
            try:
                sample_rows = df.head(min(10, len(df)))
                col_data = sample_rows.iloc[:, stock_col_idx].dropna().astype(str).tolist()
                if col_data:
                    avg_len = sum(len(v) for v in col_data[:5]) / min(5, len(col_data))
                    text_lower = ' '.join(col_data[:5]).lower()
                    
                    # Проверяем, не является ли это форматом упаковки
                    format_keywords = [
                        'банка', 'can', 'кега', 'keg', 'кег', 'бутылка', 'bottle',
                        'ж/б', 'бут', 'формат', 'format', 'упаковка', 'packaging',
                        'тара', 'фасовка', 'фасовки', 'тип фасовки'
                    ]
                    has_format_keywords = any(
                        keyword in text_lower for keyword in format_keywords
                    )
                    
                    # Если содержит ключевые слова формата и короткая - это формат, а не остатки
                    if has_format_keywords and avg_len <= 15:
                        logger.debug(f"Лист {sheet_name}: Колонка {stock_col_idx} переопределена: stock -> format_type (содержит ключевые слова формата упаковки)")
                        if 'format_type' not in col_mapping:
                            col_mapping['format_type'] = stock_col_idx
                        del col_mapping['stock']
                    else:
                        # Проверяем числовые значения только если это не формат
                        numeric_values = []
                        for val in col_data[:5]:
                            try:
                                num_val = float(str(val).replace(',', '.').replace(' ', ''))
                                numeric_values.append(num_val)
                            except (ValueError, TypeError):
                                pass
                        
                        # Проверяем, не является ли это ценой
                        # Цена обычно в диапазоне 50-10000 и может совпадать с остатками
                        if numeric_values and 'price' in col_mapping:
                            price_col_idx = col_mapping['price']
                            try:
                                price_data = sample_rows.iloc[:, price_col_idx].dropna().astype(str).tolist()
                                price_numeric = []
                                for val in price_data[:5]:
                                    try:
                                        num_val = float(str(val).replace(',', '.').replace(' ', ''))
                                        price_numeric.append(num_val)
                                    except (ValueError, TypeError):
                                        pass
                                
                                # Если значения совпадают с ценой - это цена, а не остатки
                                if price_numeric and numeric_values:
                                    stock_avg = sum(numeric_values) / len(numeric_values)
                                    price_avg = sum(price_numeric) / len(price_numeric)
                                    # Если средние значения очень близки (разница менее 1%) - это цена
                                    if abs(stock_avg - price_avg) / max(price_avg, 1) < 0.01:
                                        logger.debug(f"Лист {sheet_name}: Колонка {stock_col_idx} переопределена: stock -> price (значения совпадают с ценой)")
                                        del col_mapping['stock']
                            except Exception:
                                pass
                        
                        # Проверяем, не является ли это объемом
                        # Объем обычно в диапазоне 0.1-50 литров
                        if numeric_values and 'volume' in col_mapping:
                            volume_col_idx = col_mapping['volume']
                            try:
                                volume_data = sample_rows.iloc[:, volume_col_idx].dropna().astype(str).tolist()
                                volume_numeric = []
                                for val in volume_data[:5]:
                                    try:
                                        num_val = float(str(val).replace(',', '.').replace(' ', ''))
                                        volume_numeric.append(num_val)
                                    except (ValueError, TypeError):
                                        pass
                                
                                # Если значения совпадают с объемом - это объем, а не остатки
                                if volume_numeric and numeric_values:
                                    stock_avg = sum(numeric_values) / len(numeric_values)
                                    volume_avg = sum(volume_numeric) / len(volume_numeric)
                                    # Если средние значения очень близки (разница менее 1%) - это объем
                                    if abs(stock_avg - volume_avg) / max(volume_avg, 0.01) < 0.01:
                                        logger.debug(f"Лист {sheet_name}: Колонка {stock_col_idx} переопределена: stock -> volume (значения совпадают с объемом)")
                                        del col_mapping['stock']
                            except Exception:
                                pass
                    
                    # Проверяем, не является ли это ценой по диапазону значений
                    # Цена обычно в диапазоне 50-10000
                    if numeric_values and 'price' not in col_mapping:
                        stock_avg = sum(numeric_values) / len(numeric_values)
                        # Если среднее значение в диапазоне цен и нет ключевых слов остатков - это может быть цена
                        if 50 <= stock_avg <= 10000:
                            # Проверяем заголовок колонки
                            headers = df.columns.tolist() if hasattr(df.columns, 'tolist') else []
                            if stock_col_idx < len(headers):
                                header_lower = str(headers[stock_col_idx]).lower()
                                stock_header_keywords = ['остаток', 'stock', 'наличие', 'availability', 'склад', 'количество']
                                price_header_keywords = ['цена', 'price', 'стоимость', 'cost', 'руб']
                                
                                # Если в заголовке есть ключевые слова цены, но нет остатков - это цена
                                has_price_header = any(kw in header_lower for kw in price_header_keywords)
                                has_stock_header = any(kw in header_lower for kw in stock_header_keywords)
                                
                                if has_price_header and not has_stock_header:
                                    logger.debug(f"Лист {sheet_name}: Колонка {stock_col_idx} переопределена: stock -> price (заголовок указывает на цену)")
                                    if 'price' not in col_mapping:
                                        col_mapping['price'] = stock_col_idx
                                    del col_mapping['stock']
                    
                    # Проверяем, не является ли это стилем пива
                    style_keywords_check = [
                        'ipa', 'lager', 'ale', 'stout', 'porter', 'pilsner', 'wheat', 'sour',
                        'gose', 'neipa', 'hazy', 'imperial', 'double', 'triple',
                        'weisse', 'weizen', 'hopfen', 'belgian', 'witbier', 'saison',
                        'томатный', 'tomato', 'georgian', 'new england', 'grapefruit',
                        'strawberry', 'sour ale', 'barrel', 'aged', 'wild'
                    ]
                    has_style_keywords = any(
                        keyword in text_lower for keyword in style_keywords_check
                    )
                    
                    # Если содержит ключевые слова стилей - это стиль, а не остатки
                    if has_style_keywords and 3 <= avg_len <= 50:
                        logger.debug(f"Лист {sheet_name}: Колонка {stock_col_idx} переопределена: stock -> style (содержит ключевые слова стилей пива, средняя длина: {avg_len:.1f})")
                        # Если style еще не определен, используем эту колонку
                        if 'style' not in col_mapping:
                            col_mapping['style'] = stock_col_idx
                        del col_mapping['stock']
                    # Если средняя длина больше 50 символов - это описание, а не остатки
                    elif avg_len > 50:
                        description_keywords = [
                            'вкус', 'аромат', 'вкусовые', 'характеристики', 'описание',
                            'taste', 'aroma', 'flavor', 'description', 'характер',
                            'насыщенный', 'сочный', 'яркий', 'кислый', 'сладкий'
                        ]
                        has_description_keywords = any(
                            keyword in text_lower for keyword in description_keywords
                        )
                        # Если содержит описательные слова - это точно описание
                        if has_description_keywords:
                            logger.debug(f"Лист {sheet_name}: Колонка {stock_col_idx} переопределена: stock -> description (длинный текст с описательными словами, средняя длина: {avg_len:.1f})")
                            # Если description еще не определен, используем эту колонку
                            if 'description' not in col_mapping:
                                col_mapping['description'] = stock_col_idx
                            del col_mapping['stock']
            except Exception as e:
                logger.debug(f"Ошибка при проверке колонки stock для листа {sheet_name}: {str(e)}")
        
        # Если маппинг определен частично, но не все ключевые поля найдены,
        # запускаем агрессивный анализ содержимого всех колонок
        essential_fields = ['beer_name', 'price']
        missing_fields = [f for f in essential_fields if f not in col_mapping]
        
        if missing_fields and not df.empty:
            logger.debug(f"Не найдены поля: {missing_fields}, запускаем анализ содержимого")
            col_mapping = self._aggressive_column_mapping(df, col_mapping)
            logger.debug(f"Маппинг после агрессивного анализа: {col_mapping}")
        
        # Fallback: универсальный маппинг по типу данных (первая текстовая колонка = название, первая числовая в диапазоне цен = цена)
        if not col_mapping or not all(f in col_mapping for f in essential_fields):
            if not df.empty:
                fallback = self._fallback_generic_mapping(df)
                if fallback:
                    for k, v in fallback.items():
                        if k not in col_mapping:
                            col_mapping[k] = v
                    logger.info(f"Лист {sheet_name}: применён fallback-маппинг {fallback}")
        
        if not col_mapping:
            logger.warning(f"Не удалось определить маппинг колонок для листа {sheet_name}")
            return items
        
        # Различение названия и описания: если в beer_name попал длинный описательный текст — переназначить
        col_mapping = self._disambiguate_name_vs_description(df, col_mapping, sheet_name)
        
        logger.debug(f"Используется маппинг для листа {sheet_name}: {col_mapping}")
        # #region agent log
        try:
            import os
            _lp = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', '..', '.cursor', 'debug-6958ea.log'))
            _h = df.columns.tolist() if df is not None and hasattr(df, 'columns') else []
            _bn = col_mapping.get('beer_name')
            _dc = col_mapping.get('description')
            _data = {"hypothesisId": "H1", "message": "col_mapping names", "data": {"beer_name_idx": _bn, "beer_name_header": str(_h[_bn]) if _bn is not None and _bn < len(_h) else None, "description_idx": _dc, "description_header": str(_h[_dc]) if _dc is not None and _dc < len(_h) else None}, "timestamp": __import__('time').time() * 1000}
            with open(_lp, "a", encoding="utf-8") as _f:
                _f.write(json.dumps(_data, ensure_ascii=False) + "\n")
        except Exception:
            pass
        # #endregion agent log

        # Ищем все ценовые колонки в заголовках (может быть несколько: банки, кеги и т.д.)
        price_columns = {}  # {индекс_колонки: {'header': заголовок, 'format': формат, 'volume': объем}}
        if df is not None:
            # Получаем заголовки из DataFrame (они могут быть в df.columns или в строке header_row)
            if hasattr(df, 'columns'):
                headers = df.columns.tolist()
            else:
                headers = []
            
            for idx, header in enumerate(headers):
                header_lower = str(header).lower()
                # Ищем колонки с ценой
                if any(word in header_lower for word in ['цена', 'price', 'стоимость', 'cost']):
                    price_col_info = {'header': str(header), 'format': None, 'volume': None}
                    
                    # Определяем формат и объем из заголовка
                    if 'коробка' in header_lower or 'банка' in header_lower or 'can' in header_lower:
                        price_col_info['format'] = 'банка'
                        # Извлекаем объем банки из заголовка, если указан
                        # Пробуем найти объем в скобках типа "(0.33)" или "(0.45)"
                        volume_match = re.search(r'\((\d+[.,]?\d*)\)', header_lower)
                        if volume_match:
                            try:
                                vol_val = float(volume_match.group(1).replace(',', '.'))
                                # Если значение меньше 1, это литры (0.33 = 0.33л, 0.45 = 0.45л)
                                if vol_val < 1:
                                    price_col_info['volume'] = vol_val
                            except:
                                pass
                        # Если объем не нашли в скобках, ищем в тексте типа "0.45л" или "0.45 л"
                        if price_col_info['volume'] is None:
                            volume_match = re.search(r'(\d+[.,]?\d*)\s*л', header_lower, re.IGNORECASE)
                            if volume_match:
                                try:
                                    price_col_info['volume'] = float(volume_match.group(1).replace(',', '.'))
                                except:
                                    pass
                        # Если объем все еще не указан в заголовке, используем стандартный для банок (0.45л)
                        if price_col_info['volume'] is None:
                            price_col_info['volume'] = 0.45
                    elif 'кег' in header_lower or 'keg' in header_lower:
                        price_col_info['format'] = 'кег'
                        # Извлекаем объем кега из заголовка (например "Кег 20л")
                        volume_match = re.search(r'(\d+)\s*л', header_lower, re.IGNORECASE)
                        if volume_match:
                            try:
                                price_col_info['volume'] = float(volume_match.group(1))
                            except:
                                pass
                        # Если объем не указан в заголовке, оставляем None (будет показано как "-")
                        # НЕ устанавливаем стандартный объем 20л, если он не указан явно
                        if price_col_info['volume'] is None:
                            price_col_info['volume'] = None
                    
                    # Добавляем ценовую колонку, если определили формат или есть явная цена в заголовке
                    if price_col_info['format'] or any(word in header_lower for word in ['цена', 'price']):
                        # Если формат не определен, но есть цена - все равно добавляем (может быть общая цена)
                        if not price_col_info['format'] and ('коробка' in header_lower or 'банка' in header_lower or 'can' in header_lower):
                            price_col_info['format'] = 'банка'
                            price_col_info['volume'] = 0.45
                        elif not price_col_info['format'] and ('кег' in header_lower or 'keg' in header_lower):
                            price_col_info['format'] = 'кег'
                            price_col_info['volume'] = 20.0
                        
                        price_columns[idx] = price_col_info
                        logger.debug(f"Обнаружена ценовая колонка: {header} -> формат: {price_col_info['format']}, объем: {price_col_info['volume']}")
        
        # Парсим строки данных
        # Сохраняем последнюю заполненную пивоварню для заполнения пустых значений
        last_brewery = None
        # Для частного поставщика: последняя позиция с названием — чтобы подставить в строки кегов (продолжения)
        last_item_with_name = None
        
        # Определяем, является ли лист розливным (кеги)
        is_draft_sheet = ('розлив' in sheet_name.lower() or 
                         'draft' in sheet_name.lower() or 
                         'кeg' in sheet_name.lower() or
                         'keg' in sheet_name.lower())
        
        # Для пивоварни получаем название по умолчанию
        default_brewery = None
        if supplier_type_enum == SupplierType.BREWERY:
            try:
                brewery_name_to_use = brewery_name or characteristics.get('single_brewery_name')
                if brewery_name_to_use:
                    profile = BreweryProfile(brewery_name_to_use)
                    default_brewery = profile.get_default_brewery_name()
                    logger.debug(f"Установлена пивоварня по умолчанию: {default_brewery}")
            except Exception as e:
                logger.warning(f"Ошибка при установке пивоварни по умолчанию: {str(e)}", exc_info=True)
        
        # Оптимизация: используем выделенный модуль для проверки форматирования
        # Только если openpyxl доступен и файл небольшой (меньше 1000 строк)
        # Не помечаем строки как неактуальные по форматированию, если в таблице явно есть колонки "Наименование" и "Описание" — иначе теряются данные (CBD и др.)
        headers_for_skip = [str(c).lower() for c in (df.columns.tolist() if hasattr(df, 'columns') else [])]
        has_explicit_name_desc = (
            'beer_name' in col_mapping and 'description' in col_mapping
            and any('наименование' in h or 'название' in h for h in headers_for_skip if 'пивоварни' not in h)
            and any('описание' in h for h in headers_for_skip)
        )
        check_formatting = OPENPYXL_AVAILABLE and len(df) < 1000 and not has_explicit_name_desc
        formatting_checker = None
        
        if check_formatting:
            try:
                from .excel.formatting_checker import ExcelFormattingChecker
                formatting_checker = ExcelFormattingChecker(self.file_path)
                logger.debug(f"Инициализирован проверщик форматирования для листа '{sheet_name}'")
            except Exception as e:
                logger.debug(f"Не удалось инициализировать проверщик форматирования: {str(e)}")
                check_formatting = False
        
        # Определяем номер строки заголовка для правильного сопоставления с Excel
        header_row_num = None
        try:
            # Пробуем найти заголовок в первых строках
            header_row_num = self._find_header_row(df.head(20))
            if header_row_num is not None:
                header_row_num += 1  # Преобразуем в 1-based для Excel
        except Exception:
            pass
        
        for idx, row in df.iterrows():
            # Проверяем форматирование только если включено и проверщик инициализирован
            if check_formatting and formatting_checker:
                # Определяем реальный номер строки в Excel
                excel_row_num = None
                try:
                    if header_row_num is not None:
                        excel_row_num = header_row_num + int(idx) + 1
                    else:
                        excel_row_num = int(idx) + 2
                except (ValueError, TypeError):
                    pass
                
                # Проверяем, является ли строка скрытой или неактуальной
                if excel_row_num:
                    if formatting_checker.is_row_hidden_or_inactive(sheet_name, excel_row_num):
                        logger.debug(f"Пропущена строка {excel_row_num} (idx={idx}) - скрыта или помечена как неактуальная")
                        continue
            
            # Продолжаем обычный парсинг...
            try:
                # Инициализируем значения по умолчанию, чтобы избежать UnboundLocalError
                brewery_val_original = ''
                # КРИТИЧЕСКИ ВАЖНО: Сохраняем исходное значение brewery ДО извлечения данных
                # Это нужно для проверки на город после нормализации
                brewery_col_idx = col_mapping.get('brewery')
                brewery_val_original_raw = ''
                if brewery_col_idx is not None:
                    try:
                        raw_value = row.iloc[brewery_col_idx]
                        if pd.notna(raw_value):
                            brewery_val_original_raw = str(raw_value).strip()
                        logger.debug(f"Строка {idx}: brewery_val_original_raw = '{brewery_val_original_raw}'")
                    except (IndexError, KeyError) as e:
                        logger.debug(f"Ошибка при извлечении brewery из строки {idx}: {str(e)}")
                        pass
                
                # Если есть несколько ценовых колонок, создаем отдельный элемент для каждой
                if price_columns and len(price_columns) > 1:
                    # Извлекаем базовые данные один раз
                    base_item = self._extract_row_data(row, col_mapping, df, skip_price=True, default_brewery=default_brewery, supplier_type_enum=supplier_type_enum)
                    if base_item:
                        # Пропускаем строки с "ЕГАИС" вместо реальных данных
                        beer_name_base = base_item.get('beer_name', '').strip().lower() if base_item.get('beer_name') else ''
                        brewery_base = base_item.get('brewery', '').strip().lower() if base_item.get('brewery') else ''
                        if beer_name_base == 'егаис' or brewery_base == 'егаис':
                            logger.info(f"ПРОПУЩЕНА строка {idx} при обработке нескольких ценовых колонок - содержит только 'ЕГАИС'")
                            continue
                        # Для каждой ценовой колонки создаем отдельный элемент
                        for price_col_idx, price_info in price_columns.items():
                            try:
                                if price_col_idx < len(row):
                                    price_value = row.iloc[price_col_idx]
                                    if pd.notna(price_value):
                                        price_str = str(price_value).strip()
                                        # Пропускаем "XX", "хх", "-" и т.д. - это означает, что товара нет
                                        if price_str.upper() in ['XX', 'ХХ', 'Н/Д', 'Н.Д.', '-', '—', '–', '', 'NAN', 'NONE', 'NULL']:
                                            logger.debug(f"Пропущена ценовая колонка {price_col_idx} - значение '{price_str}' означает отсутствие товара")
                                            continue
                                        
                                        # Также проверяем, нет ли "XX" в других важных полях строки
                                        # Если в строке есть "XX" в нескольких местах - возможно, товара вообще нет
                                        row_str = ' '.join([str(v) for v in row.values if pd.notna(v)]).upper()
                                        if row_str.count('XX') > 1 or row_str.count('ХХ') > 1:
                                            logger.debug(f"Пропущена строка {idx} - содержит несколько 'XX', товар отсутствует")
                                            continue
                                        
                                        # Создаем копию базового элемента
                                        item = base_item.copy()
                                        # Сохраняем индекс строки
                                        item['_row_index'] = idx
                                        
                                        # Извлекаем цену из формата "330 / 6600" или "570 / 11400"
                                        if '/' in price_str:
                                            parts = [p.strip() for p in price_str.split('/')]
                                            if len(parts) >= 1:
                                                price_val = parts[0].strip()
                                                price_match = re.search(r'([\d\s.,]+)', price_val)
                                                if price_match:
                                                    item['price'] = normalize_number_str(price_match.group(1))
                                        
                                        # Устанавливаем формат и объем из информации о колонке
                                        item['format_type'] = price_info['format']
                                        
                                        # КРИТИЧЕСКИ ВАЖНО: Для банок и кегов volume устанавливается ТОЛЬКО из заголовка колонки
                                        # Не используем значения из ячеек, чтобы не перепутать с ценой
                                        if price_info['format'] == 'банка':
                                            # Для банок: volume устанавливается ТОЛЬКО из заголовка колонки (price_info['volume'])
                                            if price_info['volume'] is not None:
                                                item['volume'] = price_info['volume']
                                            else:
                                                # Если объем не указан в заголовке, используем стандартный 0.45л
                                                item['volume'] = 0.45
                                            # Помечаем, что volume для банки установлен из заголовка, чтобы не перезаписывать из ячейки
                                            item['_volume_from_header'] = True
                                        elif price_info['format'] == 'кег':
                                            # Для кегов: volume устанавливается ТОЛЬКО из заголовка колонки (price_info['volume'])
                                            if price_info['volume'] is not None:
                                                item['volume'] = price_info['volume']
                                            else:
                                                # Если объем не указан в заголовке, устанавливаем None (будет показано как "-")
                                                item['volume'] = None
                                            # Помечаем, что volume для кега установлен из заголовка, чтобы не перезаписывать из ячейки
                                            item['_volume_from_header'] = True
                                        elif price_info['volume'] is not None:
                                            # Для других форматов: устанавливаем объем из заголовка, если указан
                                            if 'volume' not in item or item.get('volume', 0) < price_info['volume']:
                                                item['volume'] = price_info['volume']
                                        
                                        # Если после всех проверок объема все еще нет - пропускаем элемент
                                        # НО: для кегов разрешаем None (будет показано как "-")
                                        if price_info['format'] != 'кег' and ('volume' not in item or item.get('volume') is None):
                                            logger.debug(f"Пропущен элемент - нет объема для формата {price_info['format']}")
                                            continue
                                        
                                        # Применяем стандартную обработку элемента
                                        # _process_extracted_item может вернуть None, если элемент не валиден (например, нет объема)
                                        processed_item = self._process_extracted_item(
                                            item, brewery_val_original_raw, supplier_type_enum, 
                                            default_brewery, is_draft_sheet, last_brewery, sheet_name
                                        )
                                        if processed_item:
                                            items.append(processed_item)
                                            if processed_item.get('brewery'):
                                                last_brewery = processed_item['brewery']
                                            if supplier_type_enum == SupplierType.BREWERY and (processed_item.get('beer_name') or '').strip():
                                                last_item_with_name = {k: processed_item.get(k) for k in ('beer_name', 'style', 'description', 'abv', 'ibu')}
                                                # Нормализуем описание: убираем название из начала
                                                desc = last_item_with_name.get('description') or ''
                                                beer_name = last_item_with_name.get('beer_name') or ''
                                                if desc and beer_name:
                                                    desc_lower = desc.lower()
                                                    name_lower = beer_name.lower()
                                                    if desc_lower.startswith(name_lower):
                                                        desc_normalized = desc[len(beer_name):].strip()
                                                        desc_normalized = re.sub(r'^[:\-\s]+', '', desc_normalized)
                                                        last_item_with_name['description'] = desc_normalized
                                        else:
                                            logger.debug(f"Элемент пропущен после обработки (строка {idx}, ценовая колонка {price_col_idx})")
                            except Exception as e:
                                logger.debug(f"Ошибка при обработке ценовой колонки {price_col_idx}: {str(e)}")
                                continue
                    else:
                        # Если базовый элемент не валиден, пробуем стандартный путь
                        item = self._extract_row_data(row, col_mapping, df, default_brewery=default_brewery, supplier_type_enum=supplier_type_enum)
                        if item:
                            processed_item = self._process_extracted_item(
                                item, brewery_val_original_raw, supplier_type_enum, 
                                default_brewery, is_draft_sheet, last_brewery, sheet_name
                            )
                            if processed_item:
                                items.append(processed_item)
                                if processed_item.get('brewery'):
                                    last_brewery = processed_item['brewery']
                                if supplier_type_enum == SupplierType.BREWERY and (processed_item.get('beer_name') or '').strip():
                                    last_item_with_name = {k: processed_item.get(k) for k in ('beer_name', 'style', 'description', 'abv', 'ibu')}
                                    # Нормализуем описание: убираем название из начала
                                    desc = last_item_with_name.get('description') or ''
                                    beer_name = last_item_with_name.get('beer_name') or ''
                                    if desc and beer_name:
                                        desc_lower = desc.lower()
                                        name_lower = beer_name.lower()
                                        if desc_lower.startswith(name_lower):
                                            desc_normalized = desc[len(beer_name):].strip()
                                            desc_normalized = re.sub(r'^[:\-\s]+', '', desc_normalized)
                                            last_item_with_name['description'] = desc_normalized
                else:
                    # Стандартная обработка (одна ценовая колонка)
                    item = self._extract_row_data(row, col_mapping, df, default_brewery=default_brewery, supplier_type_enum=supplier_type_enum)
                    if item:
                        # Сохраняем индекс строки для корректного экспорта заказа
                        item['_row_index'] = idx
                        # Для частного поставщика: строка кега без названия — копируем название/описание/стиль из предыдущей строки
                        if supplier_type_enum == SupplierType.BREWERY and last_item_with_name:
                            has_name = bool((item.get('beer_name') or '').strip())
                            has_format_and_price = bool((item.get('format_type') or '').strip()) and bool(item.get('price'))
                            if not has_name and has_format_and_price:
                                for key in ('beer_name', 'style', 'abv', 'ibu'):
                                    if not (item.get(key) or '').strip() and last_item_with_name.get(key):
                                        item[key] = last_item_with_name.get(key) or ''
                                # Для описания: копируем и нормализуем (убираем название из начала, если оно там есть)
                                if not (item.get('description') or '').strip() and last_item_with_name.get('description'):
                                    desc = last_item_with_name.get('description') or ''
                                    beer_name_from_last = (last_item_with_name.get('beer_name') or '').strip()
                                    if beer_name_from_last:
                                        desc_lower = desc.lower()
                                        name_lower = beer_name_from_last.lower()
                                        # Убираем название из начала описания
                                        if desc_lower.startswith(name_lower):
                                            desc = desc[len(beer_name_from_last):].strip()
                                            desc = re.sub(r'^[:\-\s]+', '', desc)
                                    item['description'] = desc
                        # КРИТИЧЕСКИ ВАЖНО: Проверяем валидность элемента ПЕРЕД обработкой
                        # Используем ИСХОДНОЕ значение brewery из строки DataFrame (до нормализации)
                        brewery_val_original = brewery_val_original_raw if brewery_val_original_raw else (item.get('brewery_original', '').strip() if item.get('brewery_original') else '')
                        # Если все еще пустое, используем текущее значение brewery (но это уже нормализованное)
                        if not brewery_val_original:
                            brewery_val_original = item.get('brewery', '').strip() if item.get('brewery') else ''
                    
                        beer_name_val_check = item.get('beer_name', '').strip() if item.get('beer_name') else ''
                        style_val_check = item.get('style', '').strip() if item.get('style') else ''
                        description_val_check = item.get('description', '').strip() if item.get('description') else ''
                        format_type_val_check = item.get('format_type', '').strip() if item.get('format_type') else ''
                        
                        # Функция проверки пустых значений
                        def is_empty_or_dash(val):
                            """Проверяет, является ли значение пустым или только тире"""
                            if val is None:
                                return True
                            if not val:
                                return True
                            val_str = str(val).strip()
                            if not val_str:
                                return True
                            val_str_lower = val_str.lower()
                            # Проверяем различные варианты тире и пустых значений
                            return val_str_lower in ['-', '—', '–', '', 'nan', 'none', 'null', 'n/a', 'na']
                        
                        # КРИТИЧЕСКАЯ ПРОВЕРКА: пропускаем строки с "ЕГАИС" в beer_name или brewery
                        # "ЕГАИС" - это служебный код, а не название пивоварни или пива
                        beer_name_lower = beer_name_val_check.lower() if beer_name_val_check else ''
                        brewery_val_check = item.get('brewery', '').strip().lower() if item.get('brewery') else ''
                        
                        # Проверяем, не является ли название ЕГАИС кодом (например, "Мантра крепкий Эль-9", "Эль-8")
                        # Паттерны ЕГАИС кодов: "крепкий Эль-X", "Эль-X", где X - число
                        egais_patterns = [
                            r'крепкий\s+эль[-\s]\d+',  # "крепкий Эль-9"
                            r'эль[-\s]\d+',  # "Эль-9", "Эль-8"
                            r'эль\s*\d+',  # "Эль9", "Эль 9"
                            r'ale[-\s]\d+',  # "Ale-9"
                        ]
                        is_egais_code = False
                        if beer_name_val_check:
                            for pattern in egais_patterns:
                                if re.search(pattern, beer_name_lower, re.IGNORECASE):
                                    is_egais_code = True
                                    logger.debug(f"Обнаружен ЕГАИС код в названии: '{beer_name_val_check}' (строка {idx})")
                                    break
                        
                        # Если beer_name или brewery содержит только "ЕГАИС" - пропускаем строку
                        if beer_name_lower == 'егаис' or brewery_val_check == 'егаис':
                            logger.info(f"ПРОПУЩЕНА строка {idx} - содержит только 'ЕГАИС' (это служебный код, а не данные о пиве)")
                            continue
                        
                        # Если beer_name - это ЕГАИС код (паттерн типа "Эль-X"), а нет описания с реальным названием - пропускаем
                        if is_egais_code:
                            # Проверяем, есть ли в описании реальное название пива (не ЕГАИС код)
                            description_text = item.get('description', '').strip().lower() if item.get('description') else ''
                            # Если в описании есть двоеточие и английские слова - возможно, это реальное название
                            has_real_name_in_desc = ':' in str(item.get('description', '')) and re.search(r'[a-zA-Z]{3,}', str(item.get('description', '')))
                            if not has_real_name_in_desc:
                                logger.info(f"ПРОПУЩЕНА строка {idx} - название '{beer_name_val_check}' является ЕГАИС кодом, нет реального названия")
                                continue
                        
                        # Если beer_name или brewery начинается с "ЕГАИС" и нет других данных - тоже пропускаем
                        if (beer_name_lower.startswith('егаис') or brewery_val_check.startswith('егаис')) and is_empty_or_dash(description_val_check) and is_empty_or_dash(style_val_check):
                            # Но если есть описание или стиль - возможно это реальные данные
                            logger.info(f"ПРОПУЩЕНА строка {idx} - начинается с 'ЕГАИС' и нет других данных")
                            continue
                    
                    # КРИТИЧЕСКАЯ ПРОВЕРКА: если brewery содержит город и все ключевые поля пустые - пропускаем
                    if brewery_val_original:
                        brewery_lower = str(brewery_val_original).lower()
                        # Расширенный список паттернов для городов
                        city_patterns = [
                            'г.', 'г ', 'город', 'city',
                            'владимир', 'москва', 'санкт-петербург', 'spb', 'мск',
                            'санкт-петербург', 'петербург', 'питер', 'spb',
                            'санкт петербург', 'санктпетербург'
                        ]
                        has_city_in_brewery = any(pattern in brewery_lower for pattern in city_patterns)
                        
                        if has_city_in_brewery:
                            is_empty_beer_name = is_empty_or_dash(beer_name_val_check)
                            is_empty_style = is_empty_or_dash(style_val_check)
                            is_empty_description = is_empty_or_dash(description_val_check)
                            is_empty_format = is_empty_or_dash(format_type_val_check)
                            
                            logger.debug(f"Строка {idx}: brewery='{brewery_val_original}', has_city={has_city_in_brewery}, "
                                       f"beer_name_empty={is_empty_beer_name}, style_empty={is_empty_style}, "
                                       f"description_empty={is_empty_description}, format_empty={is_empty_format}")
                            
                            # Если все эти поля пустые - точно пропускаем
                            if is_empty_beer_name and is_empty_style and is_empty_description and is_empty_format:
                                logger.info(f"ПРОПУЩЕНА строка {idx} с brewery '{brewery_val_original}' - содержит город и все поля (beer_name, style, description, format_type) пустые")
                                continue
                            
                            # Для позиций в "Прочее" ослабляем проверку - если есть хотя бы style или description, не пропускаем
                            # Если хотя бы beer_name пустой И нет других данных - пропускаем
                            if is_empty_beer_name and is_empty_style and is_empty_description:
                                logger.info(f"ПРОПУЩЕНА строка {idx} с brewery '{brewery_val_original}' - содержит город и нет beer_name, style, description")
                                continue
                    
                        # Обработка пивоварни в зависимости от типа поставщика
                        if supplier_type_enum == SupplierType.BREWERY:
                            # Для частного поставщика ВСЕГДА устанавливаем название пивоварни
                            # (даже если в данных есть brewery - заменяем на переданное)
                            if default_brewery:
                                # Нормализуем brewery (удаляем город) перед установкой
                                from parser_app.domain.services.normalization import DataNormalizer
                                normalizer = DataNormalizer()
                                brewery_normalized = normalizer.normalize_brewery(default_brewery)
                                item['brewery'] = brewery_normalized if brewery_normalized else default_brewery
                                logger.debug(f"Установлена пивоварня для частного поставщика: {item['brewery']}")
                        else:
                            # Для дистрибьютора используем логику с предыдущей пивоварней
                            if not item:
                                logger.debug(f"Пропуск строки {idx} - пустой item после извлечения")
                                continue
                            if not item.get('brewery') and last_brewery:
                                # Нормализуем brewery (удаляем город) перед установкой
                                from parser_app.domain.services.normalization import DataNormalizer
                                normalizer = DataNormalizer()
                                brewery_normalized = normalizer.normalize_brewery(last_brewery)
                                item['brewery'] = brewery_normalized if brewery_normalized else last_brewery
                            # Сохраняем текущую пивоварню, если она заполнена
                            elif item.get('brewery'):
                                # Нормализуем brewery перед сохранением в last_brewery
                                from parser_app.domain.services.normalization import DataNormalizer
                                normalizer = DataNormalizer()
                                brewery_normalized = normalizer.normalize_brewery(item['brewery'])
                                last_brewery = brewery_normalized if brewery_normalized else item['brewery']
                    
                        # Нормализуем brewery (удаляем город) сразу после установки
                        # Это критически важно для удаления городов из названий пивоварен
                        if item.get('brewery'):
                            from parser_app.normalizers import DataNormalizer
                            normalizer = DataNormalizer()
                            normalized_brewery = normalizer.normalize_brewery(item['brewery'])
                            if normalized_brewery:
                                item['brewery'] = normalized_brewery
                            else:
                                # Если после нормализации brewery стал пустым, оставляем исходное значение
                                # но логируем это для отладки
                                logger.debug(f"Brewery стал пустым после нормализации: '{item['brewery']}'")
                    
                        # Применяем стандартную обработку элемента
                        processed_item = self._process_extracted_item(
                            item, brewery_val_original, supplier_type_enum, 
                            default_brewery, is_draft_sheet, last_brewery, sheet_name
                        )
                        if processed_item:
                            items.append(processed_item)
                            if processed_item.get('brewery'):
                                last_brewery = processed_item['brewery']
                            # Запоминаем последнюю позицию с названием для подстановки в следующие строки (кеги)
                            if supplier_type_enum == SupplierType.BREWERY and (processed_item.get('beer_name') or '').strip():
                                last_item_with_name = {k: processed_item.get(k) for k in ('beer_name', 'style', 'description', 'abv', 'ibu')}
                                # Нормализуем описание: убираем название из начала, если оно там есть
                                desc = last_item_with_name.get('description') or ''
                                beer_name = last_item_with_name.get('beer_name') or ''
                                if desc and beer_name:
                                    desc_lower = desc.lower()
                                    name_lower = beer_name.lower()
                                    if desc_lower.startswith(name_lower):
                                        desc_normalized = desc[len(beer_name):].strip()
                                        desc_normalized = re.sub(r'^[:\-\s]+', '', desc_normalized)
                                        last_item_with_name['description'] = desc_normalized
            except Exception as e:
                logger.error(f"Ошибка при обработке строки {idx} в листе {sheet_name}: {str(e)}", exc_info=True)
                continue
        
        # Закрываем проверщик форматирования если был инициализирован
        if formatting_checker:
            try:
                formatting_checker.close()
                logger.debug(f"Закрыт проверщик форматирования для листа '{sheet_name}'")
            except Exception:
                pass
        
        # Дедупликация: убираем дубликаты (одинаковые brewery, beer_name, style, format_type, price)
        # Применяем для всех типов поставщиков
        # Используем выделенный модуль для дедупликации
        from ....domain.services.deduplication import Deduplicator
        return Deduplicator.deduplicate(items)
    
    def _process_extracted_item(self, item: Dict, brewery_val_original: str, 
                                supplier_type_enum: SupplierType, default_brewery: Optional[str],
                                is_draft_sheet: bool, last_brewery: Optional[str], 
                                sheet_name: str) -> Optional[Dict]:
        """
        Обрабатывает извлеченный элемент: нормализует brewery, устанавливает формат и т.д.
        
        Args:
            item: Словарь с данными позиции
            brewery_val_original: Исходное значение brewery (до нормализации)
            supplier_type_enum: Тип поставщика
            default_brewery: Название пивоварни по умолчанию (для частного поставщика)
            is_draft_sheet: Является ли лист розливным
            last_brewery: Последняя заполненная пивоварня (для дистрибьютора)
            sheet_name: Имя листа
            
        Returns:
            Обработанный элемент или None, если элемент не валиден
        """
        # Обработка пивоварни в зависимости от типа поставщика
        if supplier_type_enum == SupplierType.BREWERY:
            # Для частного поставщика ВСЕГДА устанавливаем название пивоварни
            if default_brewery:
                from parser_app.normalizers import DataNormalizer
                normalizer = DataNormalizer()
                brewery_normalized = normalizer.normalize_brewery(default_brewery)
                item['brewery'] = brewery_normalized if brewery_normalized else default_brewery
                logger.debug(f"Установлена пивоварня для частного поставщика: {item['brewery']}")
        else:
            # Для дистрибьютора используем логику с предыдущей пивоварней
            if not item.get('brewery') and last_brewery:
                from parser_app.normalizers import DataNormalizer
                normalizer = DataNormalizer()
                brewery_normalized = normalizer.normalize_brewery(last_brewery)
                item['brewery'] = brewery_normalized if brewery_normalized else last_brewery
        
        # Нормализуем brewery (удаляем город) сразу после установки
        if item.get('brewery'):
            from parser_app.normalizers import DataNormalizer
            normalizer = DataNormalizer()
            normalized_brewery = normalizer.normalize_brewery(item['brewery'])
            if normalized_brewery:
                item['brewery'] = normalized_brewery
        
        # Для розливных листов автоматически устанавливаем формат "кега"
        if is_draft_sheet:
            current_format = item.get('format_type', '').lower().strip()
            if not current_format or current_format not in ['кега', 'keg', 'кег']:
                item['format_type'] = 'кега'
                logger.debug(f"Установлен формат 'кега' для разливного пива (лист: {sheet_name})")
        
        # Пытаемся извлечь формат и объем из названия или описания, если они не найдены в колонках
        if not item.get('format_type') or not item.get('format_type').strip():
            # Ищем формат в названии
            beer_name_text = (item.get('beer_name') or '').lower()
            description_text = (item.get('description') or '').lower()
            combined_text = f"{beer_name_text} {description_text}"
            
            # Ищем упоминания формата в тексте
            if any(x in combined_text for x in ('банка', 'банки', 'can', 'cans', 'бутылка', 'bottle')):
                if 'банка' in combined_text or 'can' in combined_text:
                    item['format_type'] = 'банка'
                elif 'бутылка' in combined_text or 'bottle' in combined_text:
                    item['format_type'] = 'бутылка'
                logger.debug(f"Извлечен формат из названия/описания: {item['format_type']}")
            
            # Ищем объем в названии или описании (например, "0.5л", "0.45л", "20л")
            volume_match = re.search(r'(\d+[.,]?\d*)\s*л', combined_text, re.IGNORECASE)
            if volume_match:
                try:
                    volume_val = float(volume_match.group(1).replace(',', '.'))
                    if volume_val > 0:
                        item['volume'] = volume_val
                        logger.debug(f"Извлечен объем из названия/описания: {volume_val}л")
                except (ValueError, TypeError):
                    pass
        
        # Устанавливаем объем по умолчанию на основе формата (если возможно)
        # Если объем отсутствует и формат не определен - элемент попадет в "Прочее"
        if 'volume' not in item or item.get('volume') is None:
            format_lower = str(item.get('format_type', '')).lower().strip()
            if 'банка' in format_lower or 'can' in format_lower:
                item['volume'] = 0.45  # Стандартный объем банки
                logger.debug(f"Установлен объем по умолчанию 0.45л для банки")
            elif 'кег' in format_lower or 'keg' in format_lower:
                # Для кегов: если объем не указан, оставляем None (будет показано как "-")
                item['volume'] = None
                logger.debug(f"Объем кега не указан, устанавливаем None (будет показано как '-')")
            # Если формат не определен - оставляем volume=None, элемент попадет в "Прочее"
        
        # Добавляем информацию об источнике данных (если есть idx в item)
        if 'raw_source_location' not in item:
            item['raw_source_location'] = {
                'sheet': sheet_name,
                'row': item.get('_row_index', 0) + 1
            }
        
        # Для частных поставщиков (Paradox и др.): вкладки по формату — Банки / Кеги / Прочее
        if supplier_type_enum == SupplierType.BREWERY:
            format_lower = (item.get('format_type') or '').lower().strip()
            has_format = bool(format_lower)
            has_volume = item.get('volume') is not None
            
            # Банки: банка, бутылка, can, коробка
            if any(x in format_lower for x in ('банка', 'банки', 'бутылка', 'bottle', 'can', 'cans', 'коробка')):
                item['raw_source_location']['original_sheet'] = sheet_name
                item['raw_source_location']['sheet'] = 'Банки'
            # Кеги: кег, кега, keg, разлив, "20 л" / "30 л" в формате (типичный объём кега)
            elif any(x in format_lower for x in ('кег', 'кеги', 'кега', 'keg', 'kegs', 'разлив')):
                item['raw_source_location']['original_sheet'] = sheet_name
                item['raw_source_location']['sheet'] = 'Кеги'
            elif re.search(r'\d+\s*л', format_lower) and not any(x in format_lower for x in ('банка', 'бутылка', 'can', 'шт')):
                # "20 л", "30 л" без банки/бутылки — обычно кег
                item['raw_source_location']['original_sheet'] = sheet_name
                item['raw_source_location']['sheet'] = 'Кеги'
            else:
                # Прочее: позиции без формата, без объема, или позиции пива/безалкогольных напитков
                # которые не являются банками или кегами
                item['raw_source_location']['original_sheet'] = sheet_name
                item['raw_source_location']['sheet'] = 'Прочее'
                logger.debug(f"Позиция попадает в 'Прочее': формат='{format_lower}', объем={item.get('volume')}, название='{item.get('beer_name', '')[:50]}'")
        
        return item
    
    def _find_header_row(self, df: pd.DataFrame) -> Optional[int]:
        """
        Ищет строку с заголовками в DataFrame.
        
        Args:
            df: DataFrame для поиска
            
        Returns:
            Индекс строки с заголовками или None
        """
        # Паттерны из конфига (расширяемо без правок кода)
        header_patterns = get_header_patterns()
        if not header_patterns:
            header_patterns = [
                'пивоварня', 'название', 'цена', 'price', 'стоимость', 'наименование',
                'стиль', 'объем', 'фасовка', 'формат', 'остаток', 'наличие'
            ]
        
        best_match = None
        best_score = 0
        
        # Поиск до 25 строк — сложные прайсы могут иметь заголовок на 5–10 строке
        for idx in range(min(25, len(df))):
            try:
                row_values = [str(val).lower().strip() if pd.notna(val) else '' 
                             for val in df.iloc[idx].tolist()[:20]]  # Только первые 20 колонок
                
                # Пропускаем строки-шапки прайса (типа "Прайс обновлён...", "Минимальный заказ...")
                first_cell = row_values[0] if row_values else ''
                if first_cell and len(first_cell) > 80 and any(
                    k in first_cell for k in ('прайс обновлён', 'минимальный заказ', 'развоз по будням', 'канал cbd')
                ):
                    continue
                
                # Считаем количество совпадений с паттернами
                matches = 0
                matched_patterns = set()
                for pattern in header_patterns:
                    for val in row_values:
                        if pattern in val and len(val) > 2:
                            if pattern not in matched_patterns:
                                matches += 1
                                matched_patterns.add(pattern)
                                break
                
                # Также проверяем, что в строке нет слишком много пустых значений
                non_empty = sum(1 for v in row_values if v and len(v) > 2)
                
                # Специальные проверки для точного определения заголовков (дистрибьюторы часто используют "Бренд"/"Марка")
                has_brewery_header = any(
                    'наименование пивоварни' in v or
                    ('наименование' in v and 'пивоварни' in v) or
                    'пивоварня' in v or 'brewery' in v or
                    'бренд' in v or 'brand' in v or 'марка' in v or 'производитель' in v
                    for v in row_values
                )
                has_product_header = any(
                    ('наименование' in v and 'пивоварни' not in v) or
                    'название' in v or 'beer' in v or 'name' in v
                    for v in row_values
                )
                has_price_header = any(
                    'цена' in v or 'price' in v or 'стоимость' in v or 'руб' in v
                    for v in row_values
                )
                
                # Если найдены все три ключевых заголовка - это точно заголовки
                if has_brewery_header and has_product_header and has_price_header:
                    logger.debug(f"Найдена строка заголовков (точное совпадение) на строке {idx + 1}")
                    return idx
                
                # Для частных поставщиков: заголовок может быть без колонки "пивоварня"
                # Достаточно название/наименование + цена
                if has_product_header and has_price_header:
                    logger.debug(f"Найдена строка заголовков (название+цена) на строке {idx + 1}")
                    return idx
                
                # Проверяем комбинации из двух ключевых заголовков
                if (has_brewery_header and has_price_header) or \
                   (has_product_header and has_price_header):
                    score = matches * 3 + (non_empty if non_empty < 20 else 0)
                    if score > best_score:
                        best_score = score
                        best_match = idx
                
                # Считаем общий скор: количество совпадений + количество непустых значений
                score = matches * 2 + (non_empty if non_empty < 20 else 0)
                
                if score > best_score and matches >= 2:
                    best_score = score
                    best_match = idx
            except Exception as e:
                logger.debug(f"Ошибка при проверке строки {idx}: {str(e)}")
                continue
        
        if best_match is not None:
            logger.debug(f"Найдена строка заголовков (лучшее совпадение) на строке {best_match + 1}, скор: {best_score}")
        
        return best_match
    
    def _map_columns(self, headers: List[str], df: pd.DataFrame = None) -> Dict[str, int]:
        """
        Определяет соответствие колонок заголовкам.
        
        Args:
            headers: Список заголовков колонок
            
        Returns:
            Словарь {название_поля: индекс_колонки}
        """
        mapping = {}
        
        # Синонимы из конфига (column_patterns.json) — можно дополнять без правок кода
        field_patterns = get_field_patterns()
        if not field_patterns:
            field_patterns = {
                'brewery': ['пивоварня', 'brewery', 'производитель'],
                'beer_name': ['название', 'наименование', 'beer', 'name'],
                'price': ['цена', 'price', 'стоимость', 'руб'],
                'style': ['стиль', 'style'], 'volume': ['объем', 'volume', 'л'],
                'format_type': ['формат', 'упаковка', 'фасовка'],
                'abv': ['abv', 'крепость'], 'ibu': ['ibu', 'горечь'],
                'stock': ['остаток', 'наличие'], 'description': ['описание'],
                'currency': ['валюта'], 'supplier_name': ['поставщик'],
            }
        
        headers_lower = [str(h).lower().strip() if h else '' 
                        for h in headers]
        
        # КРИТИЧЕСКИ ВАЖНО: Сначала проверяем колонку "Название" (приоритет 1)
        # Это должно быть ДО всех других проверок, чтобы гарантировать правильный маппинг
        for idx, header in enumerate(headers_lower):
            if header and header.strip() == 'название':
                mapping['beer_name'] = idx
                logger.info(f"ПРИОРИТЕТ: Колонка '{header}' (индекс {idx}) устанавливается как beer_name")
                break
        
        # Специальная обработка для русских заголовков
        for idx, header in enumerate(headers_lower):
            if header:
                # Игнорируем колонку "ЕГАИС" (это код продукта, а не brewery или beer_name)
                if header in ['егаис', 'егаис код', 'код егаис', 'код товара', 'артикул']:
                    continue
                
                # Если beer_name уже маппирован (колонка "Название"), пропускаем дальнейшие проверки для beer_name
                if 'beer_name' in mapping and idx != mapping['beer_name']:
                    # Продолжаем только для других полей
                    pass
                
                # "Наименование пивоварни" = brewery
                if 'наименование пивоварни' in header or ('наименование' in header and 'пивоварни' in header):
                    if 'brewery' not in mapping:
                        mapping['brewery'] = idx
                        continue
                
                # "Название" = beer_name (если еще не маппировано)
                if header.strip() == 'название' or header.strip() == 'название товара':
                    if 'beer_name' not in mapping:
                        mapping['beer_name'] = idx
                        logger.debug(f"Колонка '{header}' (индекс {idx}) маппится как beer_name")
                        continue
                
                # "Наименование" (без пивоварни) = beer_name
                elif 'наименование' in header and 'пивоварни' not in header and 'егаис' not in header:
                    if 'beer_name' not in mapping:
                        mapping['beer_name'] = idx
                        continue
                # Пивоварня - общие паттерны (у дистров часто колонки "Бренд", "Марка")
                elif any(p in header for p in ['пивоварня', 'brewery', 'производитель', 'manufacturer', 'бренд', 'brand', 'марка']):
                    if 'brewery' not in mapping:
                        mapping['brewery'] = idx
                        continue
                # Название пива (игнорируем ЕГАИС и ЕГАИС коды)
                # ВАЖНО: Эта проверка только если beer_name еще не маппирован (колонка "Название" уже обработана выше)
                elif 'beer_name' not in mapping and any(p in header for p in ['название', 'beer', 'name']) and 'егаис' not in header:
                    # Пропускаем, если в заголовке есть слова, указывающие на ЕГАИС коды
                    if 'крепость' not in header and 'эль' not in header.lower() and 'код' not in header:
                        mapping['beer_name'] = idx
                        logger.debug(f"Колонка '{header}' (индекс {idx}) маппится как beer_name (fallback)")
                        continue
                # Тип фасовки / формат (в т.ч. "Тип фасовки / кол-во в уп" как в CBD прайсе)
                if 'тип фасовки' in header or ('фасовки' in header and 'кол-во' in header) or 'кол-во в уп' in header:
                    if 'format_type' not in mapping:
                        mapping['format_type'] = idx
                        continue
                # Объем / литраж - приоритетная обработка перед стилем
                # Проверяем явно "литраж" первой
                if 'литраж' in header:
                    if 'volume' not in mapping:
                        mapping['volume'] = idx
                        continue
                # Проверяем другие паттерны для объема, но только если это не "Стиль"
                if any(p in header for p in ['объём', 'volume', 'литр', 'литров']) and 'стиль' not in header:
                    if 'volume' not in mapping:
                        mapping['volume'] = idx
                        continue
                # Цена (в т.ч. "Цена за литр" на листе Розлив)
                if any(p in header for p in ['цена', 'price', 'стоимость', 'cost', 'цена за литр']):
                    if 'price' not in mapping:
                        mapping['price'] = idx
                        continue
                # Стиль - только если это не литраж
                if ('стиль' in header or 'style' in header) and 'литраж' not in header:
                    if 'style' not in mapping:
                        mapping['style'] = idx
                        continue
                # Описание
                if 'описание' in header or 'description' in header:
                    if 'description' not in mapping:
                        mapping['description'] = idx
                        continue
                # ABV / IBU (может быть в одной колонке типа "ABV / OG / IBU")
                if 'abv' in header or 'крепость' in header:
                    if 'abv' not in mapping:
                        mapping['abv'] = idx
                if 'ibu' in header or 'горечь' in header:
                    if 'ibu' not in mapping:
                        mapping['ibu'] = idx
                        continue
        
        # Если передан DataFrame, анализируем содержимое колонок для более точного определения
        if df is not None and not df.empty:
            sample_rows = df.head(min(10, len(df)))
            egais_pattern = re.compile(r'эль[-\s]?\d+|крепкий\s+эль[-\s]?\d+|берлинер\s+вайссе\s+\d+|гозэ\s+\d+', re.IGNORECASE)
            
            # Сначала проверяем все колонки на реальные названия и ЕГАИС коды
            real_names_candidates = []  # Список колонок с реальными названиями
            
            for idx, header in enumerate(headers_lower):
                if idx >= len(df.columns):
                    continue
                
                try:
                    col_data = sample_rows.iloc[:, idx].dropna().astype(str).tolist()
                    if not col_data:
                        continue
                    
                    # Проверяем на ЕГАИС коды
                    egais_count = sum(1 for val in col_data[:5] if egais_pattern.search(str(val).lower()))
                    is_egais_col = egais_count >= len(col_data[:5]) * 0.6
                    
                    # Проверяем на реальные названия пива
                    # Реальные названия содержат английские слова с заглавными буквами, двоеточия и т.д.
                    # Например: "Master of Pastries: Dubai Chocolate", "Eternal Nightmare"
                    real_name_patterns = [':', ' - ', '|', ' & ', ' of ', ' with ', 'Master', 'Chocolate', 'Pastries', 'Dubai', 'Eternal', 'Nightmare', 'Magic', 'Messiah', 'Croissant', 'Adept', 'Fracture', 'Lemonade', 'Shake', 'Bus']
                    # Паттерн для английских названий: заглавная буква, слово, пробел, заглавная буква, слово
                    # Например: "Master of Pastries", "Eternal Nightmare"
                    has_english_words = any(re.search(r'[A-Z][a-z]+\s+[A-Z][a-z]+', str(val)) for val in col_data[:5])
                    # Проверяем наличие паттернов из реальных названий
                    has_real_name_patterns = any(any(p in str(val) for p in real_name_patterns) for val in col_data[:5])
                    # Также проверяем, что это не короткие ЕГАИС коды (обычно 1-2 слова с числами)
                    has_multiple_words = any(len(str(val).split()) >= 2 for val in col_data[:5])
                    # Проверяем наличие двоеточий (характерно для реальных названий)
                    has_colon = any(':' in str(val) for val in col_data[:5])
                    
                    # Колонка "Описание" никогда не должна маппиться на beer_name — только на description
                    if 'описание' in header or header.strip() == 'description':
                        continue
                    # Колонка "Стиль" — только стиль, не название пива
                    if mapping.get('style') == idx or 'стиль' in header or header.strip() == 'style':
                        continue
                    # Значения выглядят как стили (IPA, Ale, Altbier...), а не названия — не подставлять в beer_name
                    if self._looks_like_style_only_column(col_data):
                        continue
                    # Длинный текст в колонке — это описание, не название (названия обычно до ~80 символов)
                    avg_len = sum(len(str(v)) for v in col_data[:5]) / min(5, len(col_data)) if col_data else 0
                    if avg_len > 80:
                        continue
                    # Заголовок длиннее 50 символов — скорее всего это значение ячейки, а не заголовок колонки
                    if len(header) > 50:
                        continue
                    # Если это колонка с реальными названиями, добавляем в кандидаты
                    is_real_name_col = (has_english_words or has_real_name_patterns or has_colon) and not is_egais_col
                    if is_real_name_col:
                        # Приоритет: "Наименование"/"Название" = 25 (явная колонка названия), с двоеточием = 15, остальные = 10
                        priority = 25 if ('наименование' in header or 'название' in header) and 'пивоварни' not in header else (15 if has_colon else 10)
                        real_names_candidates.append((idx, header, priority))
                        logger.info(f"Колонка {idx} '{header}' - кандидат на beer_name (приоритет {priority}, англ. слова: {has_english_words}, паттерны: {has_real_name_patterns}, двоеточие: {has_colon})")
                    
                    # Если текущий beer_name маппится на колонку с ЕГАИС кодами - удаляем маппинг
                    if is_egais_col and mapping.get('beer_name') == idx:
                        logger.info(f"Колонка {idx} '{header}' содержит ЕГАИС коды ({egais_count}/{len(col_data[:5])}), удаляем маппинг beer_name")
                        del mapping['beer_name']
                        
                except Exception as e:
                    logger.debug(f"Ошибка при проверке колонки {idx}: {str(e)}")
                    continue
            
            # Если нашли колонки с реальными названиями, выбираем лучшую
            if real_names_candidates:
                # Исключаем колонку "Описание" — она не должна быть beer_name
                real_names_candidates = [(i, h, p) for i, h, p in real_names_candidates if 'описание' not in h and h.strip() != 'description']
                if real_names_candidates:
                    real_names_candidates.sort(key=lambda x: x[2], reverse=True)
                    best_candidate = real_names_candidates[0]
                    best_idx, best_header, _ = best_candidate
                    current_beer_name_idx = mapping.get('beer_name')
                    current_header = headers_lower[current_beer_name_idx] if current_beer_name_idx is not None and current_beer_name_idx < len(headers_lower) else ''
                    # Нужно название пива, не стиль: колонка "Стиль" не должна быть beer_name
                    if current_header and ('стиль' in current_header or current_header.strip() == 'style'):
                        mapping['beer_name'] = best_idx
                        logger.info(f"Колонка beer_name заменена с '{current_header}' (стиль) на '{best_header}' — в названии должно быть название пива")
                    # Не перезаписываем beer_name, если уже маппирован на явную колонку "Название"/"Наименование" (не "Наименование пивоварни")
                    elif current_header and any(k in current_header for k in ('название', 'наименование', 'назва')) and 'пивоварни' not in current_header:
                        logger.info(f"Оставляем beer_name на колонке {current_beer_name_idx} '{current_header}' (явная колонка названия)")
                    elif current_beer_name_idx != best_idx:
                        logger.info(f"Устанавливаем колонку {best_idx} '{best_header}' как beer_name (вместо индекса {current_beer_name_idx})")
                        mapping['beer_name'] = best_idx
            
            # Продолжаем анализ остальных колонок для других полей
            for idx, header in enumerate(headers_lower):
                if idx >= len(df.columns):
                    continue
                
                try:
                    col_data = sample_rows.iloc[:, idx].dropna().astype(str).tolist()
                    if not col_data:
                        continue
                    
                    # Анализируем содержимое колонки
                    col_analysis = self._analyze_column_content(col_data, idx)
                    
                    # Применяем результаты анализа, если поле еще не определено
                    # Колонка "Описание" никогда не должна маппиться на beer_name
                    is_description_header = 'описание' in header or header.strip() == 'description'
                    for field_name, should_map in col_analysis.items():
                        if should_map and field_name not in mapping:
                            if field_name == 'beer_name' and is_description_header:
                                continue
                            mapping[field_name] = idx
                            logger.debug(f"Определено поле {field_name} по содержимому колонки {idx}")
                            break
                except Exception as e:
                    logger.debug(f"Ошибка при анализе колонки {idx}: {str(e)}")
                    continue
        
        # Затем ищем по общим паттернам для остальных полей
        for field, patterns in field_patterns.items():
            if field not in mapping:  # Пропускаем уже найденные поля
                for idx, header in enumerate(headers_lower):
                    if header and any(pattern in header for pattern in patterns):
                        # Для volume пропускаем, если это "Стиль" или другие не-объемные колонки
                        if field == 'volume':
                            # Не маппим volume к колонкам со стилем
                            if 'стиль' in header or 'style' in header:
                                continue
                        mapping[field] = idx
                        break
        
        # Дополнительная проверка: если есть колонка с похожим названием
        # Например, "Название товара" или "Наименование"
        for idx, header in enumerate(headers_lower):
            if header:
                # Колонка "Описание" никогда не должна маппиться на beer_name
                if 'описание' in header or header.strip() == 'description':
                    continue
                # Проверяем более широкие паттерны (игнорируем ЕГАИС)
                if ('назв' in header or 'наимен' in header or 'товар' in header) and 'beer_name' not in mapping:
                    # Но только если это не пивоварня и не ЕГАИС
                    if 'пивоварни' not in header and 'егаис' not in header:
                        mapping['beer_name'] = idx
                elif ('произв' in header or 'бренд' in header or 'марка' in header) and 'brewery' not in mapping:
                    mapping['brewery'] = idx
                elif ('цена' in header or 'стоим' in header or 'руб' in header) and 'price' not in mapping:
                    mapping['price'] = idx
                elif ('фасов' in header or 'формат' in header or 'упаковка' in header) and 'format_type' not in mapping:
                    mapping['format_type'] = idx
        
        # КРИТИЧНО: название и описание не должны указывать на одну колонку — иначе описание копируется в название
        if mapping.get('beer_name') is not None and mapping.get('description') is not None:
            if mapping['beer_name'] == mapping['description']:
                logger.warning(f"Маппинг: beer_name и description указывают на одну колонку (idx={mapping['beer_name']}), сбрасываем beer_name")
                del mapping['beer_name']
        
        return mapping
    
    def _analyze_column_content(self, col_data: List[str], col_idx: int) -> Dict[str, bool]:
        """
        Анализирует содержимое колонки для определения её назначения.
        
        Args:
            col_data: Список значений из колонки (первые 10 строк)
            col_idx: Индекс колонки
            
        Returns:
            Словарь {название_поля: True/False} - должно ли это поле маппиться к этой колонке
        """
        analysis = {
            'brewery': False,
            'beer_name': False,
            'style': False,
            'abv': False,
            'ibu': False,
            'price': False,
            'volume': False,
            'format_type': False,
            'stock': False,
            'description': False,
        }
        
        if not col_data:
            return analysis
        
        # Анализируем первые несколько значений
        sample_values = col_data[:min(10, len(col_data))]
        
        # Проверяем на числа (цена, объем, ABV, IBU)
        numeric_count = 0
        numeric_values = []
        for val in sample_values:
            try:
                # Пробуем преобразовать в число
                num_val = float(str(val).replace(',', '.').replace(' ', ''))
                numeric_count += 1
                numeric_values.append(num_val)
            except (ValueError, TypeError):
                pass
        
        numeric_ratio = numeric_count / len(sample_values) if sample_values else 0
        
        # Если больше 70% чисел - это числовая колонка
        if numeric_ratio > 0.7:
            if numeric_values:
                avg_value = sum(numeric_values) / len(numeric_values)
                max_value = max(numeric_values)
                
                # Цена обычно в диапазоне 50-10000
                if 50 <= avg_value <= 10000 and max_value < 50000:
                    analysis['price'] = True
                # Объем обычно 0.1-50 литров
                elif 0.1 <= avg_value <= 50:
                    analysis['volume'] = True
                # ABV обычно 0-15%
                elif 0 <= avg_value <= 15:
                    analysis['abv'] = True
                # IBU обычно 0-120
                elif 0 <= avg_value <= 120:
                    analysis['ibu'] = True
        
        # Анализируем текстовое содержимое
        text_values = [str(v).strip() for v in sample_values if v and str(v).strip()]
        if text_values:
            # Проверяем длину текста
            avg_length = sum(len(v) for v in text_values) / len(text_values)
            
            # Название пива обычно длинное (более 15 символов)
            # Может содержать описание, но это все равно название позиции
            if avg_length > 15:
                # Проверяем признаки названия пива:
                # - содержит слова о пиве, стилях, вкусах
                # - обычно первая колонка с длинным текстом
                # - может начинаться с заглавной буквы (название бренда)
                beer_name_indicators = [
                    'ipa', 'lager', 'ale', 'stout', 'porter', 'pilsner', 'wheat', 'sour',
                    'пиво', 'beer', 'пивоварня', 'brewery', 'gose', 'neipa', 'hazy',
                    'томатный', 'клубничный', 'базилик', 'чеснок', 'перцы',  # Примеры из вашего файла
                    'безалкогольный', 'non-alcoholic', 'new england', 'grapefruit',
                    'alisperi', 'back to balance', 'berry forward', 'bitter joy'  # Примеры названий
                ]
                has_beer_keywords = any(
                    keyword in ' '.join(text_values).lower() 
                    for keyword in beer_name_indicators
                )
                
                # Если содержит ключевые слова о пиве - это название
                if has_beer_keywords:
                    analysis['beer_name'] = True
                # Если очень длинное (более 150 символов) и не содержит ключевых слов - это описание
                elif avg_length > 150:
                    # Но только если это не единственная длинная колонка
                    analysis['description'] = True
                # Иначе считаем названием (может содержать описание в названии)
                # Это более агрессивный подход - любая длинная текстовая колонка может быть названием
                # НО: проверяем, не является ли это ЕГАИС кодом
                else:
                    # Проверяем на ЕГАИС коды
                    egais_pattern = re.compile(r'эль[-\s]?\d+|крепкий\s+эль[-\s]?\d+|берлинер\s+вайссе\s+\d+|гозэ\s+\d+', re.IGNORECASE)
                    egais_count = sum(1 for val in col_data[:5] if egais_pattern.search(str(val).lower()))
                    is_egais_col = egais_count >= len(col_data[:5]) * 0.6
                    
                    # Устанавливаем как beer_name только если это НЕ ЕГАИС колонка
                    if not is_egais_col:
                        analysis['beer_name'] = True
                    else:
                        logger.debug(f"Колонка {col_idx} не маппится как beer_name - содержит ЕГАИС коды")
            
            # Пивоварня обычно короткая (5-30 символов) и содержит названия компаний
            elif 5 <= avg_length <= 30:
                has_brewery_keywords = any(
                    keyword in ' '.join(text_values).lower()
                    for keyword in ['brewery', 'пивоварня', 'brewing', 'company', 'компания']
                )
                if has_brewery_keywords:
                    analysis['brewery'] = True
            
            # Стиль обычно короткий (3-50 символов) и содержит названия стилей
            if 3 <= avg_length <= 50:
                style_keywords = [
                    'ipa', 'lager', 'ale', 'stout', 'porter', 'pilsner', 'wheat', 'sour',
                    'gose', 'neipa', 'hazy', 'imperial', 'double', 'triple',
                    'стиль', 'style', 'томатный', 'tomato', 'georgian',
                    'new england', 'grapefruit', 'strawberry', 'sour ale'
                ]
                has_style_keywords = any(
                    keyword in ' '.join(text_values).lower()
                    for keyword in style_keywords
                )
                # Также проверяем, что это не название пива (название обычно длиннее)
                if has_style_keywords and avg_length < 50:
                    analysis['style'] = True
            
            # Формат обычно очень короткий (1-15 символов) и содержит ключевые слова формата
            # Важно: формат должен определяться ДО остатков, чтобы не путаться
            if avg_length <= 15:
                format_keywords = [
                    'банка', 'can', 'кега', 'keg', 'кег', 'бутылка', 'bottle',
                    'ж/б', 'бут', 'л', 'ml', 'мл', 'формат', 'format', 'упаковка',
                    'packaging', 'тара', 'фасовка', 'фасовки', 'тип фасовки'
                ]
                has_format_keywords = any(
                    keyword in ' '.join(text_values).lower()
                    for keyword in format_keywords
                )
                # Формат определяется если есть ключевые слова формата
                # И НЕ является числовой колонкой (числа - это остатки, цена или объем)
                if has_format_keywords:
                    # Если это не чисто числовая колонка (меньше 70% чисел) - это формат
                    if numeric_ratio < 0.7:
                        analysis['format_type'] = True
                    # Если числовая, но содержит ключевые слова формата - тоже формат
                    elif numeric_ratio >= 0.7 and has_format_keywords:
                        # Проверяем, не является ли это объемом или ценой
                        if numeric_values:
                            avg_value = sum(numeric_values) / len(numeric_values)
                            # Если значения в диапазоне объемов (0.1-50) и есть ключевые слова формата - это формат
                            if 0.1 <= avg_value <= 50:
                                analysis['format_type'] = True
                            # Если значения в диапазоне цен, но есть ключевые слова формата - это формат
                            elif 50 <= avg_value <= 10000 and ('банка' in text_lower or 'can' in text_lower or 'бутылка' in text_lower or 'bottle' in text_lower):
                                analysis['format_type'] = True
            
            # Остатки обычно содержат слова типа "много", "мало", "достаточно"
            # Или короткие числовые значения (количество на складе)
            # Важно: остатки НЕ должны быть длинным текстом (это описание)
            # И НЕ должны быть стилями пива (стили могут быть короткими, но это не остатки)
            # И НЕ должны быть заголовками ценовых колонок (например, "Коробка 20шт Цена за банку / Ящик")
            stock_keywords = [
                'много', 'мало', 'достаточно', 'нет', 'есть', 'в наличии',
                'many', 'few', 'enough', 'available', 'stock', 'остаток',
                'наличие', 'склад', 'количество', 'кол-во', 'шт', 'штук'
            ]
            has_stock_keywords = any(
                keyword in ' '.join(text_values).lower()
                for keyword in stock_keywords
            )
            
            # Проверяем, не является ли это заголовком ценовой колонки
            price_header_keywords = ['цена', 'price', 'стоимость', 'cost', 'руб', 'банку', 'ящик', 'коробка', 'кег']
            has_price_header = any(
                keyword in ' '.join(text_values).lower()
                for keyword in price_header_keywords
            )
            
            # Если содержит заголовок ценовой колонки - это НЕ остаток
            if has_price_header:
                has_stock_keywords = False
            
            # Проверяем, не является ли это стилем пива
            style_keywords_check = [
                'ipa', 'lager', 'ale', 'stout', 'porter', 'pilsner', 'wheat', 'sour',
                'gose', 'neipa', 'hazy', 'imperial', 'double', 'triple',
                'weisse', 'weizen', 'hopfen', 'belgian', 'witbier', 'saison',
                'томатный', 'tomato', 'georgian', 'new england', 'grapefruit',
                'strawberry', 'sour ale', 'barrel', 'aged', 'wild'
            ]
            has_style_keywords = any(
                keyword in ' '.join(text_values).lower()
                for keyword in style_keywords_check
            )
            
            # Остатки должны быть короткими (не более 50 символов в среднем)
            # И содержать ключевые слова остатков
            # И НЕ должны содержать ключевые слова стилей пива
            if has_stock_keywords and avg_length <= 50 and not has_style_keywords:
                analysis['stock'] = True
            # Также остатки могут быть чисто числовыми значениями (количество на складе)
            # НО только если это НЕ цена и НЕ объем
            elif numeric_ratio > 0.7 and avg_length <= 10 and not has_style_keywords:
                if numeric_values:
                    avg_value = sum(numeric_values) / len(numeric_values)
                    # Остатки могут быть любыми числами, но если они в диапазоне цен (50-10000)
                    # или объемов (0.1-50), нужно проверить заголовок колонки
                    is_price_range = 50 <= avg_value <= 10000
                    is_volume_range = 0.1 <= avg_value <= 50
                    
                    # Если не в диапазоне цен и объемов - это остатки
                    if not is_price_range and not is_volume_range and 0 <= avg_value <= 10000:
                        analysis['stock'] = True
                    # Если в диапазоне цен или объемов, но есть ключевые слова остатков - это остатки
                    elif (is_price_range or is_volume_range) and has_stock_keywords:
                        analysis['stock'] = True
            
            # Описание обычно длинное (более 50 символов) и содержит описательные слова
            # НЕ должно содержать ключевые слова остатков
            description_keywords = [
                'вкус', 'аромат', 'вкусовые', 'характеристики', 'описание',
                'taste', 'aroma', 'flavor', 'description', 'характер',
                'насыщенный', 'сочный', 'яркий', 'кислый', 'сладкий'
            ]
            has_description_keywords = any(
                keyword in ' '.join(text_values).lower()
                for keyword in description_keywords
            )
            # Описание должно быть длинным и содержать описательные слова
            # И НЕ должно содержать ключевые слова остатков
            if avg_length > 50 and has_description_keywords and not has_stock_keywords:
                analysis['description'] = True
            # Если очень длинное (более 150 символов) и не содержит ключевых слов остатков
            elif avg_length > 150 and not has_stock_keywords:
                analysis['description'] = True
        
        return analysis
    
    def _aggressive_column_mapping(self, df: pd.DataFrame, 
                                   existing_mapping: Dict[str, int]) -> Dict[str, int]:
        """
        Агрессивно определяет маппинг колонок, анализируя содержимое всех колонок.
        
        Args:
            df: DataFrame с данными
            existing_mapping: Существующий маппинг (может быть частичным)
            
        Returns:
            Обновленный маппинг колонок
        """
        mapping = existing_mapping.copy() if existing_mapping else {}
        
        if df.empty:
            return mapping
        
        # Анализируем все колонки
        sample_rows = df.head(min(20, len(df)))
        used_indices = set(mapping.values())
        
        # Собираем анализ для всех колонок
        column_analyses = []
        for col_idx in range(len(df.columns)):
            if col_idx in used_indices:
                continue
            
            try:
                col_data = sample_rows.iloc[:, col_idx].dropna().astype(str).tolist()
                if not col_data:
                    continue
                
                analysis = self._analyze_column_content(col_data, col_idx)
                column_analyses.append((col_idx, analysis))
            except Exception as e:
                logger.debug(f"Ошибка при анализе колонки {col_idx}: {str(e)}")
                continue
        
        # Определяем приоритеты полей (важные поля первыми)
        # format_type должен быть перед stock, чтобы формат не путался с остатками
        # description должен быть перед stock, чтобы длинные тексты определялись как описание
        field_priority = [
            'beer_name', 'price', 'style', 'abv', 'ibu', 
            'brewery', 'volume', 'format_type', 'description', 'stock'
        ]
        
        # Проверяем, не была ли колонка stock неправильно определена
        # Если колонка определена как stock, но содержит стили пива, цену, объем, формат или длинный текст - переопределяем
        if 'stock' in mapping:
            stock_col_idx = mapping['stock']
            try:
                col_data = sample_rows.iloc[:, stock_col_idx].dropna().astype(str).tolist()
                if col_data:
                    avg_len = sum(len(v) for v in col_data[:5]) / min(5, len(col_data))
                    text_lower = ' '.join(col_data[:5]).lower()
                    
                    # Проверяем, не является ли это форматом упаковки
                    format_keywords = [
                        'банка', 'can', 'кега', 'keg', 'кег', 'бутылка', 'bottle',
                        'ж/б', 'бут', 'формат', 'format', 'упаковка', 'packaging',
                        'тара', 'фасовка', 'фасовки', 'тип фасовки'
                    ]
                    has_format_keywords = any(
                        keyword in text_lower for keyword in format_keywords
                    )
                    
                    # Если содержит ключевые слова формата и короткая - это формат, а не остатки
                    if has_format_keywords and avg_len <= 15:
                        logger.debug(f"Колонка {stock_col_idx} переопределена: stock -> format_type (содержит ключевые слова формата упаковки)")
                        if 'format_type' not in mapping:
                            mapping['format_type'] = stock_col_idx
                        del mapping['stock']
                        used_indices.discard(stock_col_idx)
                        used_indices.add(stock_col_idx)
                    else:
                        # Проверяем числовые значения только если это не формат
                        numeric_values = []
                        for val in col_data[:5]:
                            try:
                                num_val = float(str(val).replace(',', '.').replace(' ', ''))
                                numeric_values.append(num_val)
                            except (ValueError, TypeError):
                                pass
                        
                        # Проверяем, не является ли это ценой
                        if numeric_values and 'price' in mapping:
                            price_col_idx = mapping['price']
                            try:
                                price_data = sample_rows.iloc[:, price_col_idx].dropna().astype(str).tolist()
                                price_numeric = []
                                for val in price_data[:5]:
                                    try:
                                        num_val = float(str(val).replace(',', '.').replace(' ', ''))
                                        price_numeric.append(num_val)
                                    except (ValueError, TypeError):
                                        pass
                                
                                # Если значения совпадают с ценой - это цена, а не остатки
                                if price_numeric and numeric_values:
                                    stock_avg = sum(numeric_values) / len(numeric_values)
                                    price_avg = sum(price_numeric) / len(price_numeric)
                                    if abs(stock_avg - price_avg) / max(price_avg, 1) < 0.01:
                                        logger.debug(f"Колонка {stock_col_idx} переопределена: stock -> price (значения совпадают с ценой)")
                                        del mapping['stock']
                                        used_indices.discard(stock_col_idx)
                            except Exception:
                                pass
                        
                        # Проверяем, не является ли это объемом
                        if numeric_values and 'volume' in mapping:
                            volume_col_idx = mapping['volume']
                            try:
                                volume_data = sample_rows.iloc[:, volume_col_idx].dropna().astype(str).tolist()
                                volume_numeric = []
                                for val in volume_data[:5]:
                                    try:
                                        num_val = float(str(val).replace(',', '.').replace(' ', ''))
                                        volume_numeric.append(num_val)
                                    except (ValueError, TypeError):
                                        pass
                                
                                # Если значения совпадают с объемом - это объем, а не остатки
                                if volume_numeric and numeric_values:
                                    stock_avg = sum(numeric_values) / len(numeric_values)
                                    volume_avg = sum(volume_numeric) / len(volume_numeric)
                                    if abs(stock_avg - volume_avg) / max(volume_avg, 0.01) < 0.01:
                                        logger.debug(f"Колонка {stock_col_idx} переопределена: stock -> volume (значения совпадают с объемом)")
                                        del mapping['stock']
                                        used_indices.discard(stock_col_idx)
                            except Exception:
                                pass
                        
                        # Проверяем, не является ли это стилем пива
                        style_keywords_check = [
                            'ipa', 'lager', 'ale', 'stout', 'porter', 'pilsner', 'wheat', 'sour',
                            'gose', 'neipa', 'hazy', 'imperial', 'double', 'triple',
                            'weisse', 'weizen', 'hopfen', 'belgian', 'witbier', 'saison',
                            'томатный', 'tomato', 'georgian', 'new england', 'grapefruit',
                            'strawberry', 'sour ale', 'barrel', 'aged', 'wild'
                        ]
                        has_style_keywords = any(
                            keyword in text_lower for keyword in style_keywords_check
                        )
                        
                        # Если содержит ключевые слова стилей - это стиль, а не остатки
                        if has_style_keywords and 3 <= avg_len <= 50:
                            logger.debug(f"Колонка {stock_col_idx} переопределена: stock -> style (содержит ключевые слова стилей пива)")
                            if 'style' not in mapping:
                                mapping['style'] = stock_col_idx
                            del mapping['stock']
                            used_indices.discard(stock_col_idx)
                            used_indices.add(stock_col_idx)
                        # Если средняя длина больше 50 символов - это описание, а не остатки
                        elif avg_len > 50:
                            description_keywords = [
                                'вкус', 'аромат', 'вкусовые', 'характеристики', 'описание',
                                'taste', 'aroma', 'flavor', 'description', 'характер',
                                'насыщенный', 'сочный', 'яркий', 'кислый', 'сладкий'
                            ]
                            has_description_keywords = any(
                                keyword in text_lower for keyword in description_keywords
                            )
                            # Если содержит описательные слова - это точно описание
                            if has_description_keywords:
                                logger.debug(f"Колонка {stock_col_idx} переопределена: stock -> description (длинный текст с описательными словами)")
                                mapping['description'] = stock_col_idx
                                del mapping['stock']
                                used_indices.discard(stock_col_idx)
                                used_indices.add(stock_col_idx)
            except Exception as e:
                logger.debug(f"Ошибка при проверке колонки stock: {str(e)}")
        
        # Применяем анализ, учитывая приоритеты
        for field in field_priority:
            if field in mapping:
                continue
            
            # Ищем лучшую колонку для этого поля
            best_col = None
            best_score = 0
            
            for col_idx, analysis in column_analyses:
                if col_idx in used_indices:
                    continue
                
                if analysis.get(field, False):
                    # Оцениваем колонку
                    score = 1
                    # Дополнительные проверки для важных полей
                    if field == 'beer_name':
                        # Название пива должно быть длинным текстом
                        col_data = sample_rows.iloc[:, col_idx].dropna().astype(str).tolist()
                        if col_data:
                            avg_len = sum(len(v) for v in col_data[:5]) / min(5, len(col_data))
                            if avg_len > 20:
                                score = 2
                    elif field == 'price':
                        # Цена должна быть числом
                        col_data = sample_rows.iloc[:, col_idx].dropna()
                        numeric_count = sum(1 for v in col_data if isinstance(v, (int, float)) or 
                                           (isinstance(v, str) and v.replace(',', '.').replace(' ', '').replace('-', '').isdigit()))
                        if numeric_count > len(col_data) * 0.7:
                            score = 2
                    
                    if score > best_score:
                        best_score = score
                        best_col = col_idx
            
            if best_col is not None:
                mapping[field] = best_col
                used_indices.add(best_col)
                logger.debug(f"Агрессивно определено поле {field} -> колонка {best_col}")
        
        # Fallback: если beer_name все еще не найден, берем первую длинную текстовую колонку
        # НО пропускаем колонку "ЕГАИС" (обычно первая колонка)
        if 'beer_name' not in mapping:
            # Сначала пробуем первую колонку (обычно там название), но только если это не "ЕГАИС"
            if 0 not in used_indices:
                try:
                    # Проверяем заголовок первой колонки
                    first_col_header = str(headers[0] if headers and len(headers) > 0 else '').lower()
                    if 'егаис' not in first_col_header:
                        col_data = sample_rows.iloc[:, 0].dropna().astype(str).tolist()
                        if col_data:
                            avg_len = sum(len(str(v)) for v in col_data[:5]) / min(5, len(col_data))
                            # КРИТИЧЕСКИ ВАЖНО: Не устанавливаем beer_name на колонку 0, если это ЕГАИС
                            # Проверяем, не является ли это ЕГАИС кодом
                            egais_pattern = re.compile(r'эль[-\s]?\d+|крепкий\s+эль[-\s]?\d+|берлинер\s+вайссе\s+\d+|гозэ\s+\d+', re.IGNORECASE)
                            egais_count = sum(1 for val in col_data[:5] if egais_pattern.search(str(val).lower()))
                            is_egais_col = egais_count >= len(col_data[:5]) * 0.6
                            
                            # Устанавливаем fallback только если это НЕ ЕГАИС колонка И beer_name еще не маппирован
                            if avg_len > 10 and not is_egais_col and 'beer_name' not in mapping:
                                mapping['beer_name'] = 0
                                logger.debug(f"Fallback: beer_name -> колонка 0 (первая колонка, средняя длина: {avg_len:.1f}, не ЕГАИС)")
                            elif is_egais_col:
                                logger.debug(f"Пропускаем fallback для колонки 0 - это ЕГАИС колонка ({egais_count}/{len(col_data[:5])} значений)")
                    else:
                        logger.debug(f"Пропускаем колонку 0, так как это ЕГАИС: {first_col_header}")
                except Exception:
                    pass
            
            # Если первая колонка не подошла, ищем любую длинную текстовую колонку
            if 'beer_name' not in mapping:
                for col_idx in range(len(df.columns)):
                    if col_idx in used_indices:
                        continue
                    try:
                        col_data = sample_rows.iloc[:, col_idx].dropna().astype(str).tolist()
                        if col_data:
                            avg_len = sum(len(str(v)) for v in col_data[:5]) / min(5, len(col_data))
                            if avg_len > 15:
                                mapping['beer_name'] = col_idx
                                logger.debug(f"Fallback: beer_name -> колонка {col_idx} (средняя длина: {avg_len:.1f})")
                                break
                    except Exception:
                        continue
        
        # Fallback: если style не найден, ищем колонку с коротким текстом и ключевыми словами стилей
        if 'style' not in mapping:
            # Сначала пробуем вторую колонку (обычно там стиль, если первая - название)
            if 1 not in used_indices:
                try:
                    col_data = sample_rows.iloc[:, 1].dropna().astype(str).tolist()
                    if col_data:
                        avg_len = sum(len(str(v)) for v in col_data[:5]) / min(5, len(col_data))
                        if 3 <= avg_len <= 50:
                            text_lower = ' '.join(col_data[:5]).lower()
                            style_keywords = ['ipa', 'lager', 'ale', 'stout', 'gose', 'sour', 'style', 'стиль', 
                                            'tomato', 'georgian', 'new england', 'grapefruit', 'strawberry']
                            if any(kw in text_lower for kw in style_keywords):
                                mapping['style'] = 1
                                logger.debug(f"Fallback: style -> колонка 1 (вторая колонка)")
                except Exception:
                    pass
            
            # Если вторая колонка не подошла, ищем любую подходящую
            if 'style' not in mapping:
                for col_idx in range(len(df.columns)):
                    if col_idx in used_indices:
                        continue
                    try:
                        col_data = sample_rows.iloc[:, col_idx].dropna().astype(str).tolist()
                        if col_data:
                            avg_len = sum(len(str(v)) for v in col_data[:5]) / min(5, len(col_data))
                            if 3 <= avg_len <= 50:
                                text_lower = ' '.join(col_data[:5]).lower()
                                style_keywords = ['ipa', 'lager', 'ale', 'stout', 'gose', 'sour', 'style', 'стиль',
                                                'tomato', 'georgian', 'new england', 'grapefruit', 'strawberry']
                                if any(kw in text_lower for kw in style_keywords):
                                    mapping['style'] = col_idx
                                    logger.debug(f"Fallback: style -> колонка {col_idx}")
                                    break
                    except Exception:
                        continue
        
        return mapping
    
    def _extract_brewery_from_name(self, beer_name: str) -> Optional[str]:
        """
        Извлекает название пивоварни из начала названия пива.
        
        Примеры:
        - "ALISPERI (Fresh Brewed) Томатный гозе..." -> "ALISPERI"
        - "BACK TO BALANCE (Fresh Brewed) Безалкогольный..." -> "BACK TO BALANCE"
        - "PARADOX BERRY FORWARD..." -> "PARADOX"
        - "Бокал "Paradox", 400 ml" -> "Paradox"
        - "Кружка "Paradox", 500 ml" -> "Paradox"
        
        Args:
            beer_name: Полное название пива
            
        Returns:
            Название пивоварни или None
        """
        if not beer_name or len(beer_name) < 3:
            return None
        
        # Убираем лишние пробелы
        beer_name = beer_name.strip()
        
        # Паттерны для извлечения пивоварни:
        # 1. Пивоварня в кавычках (например, "Paradox" в "Бокал "Paradox", 400 ml")
        # 2. Слово в верхнем регистре до скобки: "ALISPERI (Fresh Brewed)"
        # 3. Несколько слов в верхнем регистре: "BACK TO BALANCE"
        # 4. Одно слово в верхнем регистре в начале
        
        # Паттерн 1: Пивоварня в кавычках (обычно после слов типа "Бокал", "Кружка", "Стакан")
        # Ищем текст в кавычках (как обычных, так и типографских)
        quote_patterns = [
            r'[""]([A-ZА-ЯЁ][A-ZА-ЯЁa-zа-яё\s]{1,30}?)[""]',  # "Paradox" или "Paradox"
            r'[«»]([A-ZА-ЯЁ][A-ZА-ЯЁa-zа-яё\s]{1,30}?)[«»]',  # «Paradox»
        ]
        for pattern in quote_patterns:
            matches = re.findall(pattern, beer_name)
            if matches:
                brewery = matches[0].strip()
                # Проверяем, что это не слишком длинное и не обычное слово
                if 2 <= len(brewery) <= 30:
                    common_words = ['ipa', 'ale', 'lager', 'stout', 'porter', 'pilsner', 'wheat', 'sour', 'ml', 'л']
                    if brewery.upper() not in [w.upper() for w in common_words]:
                        logger.debug(f"Извлечена пивоварня из кавычек: {brewery} из '{beer_name}'")
                        return brewery
        
        # Паттерн 2: До скобки
        match = re.match(r'^([A-ZА-ЯЁ][A-ZА-ЯЁ\s]+?)\s*\(', beer_name)
        if match:
            brewery = match.group(1).strip()
            # Проверяем, что это не слишком длинное (больше 50 символов - вероятно не пивоварня)
            if len(brewery) <= 50:
                return brewery
        
        # Паттерн 3: Несколько слов в верхнем регистре в начале
        # Ищем последовательность слов в верхнем регистре (минимум 2, максимум 5 слов)
        match = re.match(r'^([A-ZА-ЯЁ][A-ZА-ЯЁ\s]{2,50}?)(?:\s+[a-zа-яё]|\s*\(|\s*$)', beer_name)
        if match:
            brewery = match.group(1).strip()
            # Проверяем количество слов (пивоварня обычно 1-3 слова)
            word_count = len(brewery.split())
            if 1 <= word_count <= 5 and len(brewery) <= 50:
                return brewery
        
        # Паттерн 4: Одно слово в верхнем регистре в начале (если оно не слишком короткое)
        match = re.match(r'^([A-ZА-ЯЁ][A-ZА-ЯЁ]{2,30})\s+', beer_name)
        if match:
            brewery = match.group(1).strip()
            # Проверяем, что это не обычное слово (например, не "IPA", "ALE")
            common_words = ['ipa', 'ale', 'lager', 'stout', 'porter', 'pilsner', 'wheat', 'sour']
            if brewery.upper() not in [w.upper() for w in common_words]:
                return brewery
        
        return None
    
    def _extract_row_data(self, row: pd.Series, 
                         col_mapping: Dict[str, int],
                         df: pd.DataFrame = None,
                         skip_price: bool = False,
                         default_brewery: Optional[str] = None,
                         supplier_type_enum: Optional[SupplierType] = None) -> Optional[Dict]:
        """
        Извлекает данные из строки DataFrame.
        
        Args:
            row: Строка DataFrame
            col_mapping: Соответствие полей и индексов колонок
            df: DataFrame (опционально, для доступа к заголовкам колонок)
            skip_price: Не извлекать цену (для ветки с несколькими ценовыми колонками)
            default_brewery: Для типа brewery — подставить пивоварню, если в строке её нет
            supplier_type_enum: Тип поставщика (для валидации строк без названия, но с форматом/ценой)
            
        Returns:
            Словарь с данными позиции или None
        """
        item = {}
        
        # Проверяем, что маппинг не пустой
        if not col_mapping:
            return None
        
        # Извлекаем формат и объем из заголовков колонок, если они не определены
        if df is not None:
            # Проверяем заголовки колонок для извлечения формата
            for field, idx in col_mapping.items():
                if field == 'price' and idx < len(df.columns):
                    # Проверяем заголовок колонки цены
                    header = str(df.columns[idx]).lower() if hasattr(df.columns[idx], 'lower') else str(df.columns[idx])
                    
                    # Если заголовок содержит "коробка" или "банка" - формат банка
                    if ('коробка' in header or 'банка' in header or 'can' in header) and 'format_type' not in item:
                        item['format_type'] = 'банка'
                        # Извлекаем объем банки из заголовка, если указан (например "банка 0.33л" или "0.33")
                        volume_match = re.search(r'\(?(\d+[.,]?\d*)\)?\s*л', header, re.IGNORECASE)
                        if volume_match:
                            try:
                                vol_val = float(volume_match.group(1).replace(',', '.'))
                                if 'volume' not in item:
                                    item['volume'] = vol_val
                            except:
                                pass
                        # Если объем не указан в заголовке, но есть в скобках типа "(0.33)"
                        if 'volume' not in item:
                            volume_match = re.search(r'\((\d+[.,]?\d*)\)', header)
                            if volume_match:
                                try:
                                    vol_val = float(volume_match.group(1).replace(',', '.'))
                                    # Если значение меньше 1, это литры (0.33 = 0.33л)
                                    if vol_val < 1:
                                        item['volume'] = vol_val
                                    # Если значение больше 1, но меньше 10, это тоже может быть литры (0.33л)
                                    elif vol_val < 10:
                                        item['volume'] = vol_val / 100 if vol_val > 1 else vol_val
                                except:
                                    pass
                    
                    # Если заголовок содержит "кег" - формат кег
                    if ('кег' in header or 'keg' in header) and 'format_type' not in item:
                        item['format_type'] = 'кег'
                        # Извлекаем объем из заголовка "Кег 20л"
                        volume_match = re.search(r'(\d+)\s*л', header, re.IGNORECASE)
                        if volume_match:
                            vol_val = float(volume_match.group(1))
                            if 'volume' not in item:
                                item['volume'] = vol_val
        
        # Извлекаем значения по маппингу
        # Важно: извлекаем каждое поле только один раз, чтобы избежать дублирования
        # Сначала обрабатываем format_type, чтобы извлечь volume
        for field, idx in col_mapping.items():
            try:
                # Проверяем корректность индекса
                if idx is None or idx < 0:
                    continue
                    
                # Проверяем, что индекс в пределах строки
                if idx >= len(row):
                    continue
                
                # Проверяем, что значение не пустое
                if pd.isna(row.iloc[idx]):
                    continue
                
                value = row.iloc[idx]
                
                # Обрабатываем тире и другие обозначения пустых значений
                if isinstance(value, str):
                    value_str_check = value.strip().lower()
                    # Тире, прочерк, пустая строка считаем пустым значением
                    if value_str_check in ['-', '—', '–', '', 'nan', 'none', 'null', 'хх', 'xx']:
                        continue
                
                # Преобразуем в строку и очищаем
                # Определяем value_str для всех случаев (для использования в format_type и других полях)
                # Безопасно преобразуем в строку, учитывая все типы значений
                if value is None:
                    value_str = ''
                elif isinstance(value, (int, float)):
                    value_str = str(value)
                else:
                    value_str = str(value).strip() if value else ''
                
                # Обрабатываем числовые значения для специальных полей
                if isinstance(value, (int, float)):
                    # Для чисел сохраняем как строку, но без лишних нулей
                    if field == 'price':
                        item[field] = str(value)
                        continue
                    elif field == 'abv':
                        item[field] = str(value)
                        continue
                    elif field == 'volume':
                        # КРИТИЧЕСКИ ВАЖНО: Для банок и кегов volume должен устанавливаться ТОЛЬКО из заголовка колонки (price_info['volume'])
                        # Не перезаписываем volume из значения ячейки (цены), если это банка или кег
                        if item.get('_volume_from_header'):
                            # Volume уже установлен из заголовка колонки, пропускаем перезапись
                            format_name = item.get('format_type', 'unknown')
                            logger.debug(f"Пропущена перезапись volume для {format_name} - volume установлен из заголовка колонки, не перезаписываем значением '{value}'")
                            continue
                        format_lower = str(item.get('format_type', '')).lower().strip()
                        if 'кег' in format_lower or 'keg' in format_lower or 'банка' in format_lower or 'can' in format_lower:
                            # Для банок и кегов volume уже установлен из заголовка колонки, пропускаем перезапись
                            logger.debug(f"Пропущена перезапись volume для {format_lower} - volume должен быть из заголовка колонки, а не из значения ячейки '{value}'")
                            continue
                        # Для других форматов: volume сохраняем как число только если он еще не был извлечен
                        if 'volume' not in item:
                            item[field] = float(value)
                        continue
                    elif field == 'format_type':
                        # Для format_type числовые значения обрабатываются в блоке ниже
                        # Проверяем, не является ли это только число - тогда пропускаем
                        value_str_clean = value_str.strip() if value_str else ''
                        if value_str_clean and re.match(r'^\d+[.,]?\d*$', value_str_clean):
                            logger.debug(f"Пропущено значение format_type '{value_str}' - это только число, не формат")
                            continue
                        # Если format_type уже установлен из заголовков колонок, пропускаем
                        if item.get('format_type') and item['format_type'].lower() in ['банка', 'can', 'кег', 'кега', 'keg', 'бутылка', 'bottle']:
                            logger.debug(f"format_type уже установлен из заголовка колонки: '{item['format_type']}', пропускаем перезапись числовым значением '{value_str}'")
                            continue
                    else:
                        # Для других полей с числовыми значениями
                        item[field] = str(int(value)) if float(value).is_integer() else str(value)
                        continue
                
                # Обрабатываем строковые значения
                if value_str and value_str.lower() not in ['nan', 'none', '']:
                        # Если это поле ABV или IBU, и значение содержит "/", парсим его
                        if field == 'abv' and '/' in value_str:
                            # Парсим формат типа "13.0% / 20" или "5,1 / 12 / 27"
                            parts = [p.strip() for p in value_str.split('/')]
                            if len(parts) >= 1:
                                # Извлекаем ABV из первой части (убираем %, заменяем запятую на точку)
                                abv_val = parts[0].strip().replace('%', '').replace(',', '.')
                                try:
                                    item['abv'] = str(float(abv_val))
                                except:
                                    item['abv'] = abv_val
                            # Извлекаем IBU из второй части (если формат "ABV / IBU")
                            if len(parts) >= 2:
                                ibu_val = parts[1].strip()
                                # Если IBU уже не установлен или мапится на ту же колонку
                                if 'ibu' not in item or col_mapping.get('ibu') == col_mapping.get('abv'):
                                    item['ibu'] = ibu_val
                            # Если формат "ABV / OG / IBU" (3 части)
                            elif len(parts) >= 3:
                                ibu_val = parts[2].strip()
                                if 'ibu' not in item or col_mapping.get('ibu') == col_mapping.get('abv'):
                                    item['ibu'] = ibu_val
                        elif field == 'ibu' and '/' in value_str:
                            # Если IBU мапится на ту же колонку, что и ABV, не парсим отдельно
                            if col_mapping.get('ibu') == col_mapping.get('abv'):
                                continue
                            # Парсим IBU из формата "ABV / OG / IBU"
                            parts = [p.strip() for p in value_str.split('/')]
                            if len(parts) >= 3:
                                ibu_val = parts[2].strip()
                                item['ibu'] = ibu_val
                            elif len(parts) >= 2:
                                # Если формат "ABV / IBU"
                                ibu_val = parts[1].strip()
                                item['ibu'] = ibu_val
                            elif len(parts) >= 1:
                                item['ibu'] = parts[0].strip()
                        elif field == 'format_type':
                            # КРИТИЧЕСКИ ВАЖНО: Если format_type уже установлен из заголовков колонок (например, "банка" или "кег"),
                            # не перезаписываем его числовым значением или неправильным значением из ячейки
                            if item.get('format_type') and item['format_type'].lower() in ['банка', 'can', 'кег', 'кега', 'keg', 'бутылка', 'bottle']:
                                logger.debug(f"format_type уже установлен из заголовка колонки: '{item['format_type']}', пропускаем перезапись значением '{value_str}'")
                                continue
                            
                            # Парсим формат и объем из колонки типа "БАНКА 0,45\n (12 шт/кор)" или "ж/б 0.45 / 20"
                            # Формат: "БАНКА 0,45" (жестяная банка 0.45 л) или "банка 0.5"
                            # Важно: извлекаем объем (0.45), игнорируем количество в коробке (12 шт)
                            
                            # Сначала очищаем от количества в коробке: " (12 шт/кор)", " (20 шт/кор)" и т.д.
                            value_cleaned = re.sub(r'\s*\(\d+\s*шт[^)]*\)', '', value_str, flags=re.IGNORECASE)
                            value_cleaned = re.sub(r'\s*\(\d+\s*шт\)', '', value_cleaned, flags=re.IGNORECASE)
                            value_cleaned = value_cleaned.strip()
                            
                            # Если есть перенос строки, берем первую строку (там формат и объем)
                            if '\n' in value_cleaned:
                                lines = [l.strip() for l in value_cleaned.split('\n') if l.strip()]
                                if lines:
                                    value_cleaned = lines[0]
                            
                            if '/' in value_cleaned:
                                # Разбиваем по "/" с ограничением в 2 части (формат/объем и количество)
                                # Но сначала нужно найти последний "/" который разделяет формат и количество
                                parts = value_cleaned.rsplit('/', 1)  # Разбиваем справа только один раз
                                
                                if len(parts) >= 1:
                                    format_part = parts[0].strip()
                                    # Извлекаем формат (банка, кега, бутылка)
                                    format_lower = format_part.lower()
                                    if 'ж/б' in format_lower or 'банка' in format_lower or 'can' in format_lower:
                                        item['format_type'] = 'банка'
                                    elif 'кeg' in format_lower or 'кега' in format_lower or 'keg' in format_lower:
                                        item['format_type'] = 'кега'
                                    elif 'бут' in format_lower or 'bottle' in format_lower or 'бутылка' in format_lower:
                                        item['format_type'] = 'бутылка'
                                    else:
                                        item['format_type'] = format_part
                                    
                                    # Извлекаем объем из формата (например "0.45" или "0.5")
                                    # Ищем все числа в строке и берем последнее (обычно это объем)
                                    all_numbers = re.findall(r'(\d+[.,]?\d*)', format_part)
                                    if all_numbers:
                                        # Берем последнее число (оно обычно объем после формата)
                                        vol_val = all_numbers[-1].replace(',', '.')
                                        try:
                                            vol_float = float(vol_val)
                                            # Если объем меньше 1, предполагаем литры, иначе мл
                                            if vol_float < 1:
                                                item['volume'] = vol_float  # Сохраняем как число, не строку
                                            elif vol_float < 100:
                                                # Если число от 1 до 100, вероятно это литры (например, 5 л)
                                                item['volume'] = vol_float  # Сохраняем как число
                                            else:
                                                item['volume'] = vol_float / 1000  # мл в литры, сохраняем как число
                                        except Exception as e:
                                            pass
                                
                                if len(parts) >= 2:
                                    # Вторая часть может содержать количество в упаковке
                                    qty_part = parts[1].strip()
                                    # Если это число, можем использовать для объема или оставить как есть
                                    pass
                            else:
                                # Если нет "/", обрабатываем формат типа "БАНКА 0,45", "кег 20 л" или "банка"
                                # Извлекаем объем из формата, если он указан
                                # Паттерны: "БАНКА 0,45", "кег 20 л", "банка 0.5", "банка"
                                volume_patterns = [
                                    r'(\d+[.,]?\d*)\s*л\b',  # "20 л", "0.5 л"
                                    r'(\d+[.,]?\d*)\s*ml\b',  # "500 ml"
                                    r'(\d+[.,]?\d*)\s*мл\b',  # "500 мл"
                                    r'\b(\d+[.,]\d+)\b',  # "0,45" или "0.45" (с запятой/точкой)
                                ]
                                
                                volume_found = False
                                for pattern in volume_patterns:
                                    match = re.search(pattern, value_cleaned, re.IGNORECASE)
                                    if match:
                                        vol_val = match.group(1).replace(',', '.')
                                        try:
                                            vol_float = float(vol_val)
                                            if 'мл' in match.group(0).lower() or 'ml' in match.group(0).lower():
                                                item['volume'] = vol_float / 1000
                                            else:
                                                item['volume'] = vol_float
                                            volume_found = True
                                            break
                                        except Exception:
                                            pass
                                
                                # Если не нашли объем по паттернам, ищем число с запятой/точкой (объем) или маленькое число (< 1)
                                if not volume_found:
                                    all_numbers = re.findall(r'(\d+[.,]\d+)', value_cleaned)  # Только числа с запятой/точкой (0.45, 0,45)
                                    if all_numbers:
                                        # Берем первое число с запятой/точкой (это объем)
                                        vol_val = all_numbers[0].replace(',', '.')
                                        try:
                                            vol_float = float(vol_val)
                                            if vol_float < 1:
                                                item['volume'] = vol_float  # 0.45, 0.33 - литры
                                            elif vol_float < 10:
                                                item['volume'] = vol_float  # 5, 10 - литры для кегов
                                            # Игнорируем большие числа (12, 20 - это количество в коробке)
                                        except Exception:
                                            pass
                                    else:
                                        # Если нет чисел с запятой, ищем маленькие целые числа (< 1 или 1-10 для кегов)
                                        small_numbers = re.findall(r'\b(\d+)\b', value_cleaned)
                                        if small_numbers:
                                            for num_str in small_numbers:
                                                num_val = float(num_str)
                                                # Если число 1-10 и есть слово "л" или "л)" рядом - это объем кега
                                                if 1 <= num_val <= 50 and ('л' in value_cleaned.lower() or 'l' in value_cleaned.lower()):
                                                    item['volume'] = num_val
                                                    volume_found = True
                                                    break
                                                # Если число < 1 (не бывает, но на всякий случай)
                                                elif num_val < 1:
                                                    item['volume'] = num_val
                                                    volume_found = True
                                                    break
                                
                                # Сохраняем формат (нормализуем)
                                format_lower = value_str.lower()
                                
                                # Проверяем, не является ли значение только числом (например, "300", "350", "320")
                                # Если это только число - это не формат, а возможно объем или цена
                                if re.match(r'^\d+[.,]?\d*$', value_str.strip()):
                                    # Это только число, не формат - пропускаем
                                    logger.debug(f"Пропущено значение format_type '{value_str}' - это только число, не формат")
                                    # НЕ устанавливаем format_type, если это только число
                                    continue
                                
                                if 'ж/б' in format_lower or 'банка' in format_lower or 'can' in format_lower:
                                    item['format_type'] = 'банка'
                                elif 'кeg' in format_lower or 'кега' in format_lower or 'keg' in format_lower or 'кег' in format_lower:
                                    item['format_type'] = 'кега'
                                elif 'бут' in format_lower or 'bottle' in format_lower or 'бутылка' in format_lower:
                                    item['format_type'] = 'бутылка'
                                else:
                                    # Сохраняем оригинальное значение, но очищаем от объема
                                    # Удаляем числа с единицами измерения
                                    cleaned_format = re.sub(r'\s*\d+[.,]?\d*\s*(л|ml|мл|l)\b', '', value_str, flags=re.IGNORECASE)
                                    cleaned_format = cleaned_format.strip()
                                    # Если после очистки осталось только число - не сохраняем
                                    if cleaned_format and not re.match(r'^\d+[.,]?\d*$', cleaned_format):
                                        # Проверяем, что это не число перед установкой
                                        if not re.match(r'^\d+[.,]?\d*$', cleaned_format):
                                            item['format_type'] = cleaned_format
                                    elif value_str.strip() and not re.match(r'^\d+[.,]?\d*$', value_str.strip()):
                                        # Проверяем еще раз, что это не число
                                        if not re.match(r'^\d+[.,]?\d*$', value_str.strip()):
                                            item['format_type'] = value_str.strip()
                        elif field == 'volume':
                            # КРИТИЧЕСКИ ВАЖНО: Для банок и кегов volume должен устанавливаться ТОЛЬКО из заголовка колонки (price_info['volume'])
                            # Не перезаписываем volume из значения ячейки (цены), если это банка или кег
                            if item.get('_volume_from_header'):
                                # Volume уже установлен из заголовка колонки, пропускаем перезапись
                                format_name = item.get('format_type', 'unknown')
                                logger.debug(f"Пропущена перезапись volume для {format_name} - volume установлен из заголовка колонки, не перезаписываем значением '{value_str}'")
                                continue
                            format_lower = str(item.get('format_type', '')).lower().strip()
                            if 'кег' in format_lower or 'keg' in format_lower or 'банка' in format_lower or 'can' in format_lower:
                                # Для банок и кегов volume уже установлен из заголовка колонки, пропускаем перезапись
                                logger.debug(f"Пропущена перезапись volume для {format_lower} - volume должен быть из заголовка колонки, а не из значения ячейки '{value_str}'")
                                continue
                            # Если volume уже был извлечен из format_type или установлен из заголовков колонок, не перезаписываем
                            if 'volume' not in item or item.get('volume') is None:
                                if isinstance(value, (int, float)):
                                    vol_val = float(value)
                                    # Если это большое число (>= 10) и формат банка/бутылка - это количество в коробке, не объем
                                    format_lower = str(item.get('format_type', '')).lower().strip()
                                    if vol_val >= 10 and ('банка' in format_lower or 'бутылка' in format_lower or 'can' in format_lower or 'bottle' in format_lower):
                                        logger.debug(f"Пропущено значение volume '{vol_val}' - это количество в коробке для {format_lower}, а не объем")
                                        continue
                                    item['volume'] = vol_val
                                else:
                                    # Пробуем извлечь число из строки
                                    vol_match = re.search(r'(\d+[.,]?\d*)', str(value))
                                    if vol_match:
                                        vol_val = vol_match.group(1).replace(',', '.')
                                        try:
                                            vol_float = float(vol_val)
                                            # Если это большое число (>= 10) и формат банка/бутылка - это количество в коробке
                                            format_lower = str(item.get('format_type', '')).lower().strip()
                                            if vol_float >= 10 and ('банка' in format_lower or 'бутылка' in format_lower or 'can' in format_lower or 'bottle' in format_lower):
                                                logger.debug(f"Пропущено значение volume '{vol_float}' - это количество в коробке для {format_lower}, а не объем")
                                                continue
                                            item['volume'] = vol_float
                                        except:
                                            pass
                        elif field == 'stock':
                            # Проверяем, не является ли значение заголовком ценовой колонки
                            # Заголовки типа "Коробка 20шт Цена за банку / Ящик" не должны попадать в stock
                            price_header_keywords = ['цена', 'price', 'стоимость', 'cost', 'руб', 'банку', 'ящик', 'коробка', 'кег', 'keg']
                            value_lower = value_str.lower()
                            if any(keyword in value_lower for keyword in price_header_keywords):
                                # Это заголовок ценовой колонки, пропускаем
                                logger.debug(f"Пропущено значение stock '{value_str}' - это заголовок ценовой колонки")
                                continue
                            # Если это валидное значение stock, сохраняем его
                            item[field] = value_str
                        elif field == 'beer_name':
                            # Никогда не подставлять в название колонку "Описание" или "Стиль" (дистрибьюция: Фасовка/БА)
                            if df is not None and idx < len(df.columns):
                                h = str(df.columns[idx]).strip().lower()
                                if 'описание' in h or h == 'description':
                                    continue
                                if 'стиль' in h or h == 'style':
                                    continue
                            # В одной ячейке может быть "название + описание" (первая строка — название, остальное — описание)
                            if '\n' in value_str:
                                parts = [p.strip() for p in value_str.split('\n') if p.strip()]
                                if len(parts) >= 2:
                                    first_line = parts[0]
                                    rest = '\n'.join(parts[1:]).strip()
                                    # Первая строка — короткое название (до 120 символов), остальное — описание
                                    if len(first_line) <= 120 and len(rest) >= 15:
                                        item['beer_name'] = first_line
                                        # Убираем название из начала описания, если оно там дублируется
                                        rest_lower = rest.lower()
                                        first_line_lower = first_line.lower()
                                        if rest_lower.startswith(first_line_lower):
                                            # Описание начинается с названия - удаляем его
                                            rest = rest[len(first_line):].strip()
                                            # Убираем возможные разделители в начале
                                            rest = re.sub(r'^[:\-\s]+', '', rest)
                                        if 'description' not in item or not (item.get('description') or '').strip():
                                            item['description'] = rest
                                        continue
                            # Заголовок колонки: если явно "Название" — не подменять значение логикой "описание"
                            col_header = ''
                            if df is not None and idx < len(df.columns):
                                col_header = str(df.columns[idx]).strip().lower()
                            is_name_column = any(k in col_header for k in ('название', 'наименование', 'name', 'назва'))
                            # КРИТИЧЕСКАЯ ПРОВЕРКА: Если значение похоже на описание, а не название
                            # Описание обычно длинное (более 100 символов) и содержит описательные слова
                            value_lower = value_str.lower()
                            description_keywords = [
                                'вкус', 'аромат', 'вкусовые', 'характеристики', 'описание',
                                'taste', 'aroma', 'flavor', 'description', 'характер',
                                'авторская', 'интерпретация', 'великолепное', 'смешение',
                                'немецкий', 'сухой', 'уверенной', 'благородной',
                                'на тему', 'с добавлением', 'с сухим', 'охмелением',
                                'свежайшим', 'классических', 'благородной ароматикой',
                                'кислый эль', 'вайцен', 'пилснера'
                            ]
                            has_description_keywords = any(keyword in value_lower for keyword in description_keywords)
                            is_too_long = len(value_str) > 100
                            # #region agent log
                            try:
                                import os as _os
                                _lp = _os.path.normpath(_os.path.join(_os.path.dirname(_os.path.abspath(__file__)), '..', '..', '..', '.cursor', 'debug-6958ea.log'))
                                _entering_overwrite = not is_name_column and is_too_long and has_description_keywords
                                _data = {"hypothesisId": "H2", "message": "beer_name branch", "data": {"col_header": col_header, "is_name_column": is_name_column, "value_len": len(value_str), "is_too_long": is_too_long, "has_description_keywords": has_description_keywords, "entering_overwrite_branch": _entering_overwrite}, "timestamp": __import__('time').time() * 1000}
                                with open(_lp, "a", encoding="utf-8") as _f:
                                    _f.write(json.dumps(_data, ensure_ascii=False) + "\n")
                            except Exception:
                                pass
                            # #endregion agent log

                            # Если колонка явно "Название" — используем значение как название
                            # Но если в ячейке "название + описание" (длинный текст с «...» или описательными словами) — вытаскиваем короткое название
                            if is_name_column:
                                if len(value_str) <= 80:
                                    item['beer_name'] = value_str
                                    continue
                                # Длинное значение в колонке "Название" — пробуем взять только название (кавычки или первое предложение)
                                short_name = None
                                rest_description = None
                                quote_patterns = [
                                    r'^[«""]([^«""]{1,80}?)[»""]\s*(.*)',  # «Название» остаток
                                    r'^[""]([^""]{1,80}?)[""]\s*(.*)',
                                ]
                                for pattern in quote_patterns:
                                    match = re.match(pattern, value_str.strip(), re.DOTALL)
                                    if match:
                                        short_name = match.group(1).strip()
                                        rest_description = (match.group(2).strip() if match.lastindex >= 2 and match.group(2) else None) or None
                                        if short_name and 2 <= len(short_name) <= 80:
                                            break
                                    short_name = None
                                    rest_description = None
                                if not short_name and (is_too_long or has_description_keywords):
                                    # Первое предложение до точки или до описательной фразы (до ~60 символов)
                                    first_sentence = re.match(r'^([^.]{5,60}?)(?:\.|$)', value_str.strip())
                                    if first_sentence:
                                        candidate = first_sentence.group(1).strip()
                                        if not any(kw in candidate.lower() for kw in description_keywords):
                                            short_name = candidate
                                            rest_description = value_str[len(candidate):].strip().lstrip('.').strip() or None
                                if short_name:
                                    item['beer_name'] = short_name
                                    if rest_description and ('description' not in item or not (item.get('description') or '').strip()):
                                        item['description'] = rest_description
                                else:
                                    item['beer_name'] = value_str
                                continue
                            
                            # Если это описание (длинное И содержит описательные слова) и колонка НЕ "Название"
                            if not is_name_column and is_too_long and has_description_keywords:
                                # Пытаемся извлечь название из начала текста (до первого предложения или до точки)
                                # Ищем название в кавычках или в начале текста
                                name_match = None
                                
                                # Паттерн 1: Название в кавычках в начале (приоритет)
                                quote_patterns = [
                                    r'^[«""]([^«""]{1,50}?)[»""]',  # «Название» или "Название" в начале
                                    r'[«""]([^«""]{1,50}?)[»""]',  # «Название» или "Название" где-то в тексте
                                    r'^[""]([^""]{1,50}?)[""]',  # "Название" в начале
                                    r'[""]([^""]{1,50}?)[""]',  # "Название" где-то в тексте
                                ]
                                for pattern in quote_patterns:
                                    match = re.search(pattern, value_str)
                                    if match:
                                        name_match = match.group(1).strip()
                                        # Проверяем, что это не описание
                                        if name_match and len(name_match) >= 3:
                                            name_lower = name_match.lower()
                                            # Если название содержит описательные слова, но короткое - это все равно название
                                            if len(name_match) <= 50:
                                                break
                                        name_match = None
                                
                                # Паттерн 2: Название до точки или двоеточия (первые 30-50 символов)
                                if not name_match:
                                    # Ищем до точки, двоеточия или запятой
                                    name_end_pattern = r'^([^.:,]{10,50}?)(?:[.:,]|$)'
                                    match = re.match(name_end_pattern, value_str.strip())
                                    if match:
                                        potential_name = match.group(1).strip()
                                        # Проверяем, что это не описание (не содержит описательных слов)
                                        potential_lower = potential_name.lower()
                                        if not any(keyword in potential_lower for keyword in description_keywords):
                                            name_match = potential_name
                                
                                # Паттерн 3: Первые слова до первого описательного слова или фразы
                                if not name_match:
                                    words = value_str.split()
                                    name_words = []
                                    text_lower = value_str.lower()
                                    
                                    # Проверяем наличие описательных фраз
                                    description_phrases = [
                                        'на тему', 'с добавлением', 'с сухим охмелением',
                                        'авторская интерпретация', 'великолепное смешение',
                                        'немецкий вайцен', 'немецкий хмель', 'кислый эль'
                                    ]
                                    
                                    # Находим позицию первого описательного слова/фразы
                                    first_desc_pos = len(words)
                                    for phrase in description_phrases:
                                        phrase_pos = text_lower.find(phrase)
                                        if phrase_pos >= 0:
                                            # Подсчитываем количество слов до этой фразы
                                            words_before = len(text_lower[:phrase_pos].split())
                                            if words_before < first_desc_pos:
                                                first_desc_pos = words_before
                                    
                                    # Берем слова до первого описательного слова/фразы
                                    for i, word in enumerate(words[:min(10, first_desc_pos)]):
                                        word_lower = word.lower().strip('.,:;!?')
                                        # Проверяем отдельные слова
                                        if any(keyword in word_lower for keyword in description_keywords):
                                            break
                                        name_words.append(word)
                                    
                                    if name_words:
                                        potential_name = ' '.join(name_words).strip('.,:;!?')
                                        # Минимум 3 символа, максимум 60 символов для названия
                                        if 3 <= len(potential_name) <= 60:
                                            name_match = potential_name
                                
                                # Если нашли название - используем его, описание сохраняем отдельно
                                if name_match and len(name_match) >= 3:
                                    item['beer_name'] = name_match
                                    # Сохраняем полный текст как описание, но убираем название из начала
                                    if 'description' not in item or not item.get('description'):
                                        desc = value_str
                                        # Убираем название из начала описания, если оно там есть
                                        name_lower = name_match.lower()
                                        desc_lower = desc.lower()
                                        if desc_lower.startswith(name_lower):
                                            desc = desc[len(name_match):].strip()
                                            desc = re.sub(r'^[:\-\s]+', '', desc)
                                        item['description'] = desc
                                    logger.debug(f"Извлечено название из описания: '{name_match}' (из '{value_str[:50]}...')")
                                else:
                                    # Если не удалось извлечь название, сохраняем как описание
                                    if 'description' not in item or not item.get('description'):
                                        item['description'] = value_str
                                    # Пытаемся взять первые слова как название
                                    first_words = ' '.join(value_str.split()[:5]).strip('.,:;!?')
                                    if len(first_words) >= 3:
                                        item['beer_name'] = first_words
                                        # Убираем название из описания, если оно там есть
                                        desc = item.get('description', '')
                                        if desc.lower().startswith(first_words.lower()):
                                            desc = desc[len(first_words):].strip()
                                            desc = re.sub(r'^[:\-\s]+', '', desc)
                                            item['description'] = desc
                                    else:
                                        item['beer_name'] = value_str[:50]  # Первые 50 символов как fallback
                                    logger.debug(f"Значение определено как описание, установлено beer_name из первых слов: '{item.get('beer_name')}'")
                            else:
                                # Обычная обработка названия
                                # Извлекаем brewery из начала названия пива, если его нет отдельно
                                if 'brewery' not in col_mapping or not item.get('brewery'):
                                    brewery_from_name = self._extract_brewery_from_name(value_str)
                                    if brewery_from_name:
                                        # Нормализуем brewery (удаляем город) сразу после извлечения
                                        from parser_app.domain.services.normalization import DataNormalizer
                                        normalizer = DataNormalizer()
                                        brewery_normalized = normalizer.normalize_brewery(brewery_from_name)
                                        item['brewery'] = brewery_normalized if brewery_normalized else brewery_from_name
                                        # Удаляем brewery из названия
                                        beer_name_cleaned = value_str
                                        # Если brewery была в кавычках, удаляем кавычки с содержимым
                                        quote_patterns = [
                                            r'[""]([A-ZА-ЯЁ][A-ZА-ЯЁa-zа-яё\s]{1,30}?)[""]',
                                            r'[«»]([A-ZА-ЯЁ][A-ZА-ЯЁa-zа-яё\s]{1,30}?)[«»]',
                                        ]
                                        brewery_in_quotes = False
                                        for pattern in quote_patterns:
                                            if re.search(pattern, value_str):
                                                # Удаляем кавычки с содержимым
                                                beer_name_cleaned = re.sub(pattern, '', beer_name_cleaned)
                                                brewery_in_quotes = True
                                                break
                                        
                                        if not brewery_in_quotes:
                                            # Удаляем brewery из начала названия
                                            beer_name_cleaned = value_str.replace(brewery_from_name, '', 1).strip()
                                            # Удаляем скобки с дополнительной информацией типа "(Fresh Brewed)"
                                            beer_name_cleaned = re.sub(r'\s*\([^)]*\)\s*', ' ', beer_name_cleaned).strip()
                                        
                                        # Очищаем от лишних пробелов и запятых
                                        beer_name_cleaned = re.sub(r'\s*,\s*', ' ', beer_name_cleaned).strip()
                                        beer_name_cleaned = re.sub(r'\s+', ' ', beer_name_cleaned).strip()
                                        
                                        item['beer_name'] = beer_name_cleaned if beer_name_cleaned else value_str
                                    else:
                                        item[field] = value_str
                                else:
                                    item[field] = value_str
                        elif field == 'brewery':
                            # КРИТИЧЕСКИ ВАЖНО: сохраняем исходное значение brewery ДО нормализации
                            # Это нужно для проверки на город в фильтрации
                            if 'brewery_original' not in item:
                                item['brewery_original'] = value_str
                            # Нормализуем brewery (удаляем город) сразу при извлечении из колонки
                            # Это гарантирует, что города будут удалены для всех листов, включая "Фасовка"
                            from parser_app.normalizers import DataNormalizer
                            normalizer = DataNormalizer()
                            brewery_normalized = normalizer.normalize_brewery(value_str)
                            item[field] = brewery_normalized if brewery_normalized else value_str
                        else:
                            item[field] = value_str
            except (IndexError, KeyError, ValueError, TypeError) as e:
                logger.debug(f"Ошибка при извлечении поля {field} (индекс {idx}): {str(e)}")
                continue
        
        # Пост-обработка: исправляем неправильный маппинг
        # Сохраняем volume, если он был извлечен из format_type
        volume_val = item.get('volume')
        
        # Если в beer_name или других полях есть "brewery", перемещаем в brewery
        beer_name_val_raw = item.get('beer_name', '')
        brewery_val_raw = item.get('brewery', '')
        ibu_val = item.get('ibu', '')
        style_val_raw = item.get('style', '')
        price_val_raw = item.get('price', '')
        
        # Проверяем, не был ли volume случайно перезаписан текстовым значением
        # Если volume строка и не является числом, удаляем его
        if 'volume' in item and isinstance(item['volume'], str):
            try:
                # Пробуем преобразовать в число
                float(item['volume'])
            except (ValueError, TypeError):
                # Если не число, удаляем (был перезаписан текстовым значением)
                del item['volume']
                # Восстанавливаем из сохраненного значения
                if volume_val and isinstance(volume_val, (int, float)):
                    item['volume'] = float(volume_val)
        
        # Проверяем все текстовые поля на наличие слова "brewery"
        for field_name, field_value in [('beer_name', beer_name_val_raw), ('ibu', ibu_val), ('style', style_val_raw)]:
            if field_value and isinstance(field_value, str):
                field_lower = field_value.lower()
                # Если в значении есть "brewery" или "пивоварня", это пивоварня
                if 'brewery' in field_lower or 'пивоварня' in field_lower:
                    # Если brewery пустое, перемещаем значение туда
                    if not brewery_val_raw:
                        # Нормализуем brewery (удаляем город) перед установкой
                        from parser_app.normalizers import DataNormalizer
                        normalizer = DataNormalizer()
                        brewery_normalized = normalizer.normalize_brewery(field_value)
                        item['brewery'] = brewery_normalized if brewery_normalized else field_value
                        item[field_name] = ''  # Очищаем исходное поле
                    elif brewery_val_raw != field_value:
                        # Если brewery уже заполнено и отличается, оставляем brewery, очищаем это поле
                        item[field_name] = ''
        
        # Если brewery все еще пустое, но есть beer_name с "brewery", извлекаем brewery из beer_name
        # НО не очищаем beer_name полностью - возможно там есть и название пива
        if not brewery_val_raw and beer_name_val_raw:
            beer_name_lower = str(beer_name_val_raw).lower()
            # Проверяем, содержит ли beer_name только название пивоварни или также название пива
            # Если beer_name содержит только одно слово и это "brewery" или "пивоварня" - это только пивоварня
            # Если beer_name содержит несколько слов - возможно это "Paradox IPA" или подобное
            words = beer_name_val_raw.split()
            if len(words) == 1 and ('brewery' in beer_name_lower or 'пивоварня' in beer_name_lower):
                # Только пивоварня - перемещаем в brewery
                from parser_app.normalizers import DataNormalizer
                normalizer = DataNormalizer()
                brewery_normalized = normalizer.normalize_brewery(beer_name_val_raw)
                item['brewery'] = brewery_normalized if brewery_normalized else beer_name_val_raw
                item['beer_name'] = ''
            elif len(words) > 1:
                # Несколько слов - возможно это "Paradox IPA" или "Paradox Stout"
                # Пытаемся извлечь пивоварню из начала, но оставляем остальное как название пива
                from parser_app.normalizers import DataNormalizer
                normalizer = DataNormalizer()
                # Пробуем извлечь пивоварню из первого слова
                first_word = words[0]
                brewery_normalized = normalizer.normalize_brewery(first_word)
                if brewery_normalized and brewery_normalized != first_word:
                    # Если нормализация изменила значение - это была пивоварня с городом
                    item['brewery'] = brewery_normalized
                    # Оставляем остальные слова как название пива
                    remaining_name = ' '.join(words[1:]).strip()
                    if remaining_name:
                        item['beer_name'] = remaining_name
                    else:
                        # Если после удаления пивоварни ничего не осталось - оставляем исходное значение
                        item['beer_name'] = beer_name_val_raw
                else:
                    # Не удалось извлечь пивоварню - оставляем как есть
                    pass
        
        # Для частного поставщика (brewery): подставляем пивоварню в пустые строки (например строка кега без названия)
        if supplier_type_enum == SupplierType.BREWERY and default_brewery and not item.get('brewery'):
            from parser_app.normalizers import DataNormalizer
            normalizer = DataNormalizer()
            item['brewery'] = normalizer.normalize_brewery(default_brewery) or default_brewery
        
        # Фильтрация: отбрасываем строки-заголовки без данных о пиве
        # Если есть только brewery, но нет beer_name, price, style - это заголовок секции
        # ВАЖНО: Используем ИСХОДНОЕ значение brewery ДО нормализации для проверки на город
        brewery_val_original = item.get('brewery_original', '').strip() if item.get('brewery_original') else ''
        # Если brewery_original не было сохранено, используем текущее значение brewery
        if not brewery_val_original:
            brewery_val_original = item.get('brewery', '').strip() if item.get('brewery') else ''
        brewery_val_check = item.get('brewery', '').strip() if item.get('brewery') else ''
        beer_name_val_check = item.get('beer_name', '').strip() if item.get('beer_name') else ''
        price_val_check = item.get('price', '').strip() if item.get('price') else ''
        style_val_check = item.get('style', '').strip() if item.get('style') else ''
        description_val_check = item.get('description', '').strip() if item.get('description') else ''
        format_type_val_check = item.get('format_type', '').strip() if item.get('format_type') else ''
        
        if brewery_val_check and not beer_name_val_check and not price_val_check and not style_val_check:
            # Проверяем, не содержит ли brewery адрес (г., г., город, city)
            brewery_lower = str(brewery_val_check).lower()
            address_indicators = [
                'г.', 'г ', 'город', 'city', 'ул.', 'улица', 'street', 
                'владимир', 'москва', 'санкт-петербург', 'spb', 'мск',
                'проспект', 'пр.', 'переулок', 'пер.', 'площадь', 'пл.'
            ]
            has_address = any(indicator in brewery_lower for indicator in address_indicators)
            
            # Проверяем, не является ли brewery служебным текстом
            service_keywords = ['прайс', 'price list', 'каталог', 'менеджер', 
                              'контакт', 'телефон', 'email', 'адрес']
            has_service_keyword = any(keyword in brewery_lower for keyword in service_keywords)
            
            if has_address or has_service_keyword or len(brewery_val_check.strip()) < 3:
                # Это заголовок секции или некорректная строка, пропускаем
                logger.debug(f"Пропущена строка с brewery '{brewery_val_check}' - похоже на заголовок или содержит адрес")
                return None
        
        # Определяем функцию проверки пустых значений в начале метода
        def is_empty_value(val):
            """Проверяет, является ли значение пустым или только тире"""
            if val is None:
                return True
            if not val:
                return True
            val_str = str(val).strip()
            if not val_str:
                return True
            val_str_lower = val_str.lower()
            # Проверяем различные варианты тире и пустых значений
            return val_str_lower in ['-', '—', '–', '', 'nan', 'none', 'null', 'n/a', 'na']
        
        # КРИТИЧЕСКАЯ ПРОВЕРКА: если brewery содержит город - всегда нормализуем и проверяем
        # Это нужно для фильтрации строк типа "4 Brewers г. Владимир" или "B-Side г. Санкт-Петербург" даже если есть цена
        # Используем ИСХОДНОЕ значение brewery (до нормализации в _extract_row_data)
        if brewery_val_original:
            brewery_lower = str(brewery_val_original).lower()
            # Расширенный список паттернов для городов (включая все варианты написания)
            city_patterns = [
                'г.', 'г ', 'город', 'city',
                'владимир', 'москва', 'санкт-петербург', 'spb', 'мск',
                'санкт-петербург', 'петербург', 'питер', 'spb',
                'санкт петербург', 'санктпетербург'
            ]
            has_city_in_brewery = any(pattern in brewery_lower for pattern in city_patterns)
            
            if has_city_in_brewery:
                # Нормализуем brewery (удаляем город) - если еще не было нормализовано
                from parser_app.normalizers import DataNormalizer
                normalizer = DataNormalizer()
                brewery_normalized = normalizer.normalize_brewery(brewery_val_original)
                
                # Если после нормализации brewery все еще содержит город - пропускаем строку
                if brewery_normalized:
                    brewery_normalized_lower = brewery_normalized.lower()
                    still_has_city = any(pattern in brewery_normalized_lower for pattern in city_patterns)
                    if still_has_city:
                        logger.debug(f"Пропущена строка с brewery '{brewery_val_original}' - после нормализации все еще содержит город")
                        return None
                    # Обновляем brewery в item
                    item['brewery'] = brewery_normalized
                    brewery_val_check = brewery_normalized
                else:
                    # Если нормализация вернула пустое значение - пропускаем строку
                    logger.debug(f"Пропущена строка с brewery '{brewery_val_original}' - после нормализации стал пустым")
                    return None
                
                # КРИТИЧЕСКАЯ ПРОВЕРКА: если brewery содержит город И при этом пустые/пропущены:
                # - beer_name ИЛИ
                # - style ИЛИ  
                # - description ИЛИ
                # - format_type
                # То такая строка считается некорректной и пропускается
                is_empty_beer_name = is_empty_value(beer_name_val_check)
                is_empty_style = is_empty_value(style_val_check)
                is_empty_description = is_empty_value(description_val_check)
                is_empty_format = is_empty_value(format_type_val_check)
                
                # Если все эти поля пустые - точно пропускаем
                if is_empty_beer_name and is_empty_style and is_empty_description and is_empty_format:
                    logger.debug(f"Пропущена строка с brewery '{brewery_val_original}' - содержит город и все поля (beer_name, style, description, format_type) пустые")
                    return None
                
                # Если хотя бы beer_name пустой - тоже пропускаем (даже если есть другие поля)
                if is_empty_beer_name:
                    logger.debug(f"Пропущена строка с brewery '{brewery_val_original}' - содержит город и нет beer_name")
                return None
        
        # Строгая проверка валидности записи
        # Запись считается валидной, если:
        # 1. Есть beer_name И (price ИЛИ brewery ИЛИ style)
        # 2. ИЛИ есть brewery И price (даже без beer_name)
        # 3. НО НЕ только brewery без других данных
        
        beer_name_val = beer_name_val_check
        brewery_val = brewery_val_check
        price_val = price_val_check
        style_val = style_val_check
        
        # Подсчитываем количество заполненных полей
        filled_fields = []
        if beer_name_val and not is_empty_value(beer_name_val):
            filled_fields.append('beer_name')
        if brewery_val and not is_empty_value(brewery_val):
            # Нормализуем brewery (удаляем город) перед добавлением в filled_fields
            # Это критически важно для удаления городов из названий пивоварен
            from parser_app.normalizers import DataNormalizer
            normalizer = DataNormalizer()
            brewery_normalized = normalizer.normalize_brewery(brewery_val)
            # Обновляем значение в item только если нормализация дала результат
            if brewery_normalized:
                item['brewery'] = brewery_normalized
                filled_fields.append('brewery')
            else:
                # Если после нормализации brewery стал пустым, все равно добавляем исходное значение
                # но логируем это для отладки
                logger.debug(f"Brewery стал пустым после нормализации: '{brewery_val}'")
            filled_fields.append('brewery')
        if price_val and not is_empty_value(price_val):
            filled_fields.append('price')
        if style_val and not is_empty_value(style_val):
            filled_fields.append('style')
        
        # Если нет заполненных полей - пропускаем
        if not filled_fields:
            logger.debug("Пропущена строка без заполненных полей")
            return None
        
        # Если только brewery без других данных - это заголовок секции
        if len(filled_fields) == 1 and 'brewery' in filled_fields:
            logger.debug(f"Пропущена строка с только brewery '{brewery_val}' без других данных")
            return None
        
        # Если есть beer_name, но нет price и нет brewery и нет style - возможно некорректная запись
        if beer_name_val and not is_empty_value(beer_name_val):
            if not price_val and not brewery_val and not style_val:
                # Проверяем, не является ли beer_name служебным текстом
                beer_name_lower = beer_name_val.lower()
                service_keywords = ['прайс', 'price list', 'каталог', 'менеджер', 
                                  'контакт', 'телефон', 'email', 'адрес', 'наименование',
                                  'название', 'товар', 'продукт']
                if any(keyword in beer_name_lower for keyword in service_keywords):
                    logger.debug(f"Пропущена строка с служебным текстом в beer_name: '{beer_name_val}'")
                    return None
        
        # Если нет beer_name, но есть brewery и price - это валидная запись
        # Если есть beer_name и хотя бы одно из (price, brewery, style) - это валидная запись
        has_valid_combination = (
            (beer_name_val and not is_empty_value(beer_name_val) and 
             (price_val or brewery_val or style_val)) or
            (brewery_val and not is_empty_value(brewery_val) and price_val)
        )
        
        if not has_valid_combination:
            logger.debug(f"Пропущена строка без валидной комбинации полей: beer_name={bool(beer_name_val)}, brewery={bool(brewery_val)}, price={bool(price_val)}")
            return None
        
        # Удаляем временное поле brewery_original перед возвратом
        if 'brewery_original' in item:
            del item['brewery_original']
        
        # #region agent log
        try:
            import os as _os
            _lp = _os.path.normpath(_os.path.join(_os.path.dirname(_os.path.abspath(__file__)), '..', '..', '..', '.cursor', 'debug-6958ea.log'))
            _bn_final = item.get('beer_name', '')[:100] if item.get('beer_name') else ''
            _desc_final = item.get('description', '')[:100] if item.get('description') else ''
            _data = {"hypothesisId": "H3", "message": "final values after processing", "data": {"beer_name_final": _bn_final, "description_final": _desc_final, "beer_name_len": len(item.get('beer_name', '')) if item.get('beer_name') else 0, "description_len": len(item.get('description', '')) if item.get('description') else 0}, "timestamp": __import__('time').time() * 1000}
            with open(_lp, "a", encoding="utf-8") as _f:
                _f.write(json.dumps(_data, ensure_ascii=False) + "\n")
        except Exception:
            pass
        # #endregion agent log
        
        return item
    
    def _looks_like_style_only_column(self, col_data: List[str], max_sample: int = 10) -> bool:
        """
        Возвращает True, если значения в колонке выглядят как стили пива (IPA, Ale, Altbier...),
        а не как названия продуктов. Такие колонки не должны маппиться на beer_name.
        """
        if not col_data or len(col_data) < 2:
            return False
        sample = [str(v).strip() for v in col_data[:max_sample] if str(v).strip()]
        if len(sample) < 2:
            return False
        # Типичные окончания/маркеры стилей (не уникальные для названий)
        style_endings = (
            r'\bipa\s*$', r'\bale\s*$', r'\blager\s*$', r'\bstout\s*$', r'\bporter\s*$',
            r'\bpilsner\s*$', r'\baltbier\s*$', r'\bsour\s*$', r'\bgose\s*$', r'\bweiss(e)?\s*$',
            r'\bwheat\s*$', r'\bne\s+ipa\s*$', r'\bcold\s+ipa\s*$', r'\bimperial\s+', r'\bdouble\s+ipa',
            r'\bpale\s+ale', r'\bamber\s+ale', r'\bblonde\s+ale', r'\bred\s+ale', r'\bbrown\s+ale',
            r'\bblack\s+ipa', r'\bwhite\s+ipa', r'\bnew\s+england\s+ipa', r'\bberliner\s+weisse',
            r'\bвитбир\s*$', r'\bлагер\s*$', r'\bэль\s*$', r'\bстаут\s*$', r'\bпортер\s*$',
        )
        pattern = re.compile('|'.join(style_endings), re.IGNORECASE)
        short = [s for s in sample if len(s) <= 55]
        if len(short) < len(sample) * 0.5:
            return False
        # Названия часто содержат двоеточие ("Бренд: Название"), стили — редко
        with_colon = sum(1 for s in sample if ':' in s)
        if with_colon >= len(sample) * 0.4:
            return False
        matches = sum(1 for s in sample if pattern.search(s))
        return matches >= len(sample) * 0.5

    def _disambiguate_name_vs_description(
        self, df: pd.DataFrame, col_mapping: Dict[str, int], sheet_name: str
    ) -> Dict[str, int]:
        """
        Если колонка, маппированная на beer_name, содержит длинный описательный текст,
        переназначаем её на description и ищем для beer_name колонку с коротким названием.
        Исправляет типичную ошибку парсинга дистрибьюторов: описание попадает в «Название».
        """
        if not df.empty and 'beer_name' in col_mapping:
            bn_idx = col_mapping['beer_name']
            headers = [str(c).lower() for c in df.columns.tolist()] if hasattr(df, 'columns') else []
            # Не трогаем маппинг, если колонка названия явно подписана "Название"/"Наименование"
            if bn_idx < len(headers):
                h = headers[bn_idx]
                if ('наименование' in h and 'пивоварни' not in h) or h.strip() == 'название':
                    return col_mapping
            sample = df.head(min(15, len(df)))
            try:
                col_data = sample.iloc[:, bn_idx].dropna().astype(str).tolist()
                if not col_data:
                    return col_mapping
                avg_len = sum(len(v) for v in col_data[:10]) / min(10, len(col_data))
                text_lower = ' '.join(col_data[:5]).lower()
                desc_keywords = [
                    'вкус', 'аромат', 'насыщенный', 'сочный', 'обусловлен', 'использованием',
                    'taste', 'aroma', 'flavor', 'due to', 'combination with'
                ]
                looks_like_description = (
                    avg_len > 100
                    or (avg_len > 60 and any(kw in text_lower for kw in desc_keywords))
                )
                if not looks_like_description:
                    return col_mapping
                # Текущая «beer_name» — на самом деле описание
                used_indices = set(col_mapping.values())
                desc_idx = col_mapping.get('description')
                # Ищем колонку с коротким текстом (название товара); не подставляем "Описание" и "Стиль"
                for col_idx in range(min(12, len(df.columns))):
                    if col_idx in used_indices or col_idx == bn_idx:
                        continue
                    if col_idx < len(headers):
                        oh = headers[col_idx]
                        if ('описание' in oh or oh.strip() == 'description' or
                                'стиль' in oh or oh.strip() == 'style'):
                            continue
                    try:
                        other_data = sample.iloc[:, col_idx].dropna().astype(str).tolist()
                        if len(other_data) < 2:
                            continue
                        other_avg = sum(len(v) for v in other_data[:10]) / min(10, len(other_data))
                        if 3 <= other_avg <= 100:
                            col_mapping = dict(col_mapping)
                            col_mapping['description'] = bn_idx
                            col_mapping['beer_name'] = col_idx
                            logger.info(
                                f"Лист {sheet_name}: колонка названия и описания разведены: "
                                f"beer_name={col_idx} (avg_len≈{other_avg:.0f}), description={bn_idx} (avg_len≈{avg_len:.0f})"
                            )
                            return col_mapping
                    except Exception:
                        continue
                # Не подставляем колонку «Стиль» в название — там стиль пива (IPA, Gose), а не название товара
                # Подходящей короткой колонки не нашли — не меняем маппинг (не подставляем стиль в название)
                col_mapping = dict(col_mapping)
                col_mapping['description'] = bn_idx
                logger.info(
                    f"Лист {sheet_name}: колонка beer_name (idx={bn_idx}) помечена как описание (avg_len≈{avg_len:.0f}), "
                    f"отдельной колонки для названия не найдено"
                )
            except Exception as e:
                logger.debug(f"Ошибка при разведении name/description: {e}")
        return col_mapping
    
    def _guess_column_mapping(self, df: pd.DataFrame) -> Dict[str, int]:
        """
        Пытается угадать маппинг колонок по содержимому первых строк.
        
        Args:
            df: DataFrame с данными
            
        Returns:
            Словарь {название_поля: индекс_колонки}
        """
        mapping = {}
        
        if df.empty or len(df.columns) == 0:
            return mapping
        
        # Анализируем первые несколько строк для определения типов данных
        sample_rows = df.head(min(10, len(df)))
        
        # Ищем колонки с текстом
        text_columns = []
        for col_idx in range(min(10, len(df.columns))):
            col_name = df.columns[col_idx]
            sample_values = sample_rows[col_name].dropna().astype(str).tolist()
            
            # Проверяем паттерны в значениях
            text_values = [v for v in sample_values 
                          if len(str(v)) > 2 and 
                          not str(v).replace('.', '').replace('-', '').replace(' ', '').isdigit()]
            
            if len(text_values) > len(sample_values) * 0.5:  # Если больше половины - текст
                text_columns.append((col_idx, text_values))
        
        # Пивоварня / название / описание: длинный текст — описание, короткий — название
        for col_idx, text_values in text_columns:
            has_brewery_words = any('brewery' in str(v).lower() or 'пивоварня' in str(v).lower() 
                                   for v in text_values[:5])
            avg_len = sum(len(str(v)) for v in text_values[:5]) / min(5, len(text_values)) if text_values else 0
            
            if has_brewery_words and 'brewery' not in mapping:
                mapping['brewery'] = col_idx
            elif avg_len > 120 and 'description' not in mapping:
                mapping['description'] = col_idx
            elif avg_len <= 120 and 'beer_name' not in mapping:
                mapping['beer_name'] = col_idx
        
        # Ищем колонку с ценами (содержит числа)
        for col_idx in range(min(10, len(df.columns))):
            col_name = df.columns[col_idx]
            sample_values = sample_rows[col_name].dropna()
            
            if len(sample_values) > 0:
                numeric_values = [v for v in sample_values if isinstance(v, (int, float)) and v > 0]
                if len(numeric_values) > len(sample_values) * 0.5:  # Если больше половины - числа
                    if 'price' not in mapping and max(numeric_values) < 10000:  # Цены обычно меньше 10000
                        mapping['price'] = col_idx
                    elif 'abv' not in mapping and max(numeric_values) <= 100:  # Крепость обычно до 100
                        mapping['abv'] = col_idx
        
        return mapping
    
    def _fallback_generic_mapping(self, df: pd.DataFrame) -> Dict[str, int]:
        """
        Универсальный маппинг по типу данных: первая длинная текстовая колонка = beer_name,
        первая числовая в диапазоне цен (1–100000) = price. Позволяет распарсить любой прайс хотя бы по названию и цене.
        """
        mapping = {}
        if df.empty or len(df.columns) < 2:
            return mapping
        sample = df.head(min(20, len(df)))
        for col_idx in range(len(df.columns)):
            try:
                col_vals = sample.iloc[:, col_idx].dropna().astype(str).tolist()
                non_empty = [v for v in col_vals if str(v).strip() and str(v).lower() not in ('nan', 'none', '')]
                if len(non_empty) < 2:
                    continue
                avg_len = sum(len(str(v)) for v in non_empty[:10]) / min(10, len(non_empty))
                nums = []
                for v in non_empty[:15]:
                    try:
                        n = normalize_number_str(v)
                        if n:
                            f = float(n)
                            if 0 < f < 1e7:
                                nums.append(f)
                    except (ValueError, TypeError):
                        pass
                is_likely_numeric = len(nums) >= len(non_empty[:15]) * 0.4
                if is_likely_numeric and 1 <= max(nums) <= 100000 and (sum(nums) / len(nums) > 30 or max(nums) > 100):
                    if 'price' not in mapping:
                        mapping['price'] = col_idx
                elif avg_len >= 5 and not is_likely_numeric:
                    # Длинный текст (≥120 символов) — скорее описание, не название товара
                    if avg_len >= 120:
                        if 'description' not in mapping:
                            mapping['description'] = col_idx
                    elif 'beer_name' not in mapping:
                        mapping['beer_name'] = col_idx
            except Exception:
                continue
        return mapping
    
    def _analyze_data_content(self, df: pd.DataFrame) -> Dict[str, int]:
        """
        Анализирует содержимое данных для определения колонок.
        Более агрессивный подход: анализирует первые несколько строк.
        
        Args:
            df: DataFrame с данными
            
        Returns:
            Словарь {название_поля: индекс_колонки}
        """
        mapping = {}
        
        if df.empty or len(df.columns) == 0:
            return mapping
        
        # Берем первые 5 строк для анализа
        sample = df.head(5)
        
        # Проходим по каждой колонке и анализируем содержимое
        for col_idx in range(min(10, len(df.columns))):
            col_data = sample.iloc[:, col_idx].dropna().astype(str).tolist()
            
            if not col_data:
                continue
            
            # Проверяем наличие слова "brewery" в данных
            has_brewery = any('brewery' in str(v).lower() or 'пивоварня' in str(v).lower() 
                            for v in col_data)
            
            # Проверяем на числа (цены)
            numeric_count = sum(1 for v in col_data 
                              if isinstance(df.iloc[0, col_idx], (int, float)) 
                              or (isinstance(v, str) and v.replace('.', '').replace(',', '').isdigit()))
            
            # Проверяем длину текста: название — короткое/среднее, описание — длинное
            avg_length = sum(len(str(v)) for v in col_data[:3]) / min(3, len(col_data))
            
            if has_brewery and 'brewery' not in mapping:
                mapping['brewery'] = col_idx
            elif numeric_count > len(col_data) * 0.6 and 'price' not in mapping:
                try:
                    first_num = float(str(col_data[0]).replace(',', '.'))
                    if 10 < first_num < 10000:
                        mapping['price'] = col_idx
                except:
                    pass
            elif avg_length > 120 and 'description' not in mapping and not has_brewery:
                mapping['description'] = col_idx
            elif 10 < avg_length <= 120 and 'beer_name' not in mapping and not has_brewery:
                mapping['beer_name'] = col_idx
            elif 5 < avg_length <= 120 and 'brewery' not in mapping and 'beer_name' not in mapping:
                if 'brewery' not in mapping:
                    mapping['beer_name'] = col_idx
        
        return mapping
    
    def extract_tables(self):
        """
        Извлечение всех таблиц из Excel.
        
        Returns:
            Список таблиц (DataFrames преобразуются в списки)
        """
        tables = []
        try:
            excel_file = pd.ExcelFile(self.file_path)
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                tables.append(df.values.tolist())
        except Exception:
            pass
        return tables
    
    def extract_text(self):
        """
        Извлечение всего текста из Excel.
        
        Returns:
            Текст файла (все значения соединены)
        """
        text = ""
        try:
            excel_file = pd.ExcelFile(self.file_path)
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                text += f"Sheet: {sheet_name}\n"
                text += df.to_string() + "\n\n"
        except Exception:
            pass
        return text

